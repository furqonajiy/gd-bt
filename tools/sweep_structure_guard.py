#!/usr/bin/env python3
"""Targeted comparison: base TSL18 feed vs structure-guarded variants.

This is deliberately NOT a generic profit sweep. TSL18 already went through the
RSI x Bollinger x R:R feed sweep and the trailing/risk sweeps. The open problem
is narrower: TSL18 takes **sequential losses when it trades against the larger
structure** -- BUY signals while H1 is bearish, SELL signals while H1 is bullish.
So this tool scores the structure guard on the metrics that describe THAT failure
mode, not headline profit:

    net P&L, max drawdown, win rate, profit factor, total trades, loss count,
    max consecutive losses, max daily loss,
    BUY losses during bearish HTF, SELL losses during bullish HTF,
    filtered winners vs filtered losers (what the guard removed).

For each variant it (1) regenerates the self-scalper feed with the variant's
``--structure-*`` flags (the base variant uses none), (2) runs the SAME TSL18
execution geometry through ``backtest_hybrid`` on TICK where available, and
(3) joins each closed trade to the structure-diagnostics so a loss can be tagged
as wrong-side-HTF. The base feed's removed signals are looked up in the base
backtest to report how many filtered signals were winners vs losers.

Run June 2026 first (pure TICK), then Jan->Jun 2026:

    python tools/sweep_structure_guard.py --window june
    python tools/sweep_structure_guard.py --window jan_jun

Output: reports/STRUCTURE_GUARD_<window>/summary.md (+ per-variant workbooks).
The expected result is NOT "the guard is more profitable"; it is a clear,
auditable answer to "does the guard cut wrong-side sequential losses, and at
what cost in filtered winners".
"""
from __future__ import annotations

import argparse
import csv
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]

# C160 self-scalper feed filter -- identical to TSL18 / T818 (do not change here).
FEED_FILTER = [
    "--session-start", "0", "--session-end", "0", "--signal-tz", "7",
    "--rsi-buy-max", "70", "--rsi-sell-min", "30", "--bb-bandwidth-min", "0.0006",
    "--rr1", "1.2", "--rr2", "2.5", "--rr3", "5",
]

# TSL18 execution geometry -- identical to cli/candidate_TSL18_trailing_tick.txt.
TSL18_GEOMETRY = [
    "--sizing-mode", "risk", "--lot", "0.01", "--risk", "0.01",
    "--minimum-lot", "0.01", "--maximum-lot", "500.0", "--lot-step", "0.01",
    "--bonus-per-closed-lot", "3.0", "--entries", "8", "--entry-ladder", "range_to_sl",
    "--entry-sl-gap", "0.7", "--shared-sl", "false", "--activation-delay", "0",
    "--pending-expiry", "180", "--max-hold", "150", "--sl-multiplier", "1.8",
    "--final-target", "TP3", "--lock-after-tp1", "true", "--lock-after-tp2", "true",
    "--tp1-lock-delay-minutes", "24", "--tp2-lock-delay-minutes", "24",
    "--profit-lock-mode", "tp_levels", "--bep-trigger-distance", "3.0",
    "--tp1-lock-fraction", "0.75", "--tp2-lock-target", "TP1", "--tp3-lock-target", "TP2",
    "--runner-after-tp3", "false", "--trailing-open-distance", "0.5",
    "--trailing-close-distance", "0.5", "--trailing-close-after-stage", "2",
]

# Structure variants to compare. 'base' = no guard (the incumbent TSL18 feed).
# These are starting points keyed to the wrong-side problem, NOT a wide grid.
# Structure-guard flag blocks (wrong-side veto) and the progress-stall block
# (same-side cluster cap). Combined variants stack both. Initial stall defaults
# are the spec values; this is NOT a broad parameter grid.
_HTF = ["--structure-filter", "--structure-htf-minutes", "60",
        "--structure-ema-fast", "20", "--structure-ema-slow", "50"]
_IMPULSE = ["--structure-impulse-cooldown-bars", "5", "--structure-impulse-atr", "1.5"]
_VWAP_SCORE = ["--structure-require-vwap-side", "--structure-min-score", "2"]
PROGRESS = ["--progress-stall-filter", "--progress-htf-minutes", "60",
            "--progress-ema-fast", "20", "--progress-ema-slow", "50",
            "--progress-min-diff-atr", "0.10", "--progress-local-lookback-bars", "30",
            "--progress-stall-n", "3", "--progress-min-no-progress-bars", "20",
            "--progress-probe-interval-bars", "30", "--progress-min-atr", "0.50",
            "--progress-close-confirm-atr", "0.10", "--progress-min-points", "1.0"]

VARIANTS: dict[str, list[str]] = {
    "base": [],
    "structure_htf_only": _HTF,
    "structure_htf_impulse": _HTF + _IMPULSE,
    "structure_htf_vwap_score2": _HTF + _IMPULSE + _VWAP_SCORE,
    "progress_stall_only": PROGRESS,
    "structure_htf_only_plus_progress_stall": _HTF + PROGRESS,
    "structure_htf_impulse_plus_progress_stall": _HTF + _IMPULSE + PROGRESS,
    "structure_htf_vwap_score2_plus_progress_stall": _HTF + _IMPULSE + _VWAP_SCORE + PROGRESS,
}

# NB: backtest_hybrid treats --end-date as EXCLUSIVE, so to include all of June 30
# the end is 2026-07-01 (the day after the last day we want to keep).
WINDOWS = {
    "june": ("2026-06-01", "2026-07-01"),
    "jan_jun": ("2026-01-01", "2026-07-01"),
}

ERA_SLIP = ["--lock-tp1-exit-slippage", "2.0", "--lock-tp2-exit-slippage", "1.0"]  # R4


@dataclass
class Trade:
    date: str
    time_chart: str
    side: str
    pnl: float
    status: str
    signal_key: str = ""   # 'YYYY-MM-DD#NN' (the leg's signal, NN per-day index)

    @property
    def match_key(self) -> tuple[str, str]:
        """Cross-feed signal identity = (chart timestamp, side).

        We deliberately do NOT key on the Entry Key signal portion (``date#NN``):
        the per-day index ``NN`` RENUMBERS when the guard filters signals out, so
        ``date#NN`` is not stable between the base and a guarded feed. The chart
        bar timestamp + side IS stable (the same chart bar produces the same
        signal in every feed) and is zone-consistent, so it matches correctly even
        for signals whose feed-zone (source) date differs from their chart date
        around the GMT+7/EET midnight boundary.
        """
        return (self.time_chart, self.side)


@dataclass
class Metrics:
    variant: str
    trades: int = 0
    net: float = 0.0
    win_rate: float = 0.0
    profit_factor: float = 0.0
    losses: int = 0
    # entry-level streak (TSL18 opens up to 8 entries/signal, so this over-counts
    # the user-facing 'sequential losing TRADES' problem -- keep it but name it
    # clearly so it is not confused with the signal-level streaks below).
    max_consecutive_losing_entries: int = 0
    # signal-level streaks (entries grouped by signal; a losing signal = total
    # signal P&L < 0). These are what the operator actually experiences.
    signals: int = 0
    losing_signals: int = 0
    max_consecutive_losing_signals: int = 0
    max_consecutive_wrong_side_losing_signals: int = 0
    max_daily_loss: float = 0.0
    max_drawdown: float = 0.0
    buy_loss_bear_htf: int = 0
    sell_loss_bull_htf: int = 0
    filtered_total: int = 0
    filtered_winners: int = 0
    filtered_losers: int = 0
    # progress-stall-specific (only populated for variants with --progress-stall-filter)
    progress_stall_filtered_total: int = 0
    progress_stall_filtered_winners: int = 0
    progress_stall_filtered_losers: int = 0
    avg_bars_since_progress_when_filtered: float = 0.0
    avg_non_progressing_count_when_filtered: float = 0.0
    max_non_progressing_count_seen: int = 0
    probe_signals_allowed: int = 0


def _run(cmd: list[str]) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True)


def _gen_feed(out_txt: Path, struct_csv: Path | None, charts: list[str],
              start: str, structure_flags: list[str], prog_csv: Path | None = None) -> None:
    cmd = [sys.executable, "tools/generate_scalper_signals.py",
           "--charts", *charts, "--output", str(out_txt),
           "--start", start, "--progress-interval-seconds", "0",
           *FEED_FILTER, *structure_flags]
    if struct_csv is not None and "--structure-filter" in structure_flags:
        cmd += ["--structure-diagnostics", str(struct_csv)]
    if prog_csv is not None and "--progress-stall-filter" in structure_flags:
        cmd += ["--progress-stall-diagnostics", str(prog_csv)]
    r = _run(cmd)
    if r.returncode != 0:
        raise SystemExit(f"feed generation failed:\n{r.stderr[-1500:]}")


def _run_backtest(feed: Path, charts: list[str], ticks: list[str],
                  start: str, end: str, out_dir: Path) -> Path:
    cmd = [sys.executable, "tools/backtest_hybrid.py", "--signals", str(feed),
           "--charts", *charts, "--ticks", *ticks,
           "--sync-ticks", "false", "--sync-charts", "false",
           "--max-drawdown-limit-pct", "9999", "--progress-interval-seconds", "0",
           *TSL18_GEOMETRY, *ERA_SLIP,
           "--output-dir", str(out_dir), "--initial-capital", "50000",
           "--start-date", start, "--end-date", end]
    r = _run(cmd)
    if r.returncode != 0:
        raise SystemExit(f"backtest failed:\n{r.stderr[-1500:]}")
    xlsx = out_dir.with_suffix(".xlsx")
    if not xlsx.exists():
        raise SystemExit(f"no workbook at {xlsx}")
    return xlsx


def _read_trades(xlsx: Path) -> list[Trade]:
    from openpyxl import load_workbook

    ws = load_workbook(xlsx, data_only=True)["Per-Entry Detail"]
    hdr = {str(ws.cell(row=2, column=c).value).strip(): c for c in range(1, ws.max_column + 1)}
    out: list[Trade] = []
    for r in range(3, ws.max_row + 1):
        k = ws.cell(row=r, column=hdr["Entry Key"]).value
        pnl = ws.cell(row=r, column=hdr["P&L ($)"]).value
        if not k or pnl is None:
            continue
        key = str(k)
        out.append(Trade(
            date=str(ws.cell(row=r, column=hdr["Date"]).value),
            time_chart=str(ws.cell(row=r, column=hdr["Time (chart EET/EEST)"]).value),
            side=str(ws.cell(row=r, column=hdr["Side"]).value).upper(),
            pnl=float(pnl),
            status=str(ws.cell(row=r, column=hdr["Status"]).value),
            signal_key=key.rsplit(".", 1)[0] if "." in key else key,
        ))
    return out


def _read_htf(struct_csv: Path) -> dict[tuple[str, str], str]:
    """(chart-time 'YYYY-MM-DD HH:MM', side) -> htf_state, from the diagnostics."""
    out: dict[tuple[str, str], str] = {}
    if not struct_csv.exists():
        return out
    with struct_csv.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            t = row["time"][:16]  # 'YYYY-MM-DD HH:MM'
            out[(t, row["side"].upper())] = row["htf_state"]
    return out


def _metrics(variant: str, trades: list[Trade], htf: dict[tuple[str, str], str]) -> Metrics:
    m = Metrics(variant=variant, trades=len(trades))
    wins = [t.pnl for t in trades if t.pnl > 0]
    losses = [t.pnl for t in trades if t.pnl < 0]
    m.net = round(sum(t.pnl for t in trades), 2)
    m.losses = len(losses)
    m.win_rate = round(100.0 * len(wins) / len(trades), 1) if trades else 0.0
    gross_win, gross_loss = sum(wins), -sum(losses)
    m.profit_factor = round(gross_win / gross_loss, 2) if gross_loss > 0 else float("inf")
    # ENTRY-level max consecutive losses (trades are in fill order in the sheet).
    run = best = 0
    for t in trades:
        if t.pnl < 0:
            run += 1
            best = max(best, run)
        else:
            run = 0
    m.max_consecutive_losing_entries = best
    # max daily loss + a simple equity-curve drawdown (sequential)
    by_day: dict[str, float] = {}
    eq = peak = 0.0
    dd = 0.0
    for t in trades:
        by_day[t.date] = by_day.get(t.date, 0.0) + t.pnl
        eq += t.pnl
        peak = max(peak, eq)
        dd = min(dd, eq - peak)
    m.max_daily_loss = round(min(by_day.values()), 2) if by_day else 0.0
    m.max_drawdown = round(dd, 2)
    # entry-level wrong-side-HTF losses
    for t in trades:
        if t.pnl >= 0:
            continue
        state = htf.get((t.time_chart[:16], t.side))
        if t.side == "BUY" and state == "bear":
            m.buy_loss_bear_htf += 1
        elif t.side == "SELL" and state == "bull":
            m.sell_loss_bull_htf += 1

    # --- SIGNAL-level streaks (the user-facing sequential-loss view). Group the
    #     up-to-8 entries of a signal, aggregate their P&L; a signal LOSES when
    #     its total P&L < 0. Order signals by their (chart) entry time. A
    #     wrong-side losing signal = a losing BUY whose HTF was bearish, or a
    #     losing SELL whose HTF was bullish. ---
    agg: dict[str, dict] = {}
    for t in trades:
        sig = agg.setdefault(t.signal_key, {"pnl": 0.0, "side": t.side,
                                            "time": t.time_chart})
        sig["pnl"] += t.pnl
        sig["time"] = min(sig["time"], t.time_chart)  # earliest leg = signal time
    ordered = sorted(agg.values(), key=lambda s: s["time"])
    m.signals = len(ordered)
    m.losing_signals = sum(1 for s in ordered if s["pnl"] < 0)

    def _max_run(flags: list[bool]) -> int:
        run = best = 0
        for f in flags:
            run = run + 1 if f else 0
            best = max(best, run)
        return best

    losing_flags = [s["pnl"] < 0 for s in ordered]
    wrong_side_flags = []
    for s in ordered:
        if s["pnl"] >= 0:
            wrong_side_flags.append(False)
            continue
        state = htf.get((s["time"][:16], s["side"]))
        wrong_side_flags.append(
            (s["side"] == "BUY" and state == "bear")
            or (s["side"] == "SELL" and state == "bull"))
    m.max_consecutive_losing_signals = _max_run(losing_flags)
    m.max_consecutive_wrong_side_losing_signals = _max_run(wrong_side_flags)
    return m


def _filtered_breakdown(base_trades: list[Trade], variant_trades: list[Trade]) -> tuple[int, int, int]:
    """SIGNALS present in base but removed by the variant: winners vs losers.

    Matched at the SIGNAL level on ``Trade.match_key`` (chart timestamp + side),
    which is stable across feeds -- the Entry Key's per-day index renumbers when
    the guard drops signals, so it can't be used for cross-feed matching. A
    removed signal counts as a winner/loser by its aggregate P&L.
    """
    def by_signal(trades: list[Trade]) -> dict[tuple[str, str], float]:
        agg: dict[tuple[str, str], float] = {}
        for t in trades:
            agg[t.match_key] = agg.get(t.match_key, 0.0) + t.pnl
        return agg

    base_sig = by_signal(base_trades)
    kept = set(by_signal(variant_trades))
    removed = {k: pnl for k, pnl in base_sig.items() if k not in kept}
    winners = sum(1 for pnl in removed.values() if pnl > 0)
    losers = sum(1 for pnl in removed.values() if pnl < 0)
    return len(removed), winners, losers


def _progress_stall_metrics(m: Metrics, prog_csv: Path, base_trades: list[Trade]) -> None:
    """Fill the progress-stall-specific fields from the variant's diagnostics CSV.

    Winner/loser of a progress_stall-rejected signal is looked up in the BASE
    backtest by chart-time+side (the signal the cap removed and what it would
    have done if left in)."""
    if not prog_csv.exists():
        return
    base_sig: dict[tuple[str, str], float] = {}
    for t in base_trades:
        base_sig[t.match_key] = base_sig.get(t.match_key, 0.0) + t.pnl
    bars, counts = [], []
    max_seen = 0
    win = lose = total = probes = 0
    with prog_csv.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            try:
                max_seen = max(max_seen, int(row["non_progressing_count"]))
            except (ValueError, KeyError):
                pass
            if str(row.get("probe_allowed")) == "1":
                probes += 1
            if row.get("reject_reason") != "progress_stall":
                continue
            total += 1
            bars.append(float(row.get("bars_since_valid_progress") or 0))
            counts.append(float(row.get("non_progressing_count") or 0))
            # diag time has seconds; the backtest chart-time key is minute precision
            key = (row["time"][:16], row["side"].upper())
            pnl = base_sig.get(key)
            if pnl is not None:
                win += int(pnl > 0)
                lose += int(pnl < 0)
    m.progress_stall_filtered_total = total
    m.progress_stall_filtered_winners = win
    m.progress_stall_filtered_losers = lose
    m.avg_bars_since_progress_when_filtered = round(sum(bars) / len(bars), 1) if bars else 0.0
    m.avg_non_progressing_count_when_filtered = round(sum(counts) / len(counts), 2) if counts else 0.0
    m.max_non_progressing_count_seen = max_seen
    m.probe_signals_allowed = probes


def _write_summary(out_dir: Path, window: str, span: tuple[str, str],
                   results: list[Metrics]) -> Path:
    md = out_dir / "summary.md"
    md.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        f"# Structure-guard comparison — {window} ({span[0]}..{span[1]})",
        "",
        "Base TSL18 feed vs structure-guarded variants, SAME TSL18 geometry, TICK where available.",
        "Judge on the sequential / wrong-side columns, not net alone.",
        "",
        "| variant | signals | losing sig | maxConsec losing SIG | maxConsec wrong-side losing SIG | entries | maxConsec losing ENTRIES | net $ | win% | PF | maxDailyLoss $ | maxDD $ | BUYloss·bearHTF | SELLloss·bullHTF | filtered sig (W/L) |",
        "|---|--:|--:|--:|--:|--:|--:|--:|--:|--:|--:|--:|--:|--:|--:|",
    ]
    for m in results:
        pf = "inf" if m.profit_factor == float("inf") else f"{m.profit_factor:.2f}"
        lines.append(
            f"| {m.variant} | {m.signals} | {m.losing_signals} | "
            f"{m.max_consecutive_losing_signals} | {m.max_consecutive_wrong_side_losing_signals} | "
            f"{m.trades} | {m.max_consecutive_losing_entries} | {m.net:,.0f} | {m.win_rate} | {pf} | "
            f"{m.max_daily_loss:,.0f} | {m.max_drawdown:,.0f} | "
            f"{m.buy_loss_bear_htf} | {m.sell_loss_bull_htf} | "
            f"{m.filtered_winners}/{m.filtered_losers} (of {m.filtered_total}) |"
        )
    lines += [
        "",
        "**Read this on the SIGNAL columns first** (TSL18 opens up to 8 entries per "
        "signal, so the entry-level streak over-counts the felt pain). A good guard "
        "lowers *maxConsec losing SIG*, *maxConsec wrong-side losing SIG*, *maxDailyLoss* "
        "and *maxDD* while keeping *filtered losers >> filtered winners*. If it filters "
        "mostly winners, it is hurting — do not promote.",
    ]
    # progress-stall-specific table (only variants that ran the stall cap)
    prog = [m for m in results if m.progress_stall_filtered_total > 0
            or m.max_non_progressing_count_seen > 0]
    if prog:
        lines += [
            "",
            "## Progress-stall specifics (same-side cluster cap)",
            "",
            "| variant | ps filtered (W/L of total) | avg bars-since-progress | avg non-prog count | max non-prog seen | probes allowed |",
            "|---|--:|--:|--:|--:|--:|",
        ]
        for m in prog:
            lines.append(
                f"| {m.variant} | {m.progress_stall_filtered_winners}/"
                f"{m.progress_stall_filtered_losers} (of {m.progress_stall_filtered_total}) | "
                f"{m.avg_bars_since_progress_when_filtered} | "
                f"{m.avg_non_progressing_count_when_filtered} | "
                f"{m.max_non_progressing_count_seen} | {m.probe_signals_allowed} |"
            )
        lines += [
            "",
            "The progress-stall cap is promotable on this window only if it LOWERS "
            "*maxConsec losing SIG* vs base while *ps filtered losers > winners* and PF "
            "does not materially collapse.",
        ]
    md.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return md


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--window", choices=list(WINDOWS), default="june")
    ap.add_argument("--charts", nargs="+",
                    default=["data/XAUUSD_M1_202601_ELEV8.csv", "data/XAUUSD_M1_202602_ELEV8.csv",
                             "data/XAUUSD_M1_202603_ELEV8.csv", "data/XAUUSD_M1_202604_ELEV8.csv",
                             "data/XAUUSD_M1_202605_ELEV8.csv", "data/XAUUSD_M1_202606_ELEV8.csv"])
    ap.add_argument("--ticks", nargs="+",
                    default=["data/ticks/XAUUSD_TICK_*_ELEV8.csv"])
    ap.add_argument("--variants", nargs="+", default=list(VARIANTS),
                    help=f"subset of {list(VARIANTS)}")
    ap.add_argument("--gen-start", default="2026-01-01",
                    help="feed-generation start (indicator warmup precedes the window).")
    ap.add_argument("--out-root", default="reports")
    args = ap.parse_args(argv)

    start, end = WINDOWS[args.window]
    out_root = Path(args.out_root) / f"STRUCTURE_GUARD_{args.window}"
    out_root.mkdir(parents=True, exist_ok=True)

    base_trades: list[Trade] = []
    # One-time HTF label pass (htf_state per signal time is variant-independent):
    # a guard-ON generation whose diagnostics give each base-setup signal's
    # htf_state. Used to tag every variant's trades for the wrong-side metrics.
    print("[structure-sweep] generating HTF labels ...", flush=True)
    label_csv = out_root / "htf_labels.csv"
    _gen_feed(out_root / "_label.txt", label_csv, args.charts, args.gen_start, _HTF)
    htf = _read_htf(label_csv)

    results: list[Metrics] = []
    for v in args.variants:
        flags = VARIANTS[v]
        has_prog = "--progress-stall-filter" in flags
        feed = out_root / f"feed_{v}.txt"
        prog_csv = out_root / f"progress_{v}.csv"
        print(f"[structure-sweep] {v}: generate feed ...", flush=True)
        _gen_feed(feed, None, args.charts, args.gen_start, flags,
                  prog_csv=prog_csv if has_prog else None)
        print(f"[structure-sweep] {v}: backtest {start}..{end} ...", flush=True)
        xlsx = _run_backtest(feed, args.charts, args.ticks, start, end, out_root / f"bt_{v}")
        trades = _read_trades(xlsx)
        m = _metrics(v, trades, htf)
        if v == "base":
            base_trades = trades
        else:
            m.filtered_total, m.filtered_winners, m.filtered_losers = \
                _filtered_breakdown(base_trades, trades)
        if has_prog:
            _progress_stall_metrics(m, prog_csv, base_trades)
        results.append(m)
        print(f"[structure-sweep] {v}: signals={m.signals} losingSig={m.losing_signals} "
              f"maxConsecLosingSig={m.max_consecutive_losing_signals} "
              f"maxConsecWrongSideSig={m.max_consecutive_wrong_side_losing_signals} "
              f"entries={m.trades} net=${m.net:,.0f} psFiltL/W="
              f"{m.progress_stall_filtered_losers}/{m.progress_stall_filtered_winners}", flush=True)

    md = _write_summary(out_root, args.window, (start, end), results)
    print(f"[structure-sweep] summary: {md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
