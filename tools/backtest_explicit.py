#!/usr/bin/env python3
"""Backtest runner that requires every strategy parameter explicitly.

This mirrors ``tools/auto_explicit.py``. Use it for research/backtest commands
when you want to guarantee that the run does not silently depend on
StrategyConfig defaults.

Optional audit mode:
    pass ``--all-signals signals.txt --filter-preset high_growth_hour_side`` to
    append an "All Signals Audit" Excel sheet. The backtest still executes only
    ``--signals`` (usually the filtered executable file), but the workbook shows
    every raw provider signal with either EXECUTED result details or SKIP reason.
"""
from __future__ import annotations

import argparse
import glob
import json
import sys
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from xauusd_trading import CsvChartSource, StrategyConfig, parse_signals_file, run_backtest, write_backtest_outputs  # noqa: E402
from xauusd_trading.strategy.provider_filter import decide_provider_signal_filter  # noqa: E402


def _positive_int(raw: str) -> int:
    value = int(raw)
    if value < 0:
        raise argparse.ArgumentTypeError("must be >= 0")
    return value


def _positive_float(raw: str) -> float:
    value = float(raw)
    if value < 0:
        raise argparse.ArgumentTypeError("must be >= 0")
    return value


# Default M1 archive location + window, matching the main CLI's auto-fetch.
ARCHIVE_DIR = "data"
ARCHIVE_MONTHS = 2


def _sync_charts_from_mt5(symbol: str, server_offset: int, months_back: int) -> None:
    """Best-effort: pull the latest M1 from MT5 into data/ before backtesting.

    Mirrors the main CLI's pre-backtest auto-fetch. Soft-fails (warn + continue)
    when MetaTrader5 is unavailable (Linux/CI) or the terminal isn't reachable,
    so the run falls back to whatever CSVs already exist. Writes/merges per-month
    ``data/<symbol>_M1_<YYYYMM>.csv`` — matching a ``--charts
    data/<symbol>_M1_*.csv`` glob.
    """
    try:
        from xauusd_trading import (
            Mt5Connection, archive_m1_by_month, render_archive_summary,
        )
    except Exception as e:
        print(f"[mt5] skipped chart sync (import failed: {e})", file=sys.stderr)
        return
    try:
        with Mt5Connection() as conn:
            summary = archive_m1_by_month(
                conn, symbol, ARCHIVE_DIR,
                months_back=months_back,
                server_offset_hours=server_offset,
                overwrite=False,
            )
            print(render_archive_summary(summary))
            print()
    except Exception as e:
        print(f"[mt5] skipped chart sync ({e})", file=sys.stderr)


def _bool_text(raw: str) -> bool:
    text = str(raw).strip().lower()
    if text not in {"true", "false"}:
        raise argparse.ArgumentTypeError("must be true or false")
    return text == "true"


_TARGET_TOKENS = {"TP1", "TP2", "TP3", "RUN"}


def _parse_entry_targets(raw: str | None, entries: int) -> tuple[str, ...]:
    if not raw:
        return ()
    toks = tuple(t.strip().upper() for t in raw.split(",") if t.strip())
    bad = [t for t in toks if t not in _TARGET_TOKENS]
    if bad:
        raise SystemExit(f"--entry-targets tokens must be TP1/TP2/TP3/RUN (got: {','.join(bad)})")
    if len(toks) != entries:
        raise SystemExit(f"--entry-targets needs one token per entry (--entries {entries}); got {len(toks)}")
    return toks


def _expand_chart_paths(patterns: list[str]) -> list[Path]:
    out: list[Path] = []
    for pat in patterns:
        if any(ch in pat for ch in "*?["):
            matches = sorted(glob.glob(pat))
            if not matches:
                raise SystemExit(f"No files match pattern: {pat}")
            out.extend(Path(m) for m in matches)
        else:
            path = Path(pat)
            if not path.exists():
                raise SystemExit(f"Chart file not found: {pat}")
            out.append(path)
    if not out:
        raise SystemExit("No chart files provided")
    return out


def _fmt_duration(seconds: float) -> str:
    seconds = max(0, int(seconds))
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h:d}h {m:02d}m {s:02d}s"
    if m:
        return f"{m:d}m {s:02d}s"
    return f"{s:d}s"


class Heartbeat:
    def __init__(self, label: str, interval_seconds: float, *, enabled: bool = True):
        self.label = label
        self.interval_seconds = max(1.0, float(interval_seconds))
        self.enabled = enabled
        self._stop = threading.Event()
        self._thread: threading.Thread | None = None
        self._start = 0.0

    def __enter__(self):
        if not self.enabled:
            return self
        self._start = time.time()
        print(f"[{self.label}] started", file=sys.stderr, flush=True)
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()
        return self

    def __exit__(self, exc_type, exc, tb):
        if not self.enabled:
            return False
        self._stop.set()
        if self._thread is not None:
            self._thread.join(timeout=1.0)
        print(f"[{self.label}] finished after {_fmt_duration(time.time() - self._start)}", file=sys.stderr, flush=True)
        return False

    def _run(self) -> None:
        while not self._stop.wait(self.interval_seconds):
            print(f"[{self.label}] still running... elapsed {_fmt_duration(time.time() - self._start)}", file=sys.stderr, flush=True)


def _signal_signature(sig) -> tuple:
    """Stable identity across raw GMT+7 and filtered GMT+3 files."""
    return (
        sig.signal_time_chart.isoformat(sep=" "),
        sig.side,
        round(float(sig.range_low), 3),
        round(float(sig.range_high), 3),
        round(float(sig.sl), 3),
        round(float(sig.tp1), 3),
        round(float(sig.tp2), 3),
        round(float(sig.tp3), 3),
    )


def _source_tz_label(offset: int) -> str:
    return f"GMT+{offset}" if offset >= 0 else f"GMT{offset}"


def _build_all_signals_audit_rows(raw_signals: list, executed_signals: list, result: dict, preset: str) -> list[dict]:
    result_by_key = {r.get("signal_key"): r for r in result.get("rows", []) or []}
    executed_by_signature: dict[tuple, tuple] = {}
    for sig in executed_signals:
        executed_by_signature[_signal_signature(sig)] = (sig, result_by_key.get(sig.signal_key, {}))

    rows: list[dict] = []
    for raw in raw_signals:
        sig_key = _signal_signature(raw)
        executed = executed_by_signature.get(sig_key)
        anomalies = "; ".join(raw.anomalies or [])
        if executed is not None:
            exec_sig, exec_row = executed
            decision = "EXECUTED"
            reason = "Passed filter and present in executable filtered signal file."
            status = exec_row.get("status")
            pnl = exec_row.get("pnl")
            trading_pnl = exec_row.get("trading_pnl")
            bonus = exec_row.get("bonus")
            closed_lots = exec_row.get("closed_lots")
            equity_before = exec_row.get("equity_before")
            equity_after = exec_row.get("equity_after")
            executed_signal_key = exec_sig.signal_key
        else:
            filter_decision = decide_provider_signal_filter(
                side=raw.side,
                source_time=raw.signal_time_source,
                preset=preset,
                source_tz_offset=raw.source_tz_offset,
                chart_tz_offset=3,
            )
            decision = "SKIP"
            if filter_decision.keep:
                reason = "Passed filter preset but was not found in executable filtered signal file; check the generated filtered file or manual edits."
            else:
                reason = filter_decision.reason
            if raw.structural_anomaly:
                reason = f"{reason}; structural anomaly: {anomalies}"
            status = pnl = trading_pnl = bonus = closed_lots = equity_before = equity_after = None
            executed_signal_key = None

        rows.append({
            "raw_signal_key": raw.signal_key,
            "executed_signal_key": executed_signal_key,
            "decision": decision,
            "reason": reason,
            "raw_date": raw.source_date,
            "raw_time_source": raw.source_time_text,
            "source_tz": _source_tz_label(raw.source_tz_offset),
            "chart_time_gmt3": raw.signal_time_chart,
            "side": raw.side,
            "range_low": raw.range_low,
            "range_high": raw.range_high,
            "SL": raw.sl,
            "TP1": raw.tp1,
            "TP2": raw.tp2,
            "TP3": raw.tp3,
            "signal_status": status,
            "net_pnl": pnl,
            "trading_pnl": trading_pnl,
            "bonus": bonus,
            "closed_lots": closed_lots,
            "equity_before": equity_before,
            "equity_after": equity_after,
            "anomalies": anomalies,
        })
    return rows


def _append_audit_sheet(workbook_path: Path, audit_rows: list[dict]) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(workbook_path)
    if "All Signals Audit" in wb.sheetnames:
        del wb["All Signals Audit"]
    ws = wb.create_sheet("All Signals Audit", 0)

    headers = [
        "Raw Signal Key", "Executed Signal Key", "Decision", "Reason",
        "Raw Date", "Raw Time", "Source TZ", "Chart Time GMT+3", "Side",
        "Range Low", "Range High", "SL", "TP1", "TP2", "TP3",
        "Signal Status", "Net P&L", "Trading P&L", "Bonus", "Closed Lots",
        "Equity Before", "Equity After", "Anomalies",
    ]
    keys = [
        "raw_signal_key", "executed_signal_key", "decision", "reason",
        "raw_date", "raw_time_source", "source_tz", "chart_time_gmt3", "side",
        "range_low", "range_high", "SL", "TP1", "TP2", "TP3",
        "signal_status", "net_pnl", "trading_pnl", "bonus", "closed_lots",
        "equity_before", "equity_after", "anomalies",
    ]

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    exec_fill = PatternFill("solid", fgColor="C6EFCE")
    skip_fill = PatternFill("solid", fgColor="EAEAEA")
    loss_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(border_style="thin", color="BFBFBF")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = '"$"#,##0.00;[Red]-"$"#,##0.00'

    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = grid

    money_cols = {17, 18, 19, 21, 22}
    for r, row in enumerate(audit_rows, start=2):
        decision = row.get("decision")
        fill = exec_fill if decision == "EXECUTED" else skip_fill
        if decision == "EXECUTED" and (row.get("net_pnl") or 0) < 0:
            fill = loss_fill
        for c, key in enumerate(keys, start=1):
            value = row.get(key)
            if key == "chart_time_gmt3" and value is not None:
                value = value.strftime("%Y-%m-%d %H:%M")
            cell = ws.cell(row=r, column=c, value=value)
            cell.fill = fill
            cell.border = grid
            if c in money_cols:
                cell.number_format = money_fmt

    ws.freeze_panes = "A2"
    if audit_rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(audit_rows) + 1}"
    for column_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(10, min(45, length + 2))
    wb.save(workbook_path)


def add_required_strategy_args(p: argparse.ArgumentParser) -> None:
    strategy = p.add_argument_group("required strategy contract")
    strategy.add_argument("--initial-capital", type=_positive_float, required=True)
    strategy.add_argument("--sizing-mode", choices=["fixed", "risk"], required=True)
    strategy.add_argument("--lot", type=_positive_float, required=True)
    strategy.add_argument("--risk", type=_positive_float, required=True)
    strategy.add_argument("--minimum-lot", type=_positive_float, required=True)
    strategy.add_argument("--lot-step", type=_positive_float, required=True)
    strategy.add_argument("--bonus-per-closed-lot", type=_positive_float, required=True)
    strategy.add_argument("--entries", type=int, required=True)
    strategy.add_argument("--entry-ladder", choices=["signal_range_3", "range_uniform", "range_to_sl"], required=True)
    strategy.add_argument("--entry-sl-gap", type=_positive_float, required=True)
    strategy.add_argument("--activation-delay", type=_positive_int, required=True)
    strategy.add_argument("--pending-expiry", type=_positive_int, required=True)
    strategy.add_argument("--max-hold", type=_positive_int, required=True)
    strategy.add_argument("--sl-multiplier", type=_positive_float, required=True)
    strategy.add_argument("--final-target", choices=["TP1", "TP2", "TP3"], required=True)
    strategy.add_argument("--lock-after-tp1", type=_bool_text, required=True)
    strategy.add_argument("--lock-after-tp2", type=_bool_text, required=True)
    strategy.add_argument("--tp1-lock-delay-minutes", type=_positive_int, required=True)
    strategy.add_argument("--tp2-lock-delay-minutes", type=_positive_int, required=True)
    strategy.add_argument("--profit-lock-mode", choices=["tp_levels", "bep_plus_half_tp1"], required=True)
    strategy.add_argument("--bep-trigger-distance", type=_positive_float, required=True)
    strategy.add_argument("--tp1-lock-fraction", type=float, required=True)
    strategy.add_argument("--tp2-lock-target", choices=["TP1", "TP2"], required=True)
    strategy.add_argument("--runner-after-tp3", type=_bool_text, required=True)
    strategy.add_argument("--tp3-lock-target", choices=["TP2"], required=True)
    strategy.add_argument("--trailing-open-distance", type=_positive_float, required=True,
                          help="Virtual trailing-open entry distance in price units; 0 disables.")
    strategy.add_argument("--trailing-close-distance", type=_positive_float, required=True,
                          help="Trailing-close (ratcheting) stop distance in price units; 0 disables.")


def add_scale_out_args(p: argparse.ArgumentParser) -> None:
    """Optional multi-entry scale-out exit. All default off so omitting them keeps
    the DD40/TRAILING-0.5 contract byte-identical; sweep by passing them explicitly."""
    g = p.add_argument_group("optional scale-out exit (default off)")
    g.add_argument("--scale-out-at-tp1", type=_bool_text, default=False,
                   help="At TP1 touch, close the worst open leg (furthest from signal SL). Needs >=2 filled legs.")
    g.add_argument("--scale-out-at-tp2", type=_bool_text, default=False,
                   help="At TP2 touch, close the worst remaining open leg.")
    g.add_argument("--bep-after-tp1", type=_bool_text, default=False,
                   help="At TP1, move remaining legs' stop to entry +/- --bep-buffer.")
    g.add_argument("--bep-buffer", type=_positive_float, default=0.0,
                   help="Profit locked beyond entry (price units) when --bep-after-tp1; use >=0.40 for live.")
    g.add_argument("--trailing-close-after-stage", type=_positive_int, default=0,
                   help="Trailing-close engages only at/after this stage (0=from open, 1=after TP1, 2=after TP2).")
    g.add_argument("--runner-final-cap", choices=["tp3", "none"], default="tp3",
                   help="tp3 = trailing remainder force-closes at the final target; none = pure trail.")
    g.add_argument("--shared-sl", type=_bool_text, default=False,
                   help="All entries share ONE stop level (anchored on the first entry) instead of "
                        "each entry getting its own per-price stop. Risk-sizing uses each leg's "
                        "real distance to the shared level. Sweep by passing true/false.")
    g.add_argument("--entry-targets", default=None, metavar="T1,T2,...",
                   help="Per-entry targets, one token per entry from {TP1,TP2,TP3,RUN}; RUN holds "
                        "at TP3 then trails by --trailing-close-distance. Empty = single "
                        "--final-target. Length must equal --entries.")
    g.add_argument("--bep-after-move", type=_positive_float, default=0.0,
                   help="Per-leg break-even+ (per-entry-targets mode): once a filled leg is this "
                        "many price units in favour, move its SL to entry +/- --bep-buffer. 0=off.")
    g.add_argument("--runner-trail-from", choices=["TP1", "TP2", "TP3"], default="TP3",
                   help="RUN legs engage their trailing stop when this TP is touched, then trail "
                        "by --trailing-close-distance. Never trails from entry. Default TP3.")


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Run backtest with no hidden strategy defaults.")
    p.add_argument("--signals", required=True, help="Executable/filtered signal file to backtest.")
    p.add_argument("--all-signals", default=None, help="Optional raw provider signals.txt for All Signals Audit sheet.")
    p.add_argument("--filter-preset", default="high_growth_hour_side", choices=["all", "no_bad_hours", "best_hours", "high_growth_hour_side", "research_month_hour_side"])
    p.add_argument("--charts", required=True, nargs="+")
    p.add_argument("--start-date", default=None, metavar="YYYY-MM-DD",
                   help="Only backtest signals on/after this date (chart time GMT+3). "
                        "Equity starts at --initial-capital on the first surviving signal.")
    p.add_argument("--end-date", default=None, metavar="YYYY-MM-DD",
                   help="Only backtest signals on/before this date (chart time GMT+3, "
                        "inclusive of the whole day).")
    p.add_argument("--output-dir", required=True)
    p.add_argument("--mt5-history", default=None, metavar="FILE",
                   help="MT5 History export (.xlsx/.csv/.html). Matches live positions to "
                        "backtest entries by the order Comment (= entry_key) and adds a LIVE "
                        "column group + live R to the Per-Entry Detail.")
    p.add_argument("--exclude-structural-anomalies", action="store_true")
    p.add_argument("--max-drawdown-limit-pct", type=float, required=True)
    p.add_argument("--fail-on-drawdown-limit", action="store_true")
    p.add_argument("--progress-interval-seconds", type=float, required=True)
    add_required_strategy_args(p)
    add_scale_out_args(p)
    add_chart_sync_args(p)
    return p


def add_chart_sync_args(p: argparse.ArgumentParser) -> None:
    """Pull fresh M1 from MT5 into data/ before the backtest (default on)."""
    g = p.add_argument_group("chart sync (fetch latest M1 from MT5 before the run)")
    g.add_argument("--sync-charts", type=_bool_text, default=True,
                   help="Fetch the latest M1 from MT5 into data/ before running (default true). "
                        "Soft-fails to the existing CSVs when MT5 is unavailable (e.g. Linux/CI). "
                        "Pass false to skip the fetch.")
    g.add_argument("--mt5-symbol", default="XAUUSD",
                   help="MT5 symbol to sync (default XAUUSD); writes data/<symbol>_M1_<YYYYMM>.csv.")
    g.add_argument("--mt5-server-offset", type=int, default=3,
                   help="Broker server GMT offset used by the sync (default 3).")
    g.add_argument("--sync-months", type=_positive_int, default=ARCHIVE_MONTHS,
                   help=f"Months of M1 history to pull on sync (default {ARCHIVE_MONTHS}).")


def config_from_args(args: argparse.Namespace) -> StrategyConfig:
    if args.entries < 1:
        raise SystemExit("--entries must be >= 1")
    if args.tp1_lock_fraction < 0 or args.tp1_lock_fraction > 1:
        raise SystemExit("--tp1-lock-fraction must be between 0 and 1")
    if args.sizing_mode == "risk" and args.risk <= 0:
        raise SystemExit("--risk must be > 0 when --sizing-mode risk")
    if args.sizing_mode == "fixed" and args.lot <= 0:
        raise SystemExit("--lot must be > 0 when --sizing-mode fixed")
    if not 0 <= args.trailing_close_after_stage <= 3:
        raise SystemExit("--trailing-close-after-stage must be between 0 and 3")

    return StrategyConfig(
        initial_capital=args.initial_capital,
        sizing_mode=args.sizing_mode,
        lot_per_entry=args.lot,
        risk_per_signal=args.risk,
        minimum_lot=args.minimum_lot,
        lot_step=args.lot_step,
        bonus_per_closed_lot=args.bonus_per_closed_lot,
        entry_count=args.entries,
        entry_ladder=args.entry_ladder,
        entry_sl_gap=args.entry_sl_gap,
        activation_delay_minutes=args.activation_delay,
        pending_expiry_minutes=args.pending_expiry,
        max_hold_minutes=args.max_hold,
        sl_multiplier=args.sl_multiplier,
        final_target=args.final_target,
        lock_after_tp1=args.lock_after_tp1,
        lock_after_tp2=args.lock_after_tp2,
        tp1_lock_delay_minutes=args.tp1_lock_delay_minutes,
        tp2_lock_delay_minutes=args.tp2_lock_delay_minutes,
        profit_lock_mode=args.profit_lock_mode,
        bep_trigger_distance=args.bep_trigger_distance,
        tp1_lock_fraction=args.tp1_lock_fraction,
        tp2_lock_target=args.tp2_lock_target,
        runner_after_tp3=args.runner_after_tp3,
        tp3_lock_target=args.tp3_lock_target,
        trailing_open_distance=args.trailing_open_distance,
        trailing_close_distance=args.trailing_close_distance,
        scale_out_at_tp1=args.scale_out_at_tp1,
        scale_out_at_tp2=args.scale_out_at_tp2,
        bep_after_tp1=args.bep_after_tp1,
        bep_buffer=args.bep_buffer,
        trailing_close_after_stage=args.trailing_close_after_stage,
        runner_no_final_cap=(args.runner_final_cap == "none"),
        shared_sl=args.shared_sl,
        per_entry_targets=_parse_entry_targets(args.entry_targets, args.entries),
        bep_after_move=args.bep_after_move,
        runner_trail_from=args.runner_trail_from,
    )


def _parse_date(value: str | None) -> datetime | None:
    return datetime.strptime(value, "%Y-%m-%d") if value else None


def filter_signals_by_date(signals, start_date: str | None, end_date: str | None):
    """Keep signals whose chart-time falls within [start_date 00:00, end_date 24:00).

    Dates are 'YYYY-MM-DD' (chart time GMT+3) or None; end_date includes the whole
    day. run_backtest then starts equity at --initial-capital on the first
    surviving signal, so this is how a run begins from a specific date.
    """
    start = _parse_date(start_date)
    end = _parse_date(end_date)
    if end is not None:
        end = end + timedelta(days=1)
    kept = []
    for s in signals:
        t = s.signal_time_chart
        if start is not None and t < start:
            continue
        if end is not None and t >= end:
            continue
        kept.append(s)
    return kept


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    config = config_from_args(args)
    progress_enabled = args.progress_interval_seconds > 0

    # Sync the chart archive from MT5 first so the backtest runs on fresh bars.
    if args.sync_charts:
        _sync_charts_from_mt5(args.mt5_symbol, args.mt5_server_offset, args.sync_months)

    signals = filter_signals_by_date(
        parse_signals_file(Path(args.signals)), args.start_date, args.end_date
    )
    with Heartbeat("chart load", args.progress_interval_seconds, enabled=progress_enabled):
        chart = CsvChartSource(_expand_chart_paths(args.charts))
    with Heartbeat("backtest", args.progress_interval_seconds, enabled=progress_enabled):
        result = run_backtest(
            signals,
            chart,
            config,
            exclude_structural_anomalies=args.exclude_structural_anomalies,
        )

    # Merge real MT5 fills (matched by Comment == entry_key) into the entry rows.
    if args.mt5_history:
        from xauusd_trading.reporting.mt5_history import attach_live_history, parse_mt5_history
        live = parse_mt5_history(args.mt5_history)
        info = attach_live_history(result, live)
        print(f"[mt5-history] matched {info['matched']} live position(s) to backtest entries; "
              f"{len(info['unmatched'])} live comment(s) unmatched.", file=sys.stderr)
        if info["unmatched"]:
            print(f"[mt5-history] unmatched comments: {', '.join(info['unmatched'][:10])}"
                  f"{' ...' if len(info['unmatched']) > 10 else ''}", file=sys.stderr)

    summary = {k: v for k, v in result.items() if k not in {"rows", "entry_rows"}}
    dd_abs = abs(min(0.0, float(result.get("max_drawdown_pct", 0.0) or 0.0)))
    summary["max_drawdown_limit_pct"] = args.max_drawdown_limit_pct
    summary["passes_drawdown_limit"] = dd_abs <= args.max_drawdown_limit_pct

    audit_rows: list[dict] = []
    if args.all_signals:
        raw_signals = parse_signals_file(Path(args.all_signals))
        audit_rows = _build_all_signals_audit_rows(raw_signals, signals, result, args.filter_preset)
        summary["all_signals_raw_count"] = len(raw_signals)
        summary["all_signals_audit_executed"] = sum(1 for r in audit_rows if r["decision"] == "EXECUTED")
        summary["all_signals_audit_skipped"] = sum(1 for r in audit_rows if r["decision"] == "SKIP")

    print(json.dumps(summary, indent=2, default=str))

    with Heartbeat("Excel write", args.progress_interval_seconds, enabled=progress_enabled):
        path = write_backtest_outputs(result, Path(args.output_dir))
        if audit_rows:
            _append_audit_sheet(path, audit_rows)
    print(f"\nWrote Excel output to {path.resolve()}", file=sys.stderr)

    if args.fail_on_drawdown_limit and not summary["passes_drawdown_limit"]:
        print(
            f"Max drawdown {result.get('max_drawdown_pct', 0.0):.2f}% exceeds limit "
            f"-{args.max_drawdown_limit_pct:.2f}%.",
            file=sys.stderr,
        )
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())