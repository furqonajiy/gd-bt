#!/usr/bin/env python3
"""Per-signal LIVE vs BACKTEST reconciliation -> a colour-coded Excel workbook.

This is the agreed reconcile contract for "did the live executor do what the
backtest said, and where did it diverge, and *why*". Unlike
``tools/reconcile_report_html.py`` (which prints an aggregate + exit-type
bottom line), this produces the **per-signal table** the operator reads:

    Signal | Side | BT legs/entry/open/close/exit/$ | LV legs/entry/open/close/$
           | dEntry | dOpen(min) | dClose(min) | d$ | Discrepancy | Description

with one row per signal, a plain-English **Description** of each unique
discrepancy (including the counterfactual P&L -- e.g. "you closed this by hand;
left alone the model banks $X" / "late close earned +$Y"), and **colour coding**
so wins/losses and discrepancy classes are obvious at a glance.

Design choices (the reconcile contract):

  1. **Same lot as LIVE.** The backtest dollars are recomputed from entry/exit
     PRICES at the live fixed lot (``--lot`` -> ``--usd-per-point``), so the
     comparison is apples-to-apples no matter how the backtest workbook was
     sized (risk-sized or fixed). A XAUUSD $1.00 move = $1.00 per 0.01 lot, so
     ``--usd-per-point`` defaults to 1.0.
  2. **Same columns** as the reviewed terminal table.
  3. **A Description per unique discrepancy class** (see ``DISCREPANCY`` below).
  4. **Excel colour coding** -- d$ green (live better) / red (live worse), and
     each discrepancy class gets its own fill; a Legend sheet explains them.

It reads the **MT5 Trade History Report (HTML)** you export from the terminal
(History tab -> right-click -> Report -> HTML) and the per-strategy backtest
**Per-Entry Detail** workbook. Match key is the order comment
``[TAG-]MMDD#DD.N`` <-> backtest key ``YYYY-MM-DD#DD.N`` (matched on
month-day-signal-leg, single-report assumption).

For a clean reconcile the backtest workbook MUST cover the live window -- its
chart/tick data has to extend past the last live trade, or the live signals
won't exist in the backtest and nothing matches. Regenerate the backtest right
after you export the HTML, e.g. (fixed lot to match live exactly):

    python tools/backtest_hybrid.py --signals signals/tsl18.txt \
        --charts "data/XAUUSD_M1_*_ELEV8.csv" --ticks "data/ticks/XAUUSD_TICK_*_ELEV8.csv" \
        --sizing-mode fixed --lot 0.01 --risk 0.01  ...same strategy flags... \
        --output-dir reports/TSL18_fixed --start-date 2026-06-30 --end-date 2026-06-30

    python tools/reconcile_live_vs_backtest.py \
        --report ReportHistory.html --backtest reports/TSL18_fixed.xlsx \
        --tag TSL18 --lot 0.01 --out reports/TSL18_recon_2026-06-30.xlsx
"""
from __future__ import annotations

import argparse
import datetime as dt
import re
from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------- #
# HTML report parsing                                                         #
# --------------------------------------------------------------------------- #

def _decode_report(path: str) -> str:
    raw = open(path, "rb").read()
    for enc in ("utf-16", "utf-8", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeError:
            continue
    return raw.decode("utf-8", "ignore")


def _cells(row_html: str) -> list[str]:
    return [
        re.sub(r"<[^>]+>", "", c).replace("&nbsp;", " ").strip()
        for c in re.findall(r"<td[^>]*>(.*?)</td>", row_html, re.S)
    ]


def _parse_dt(s: str | None) -> dt.datetime | None:
    if not s:
        return None
    s = str(s).strip()
    for f in ("%Y.%m.%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y.%m.%d %H:%M", "%Y-%m-%d %H:%M"):
        try:
            return dt.datetime.strptime(s, f)
        except ValueError:
            continue
    return None


def _f(s: str) -> float | None:
    try:
        return float(str(s).replace(" ", "").replace(",", ""))
    except (ValueError, AttributeError):
        return None


# section boundaries are located by their <b>Header</b> markers
def _section_bounds(text: str) -> dict[str, tuple[int, int]]:
    order = ["Positions", "Orders", "Deals", "Open Positions", "Results"]
    starts: dict[str, int] = {}
    for name in order:
        m = re.search(r"<b>" + re.escape(name) + r"</b>", text)
        if m:
            starts[name] = m.start()
    bounds: dict[str, tuple[int, int]] = {}
    items = sorted(starts.items(), key=lambda kv: kv[1])
    for i, (name, st) in enumerate(items):
        end = items[i + 1][1] if i + 1 < len(items) else len(text)
        bounds[name] = (st, end)
    return bounds


@dataclass
class LiveFill:
    sig: int
    leg: int
    side: str
    open_t: dt.datetime | None
    open_px: float | None
    close_t: dt.datetime | None
    close_px: float | None
    profit: float
    is_open: bool = False
    exit_tail: str | None = None  # /sl, /tp3, /time, /expert, /lock-tp1 ...


_KEY_RE = re.compile(r"#(\d+)\.(\d+)")


def parse_live(text: str, tag: str) -> list[LiveFill]:
    """Closed Positions + still-Open Positions for one strategy tag."""
    bounds = _section_bounds(text)
    fills: list[LiveFill] = []
    com_re = re.compile(rf"{re.escape(tag)}-(\d{{2}})(\d{{2}})#(\d+)\.(\d+)")
    tail_by_key: dict[tuple[int, int, int], str] = {}

    # exit tail (/sl, /time, ...) lives on the close DEAL comment
    if "Deals" in bounds:
        ds, de = bounds["Deals"]
        deal_re = re.compile(rf"{re.escape(tag)}-(\d{{2}})(\d{{2}})#(\d+)/(\S+)")
        for row in re.findall(r"<tr[^>]*>(.*?)</tr>", text[ds:de], re.S):
            c = _cells(row)
            if not c:
                continue
            m = deal_re.search(c[-1])
            if m:
                mm, dd, sig, tail = m.groups()
                tail_by_key[(int(mm) * 100 + int(dd), int(sig), 0)] = tail

    # closed positions: Time|Pos|Sym|Type|Comment|Vol|OpenPx|SL|TP|CloseT|ClosePx|Comm|Swap|Profit
    if "Positions" in bounds:
        ps, pe = bounds["Positions"]
        for row in re.findall(r"<tr[^>]*>(.*?)</tr>", text[ps:pe], re.S):
            c = _cells(row)
            if len(c) < 14:
                continue
            m = com_re.match(c[4])
            if not m:
                continue
            mm, dd, sig, leg = (int(x) for x in m.groups())
            prof = _f(c[13])
            if prof is None:
                continue
            tail = tail_by_key.get((mm * 100 + dd, sig, 0))
            fills.append(LiveFill(sig, leg, c[3].upper(), _parse_dt(c[0]), _f(c[6]),
                                  _parse_dt(c[9]), _f(c[10]), prof, False, tail))

    # still-open positions: Time|Pos|Sym|Type|Vol|Price|SL|TP|MktPx|Swap|Profit|Comment
    if "Open Positions" in bounds:
        os_, oe = bounds["Open Positions"]
        for row in re.findall(r"<tr[^>]*>(.*?)</tr>", text[os_:oe], re.S):
            c = _cells(row)
            if len(c) < 12:
                continue
            m = com_re.match(c[-1])
            if not m:
                continue
            mm, dd, sig, leg = (int(x) for x in m.groups())
            prof = _f(c[10]) or 0.0
            fills.append(LiveFill(sig, leg, c[3].upper(), _parse_dt(c[0]), _f(c[5]),
                                  None, _f(c[8]), prof, True, None))
    return fills


# --------------------------------------------------------------------------- #
# Backtest Per-Entry workbook parsing                                         #
# --------------------------------------------------------------------------- #

@dataclass
class BtLeg:
    sig: int
    leg: int
    side: str
    fill_t: dt.datetime | None
    entry_px: float | None
    status: str
    exit_t: dt.datetime | None
    exit_px: float | None
    source: str = ""  # TICK / M1 -- from the Per-Entry "Data Source" column


def parse_backtest(path: str) -> list[BtLeg]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    sheet = next((s for s in wb.sheetnames if "Entry" in s), wb.sheetnames[-1])
    ws = wb[sheet]
    # header row = the one containing "Entry Key"
    hdr_row = next(
        (r for r in range(1, min(6, ws.max_row) + 1)
         if any(str(ws.cell(row=r, column=c).value).strip() == "Entry Key"
                for c in range(1, ws.max_column + 1))),
        2,
    )
    hdr = {str(ws.cell(row=hdr_row, column=c).value).strip(): c
           for c in range(1, ws.max_column + 1)}

    def col(name: str) -> int | None:
        return hdr.get(name)

    legs: list[BtLeg] = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        key = ws.cell(row=r, column=col("Entry Key")).value
        if not key:
            continue
        m = _KEY_RE.search(str(key))
        if not m:
            continue
        sig, leg = int(m.group(1)), int(m.group(2))
        gv = lambda n: (ws.cell(row=r, column=col(n)).value if col(n) else None)
        legs.append(BtLeg(
            sig, leg, str(gv("Side") or "").upper(),
            _parse_dt(gv("Fill Time")), _f(gv("Entry Price")),
            str(gv("Status") or ""), _parse_dt(gv("Exit Time")), _f(gv("Exit Price")),
            str(gv("Data Source") or "").upper(),
            ))
    return legs


# --------------------------------------------------------------------------- #
# Per-signal aggregation + discrepancy classification                         #
# --------------------------------------------------------------------------- #

def _signed_pnl(side: str, entry: float | None, exit_px: float | None,
                usd_per_point: float) -> float | None:
    if entry is None or exit_px is None:
        return None
    direction = 1.0 if side.startswith("B") else -1.0
    return (exit_px - entry) * direction * usd_per_point


@dataclass
class SignalRow:
    sig: int
    side: str
    bt_legs: int = 0
    bt_entry: float | None = None
    bt_open: dt.datetime | None = None
    bt_close: dt.datetime | None = None
    bt_exit: str = ""
    bt_src: str = ""  # TICK / M1 / mixed -- backtest data source for this signal
    bt_pnl: float = 0.0
    lv_legs: int = 0
    lv_fills: int = 0
    lv_entry: float | None = None
    lv_open: dt.datetime | None = None
    lv_close: dt.datetime | None = None
    lv_pnl: float = 0.0
    lv_open_count: int = 0
    lv_has_open: bool = False
    d_entry: float | None = None
    d_open_min: int | None = None
    d_close_min: int | None = None
    d_pnl: float = 0.0
    klass: str = ""
    description: str = ""


# discrepancy classes: key -> (label, fill colour ARGB, legend text)
DISCREPANCY = {
    "MATCH":        ("Match",            "FFD9EAD3", "Live tracked the backtest within tolerance (|d$|<$5, same leg count, exit within 3 min)."),
    "UNDER_FILL":   ("Ladder under-fill","FFFCE5CD", "Live filled fewer ladder rungs than the model (fast move -> trailing-open stops never triggered). The missing legs change the P&L."),
    "OVER_FILL":    ("Reopen / churn",   "FFEAD1DC", "Live filled the leg MORE than once (a manual close round-tripped by the reopen pass, or trailing re-arm). Extra fills move the P&L."),
    "LATE_ARM":     ("Late arm",         "FFD0E0F0", "Live armed the entry materially later than the model (executor catch-up / deployment start), so the entry basis differs."),
    "LATE_CLOSE":   ("Exit-time drift",  "FFFFF2CC", "Entry matched but live exit fired minutes later/earlier than the model, so the exit price (and P&L) differs."),
    "MANUAL_CLOSE": ("Manual close",     "FFF4CCCC", "Live closed earlier than the model exit with no engine exit tag -> a hand close. The model would have held the position."),
    "STILL_OPEN":   ("Still open live",  "FFEFEFEF", "The live leg(s) are still open; the backtest already closed. No realized comparison yet."),
    "ENTRY_DRIFT":  ("Entry drift",      "FFFFFFCC", "Entry price differs by >=2 pt (trailing-open caught a different rebound), small timing/price divergence."),
    "NO_LIVE":      ("Not traded live",  "FFE0E0E0", "The backtest filled this signal but live never did (strategy not deployed in this window, or signal filtered live)."),
    "NO_BT":        ("Not in backtest",  "FFE0E0E0", "Live traded a signal the backtest did not fill (feed/timing mismatch)."),
}


def _mins(a: dt.datetime | None, b: dt.datetime | None) -> int | None:
    if a and b:
        return round((a - b).total_seconds() / 60)
    return None


def _money(x: float) -> str:
    return f"${x:+,.2f}"


def classify(row: SignalRow) -> tuple[str, str]:
    """Return (class_key, human description) for a per-signal row."""
    bt, lv = row.bt_pnl, row.lv_pnl
    delta = lv - bt

    # ---- coverage extremes -------------------------------------------------
    if row.bt_legs and not row.lv_legs:
        return "NO_LIVE", (
            f"Backtest filled {row.bt_legs} leg(s) for {_money(bt)} but live never traded this "
            f"signal (strategy not live for this signal, or filtered). No live exposure taken."
        )
    if row.lv_legs and not row.bt_legs:
        return "NO_BT", (
            f"Live traded {row.lv_legs} leg(s) for {_money(lv)} but the backtest did not fill this "
            f"signal (feed/timing mismatch). Investigate the feed."
        )
    if row.lv_has_open:
        return "STILL_OPEN", (
            f"{row.lv_legs} live leg(s) still open ({_money(lv)} floating); the model already "
            f"closed for {_money(bt)}. Re-export once they close for a realized compare."
        )

    # ---- ladder under-fill (size mismatch) ---------------------------------
    if row.lv_legs < row.bt_legs:
        missing = row.bt_legs - row.lv_legs
        verb = "saved you" if delta > 0 else "cost you"
        return "UNDER_FILL", (
            f"Live filled only {row.lv_legs} of {row.bt_legs} rungs - a fast move left {missing} "
            f"trailing-open rung(s) untriggered. At matched lot the under-fill {verb} "
            f"{_money(abs(delta))} (live {_money(lv)} vs model {_money(bt)}). Luck, not edge - "
            f"on a fast WIN the same under-fill would cost you the gain."
        )

    # ---- late arm / deployment start (entry materially later) --------------
    if row.d_open_min is not None and row.d_open_min >= 20:
        if delta >= 0:
            tail = (f"the late entry dodged the model's loss - you booked {_money(lv)} vs the "
                    f"model's {_money(bt)} ({_money(delta)}).")
        else:
            tail = (f"the late entry gave a worse basis - you booked {_money(lv)} vs the model's "
                    f"{_money(bt)} ({_money(delta)}).")
        return "LATE_ARM", (
            f"Live armed {row.d_open_min:+d} min after the model (executor catch-up / first "
            f"signal after deployment). " + tail
        )

    # ---- reopen / re-arm churn (more fills than legs) ----------------------
    if row.lv_fills > row.lv_legs:
        extra = row.lv_fills - row.lv_legs
        return "OVER_FILL", (
            f"Live re-filled {extra} extra time(s) beyond {row.lv_legs} leg(s) - either a hand "
            f"close round-tripped by the reopen pass, or a trailing-open re-arm after a stop. "
            f"If it was your manual close, leaving it alone gives the model's "
            f"{row.bt_exit or 'exit'} = {_money(bt)}; live got {_money(lv)} ({_money(delta)})."
        )

    # ---- manual close (early, no engine tail) ------------------------------
    if (row.d_close_min is not None and row.d_close_min <= -5
            and not (row.bt_exit and row.bt_exit.startswith(("TP", "SL")))):
        return "MANUAL_CLOSE", (
            f"Live closed {abs(row.d_close_min)} min EARLIER than the model with no engine exit "
            f"tag - a hand close. Left alone the model holds to {row.bt_exit or 'exit'} for "
            f"{_money(bt)} (you got {_money(lv)}, {_money(delta)})."
        )

    # ---- exit-time drift ---------------------------------------------------
    if row.d_close_min is not None and abs(row.d_close_min) >= 5:
        better = "earned" if delta >= 0 else "lost"
        when = "later" if row.d_close_min > 0 else "earlier"
        return "LATE_CLOSE", (
            f"Entry matched but live exit fired {abs(row.d_close_min)} min {when} than the model -> "
            f"{better} {_money(abs(delta))} ({'late close better' if delta >= 0 and row.d_close_min > 0 else 'exit-timing drift'}): "
            f"live {_money(lv)} vs model {_money(bt)}."
        )

    # ---- small entry drift -------------------------------------------------
    if row.d_entry is not None and abs(row.d_entry) >= 2.0:
        return "ENTRY_DRIFT", (
            f"Entry differs {row.d_entry:+.2f} pt (trailing-open caught a different rebound); "
            f"net {_money(delta)} (live {_money(lv)} vs model {_money(bt)})."
        )

    # ---- otherwise a match -------------------------------------------------
    return "MATCH", (
        f"Live tracked the model: entry {(row.d_entry or 0.0):+.2f} pt, exit "
        f"{(row.d_close_min or 0):+d} min, P&L {_money(delta)} "
        f"(live {_money(lv)} vs model {_money(bt)})."
    )


def build_rows(live: list[LiveFill], bt: list[BtLeg], usd_per_point: float,
               date_prefix_mmdd: set[int] | None) -> list[SignalRow]:
    def in_scope_sig(sig_keys: set[int]) -> bool:
        return True  # signals are per-day; filtering is by mmdd on the legs

    # group
    lv_by_sig: dict[int, list[LiveFill]] = {}
    for f in live:
        lv_by_sig.setdefault(f.sig, []).append(f)
    bt_by_sig: dict[int, list[BtLeg]] = {}
    for b in bt:
        bt_by_sig.setdefault(b.sig, []).append(b)

    rows: list[SignalRow] = []
    for sig in sorted(set(lv_by_sig) | set(bt_by_sig)):
        bvals = bt_by_sig.get(sig, [])
        lvals = lv_by_sig.get(sig, [])
        bvals = [b for b in bvals if b.fill_t is not None]  # filled only
        side = (bvals[0].side if bvals else lvals[0].side) if (bvals or lvals) else ""

        row = SignalRow(sig=sig, side=side)

        # backtest aggregate (dollars recomputed at the LIVE lot from prices)
        if bvals:
            row.bt_legs = len(bvals)
            row.bt_entry = sum(b.entry_px for b in bvals if b.entry_px) / row.bt_legs
            row.bt_open = min((b.fill_t for b in bvals if b.fill_t), default=None)
            row.bt_close = max((b.exit_t for b in bvals if b.exit_t), default=None)
            row.bt_exit = ",".join(sorted({(b.status or "")[:4] for b in bvals}))
            srcs = {b.source for b in bvals if b.source}
            row.bt_src = next(iter(srcs)) if len(srcs) == 1 else ("mixed" if srcs else "?")
            row.bt_pnl = sum(
                (_signed_pnl(b.side, b.entry_px, b.exit_px, usd_per_point) or 0.0)
                for b in bvals
            )

        # live aggregate (report profit already at live lot; all fills incl re-arms)
        if lvals:
            legkeys = {f.leg for f in lvals}
            row.lv_legs = len(legkeys)
            row.lv_fills = len(lvals)
            first_per_leg = [
                sorted([f for f in lvals if f.leg == lk],
                       key=lambda x: (x.open_t or dt.datetime.min))[0]
                for lk in legkeys
            ]
            row.lv_entry = sum(f.open_px for f in first_per_leg if f.open_px) / max(1, len(first_per_leg))
            row.lv_open = min((f.open_t for f in first_per_leg if f.open_t), default=None)
            row.lv_close = max((f.close_t for f in lvals if f.close_t), default=None)
            row.lv_pnl = sum(f.profit for f in lvals)
            row.lv_has_open = any(f.is_open for f in lvals)

        if not row.bt_legs and not row.lv_legs:
            continue  # signal filled on neither side (no-fill backtest signal) -> skip

        row.d_entry = (row.lv_entry - row.bt_entry) if (row.lv_entry and row.bt_entry) else None
        row.d_open_min = _mins(row.lv_open, row.bt_open)
        row.d_close_min = _mins(row.lv_close, row.bt_close)
        row.d_pnl = row.lv_pnl - row.bt_pnl
        row.klass, row.description = classify(row)
        rows.append(row)
    return rows


# --------------------------------------------------------------------------- #
# Excel output                                                                #
# --------------------------------------------------------------------------- #

_COLS = [
    ("Signal", 8), ("Side", 6),
    ("BT legs", 8), ("BT src", 8), ("BT entry", 10), ("BT open", 9), ("BT close", 9),
    ("BT exit", 9), ("BT $", 11),
    ("LV legs", 8), ("LV entry", 10), ("LV open", 9), ("LV close", 9), ("LV $", 11),
    ("dEntry (pt)", 11), ("dOpen (min)", 11), ("dClose (min)", 12), ("d$", 11),
    ("Discrepancy", 18), ("Description", 90),
]
# 1-based column index by header name (so colour/format code never hardcodes numbers)
_CI = {name: i for i, (name, _) in enumerate(_COLS, start=1)}


def _hm(d: dt.datetime | None) -> str:
    return d.strftime("%H:%M") if isinstance(d, dt.datetime) else "-"


def write_excel(rows: list[SignalRow], out_path: str, *, tag: str, lot: float,
                window: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconcile"

    thin = Side(style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="FF1F4E78")
    hdr_font = Font(bold=True, color="FFFFFFFF", size=10)
    green = PatternFill("solid", fgColor="FFD9EAD3")
    red = PatternFill("solid", fgColor="FFF4CCCC")

    # title
    ws.cell(row=1, column=1, value=f"LIVE vs BACKTEST reconcile - {tag} - {window} - matched lot {lot}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLS))

    hrow = 3
    for ci, (name, width) in enumerate(_COLS, start=1):
        c = ws.cell(row=hrow, column=ci, value=name)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width

    r = hrow + 1
    tot_bt = tot_lv = 0.0
    for row in rows:
        tot_bt += row.bt_pnl
        tot_lv += row.lv_pnl
        klass_fill = PatternFill("solid", fgColor=DISCREPANCY[row.klass][1])
        vals = [
            row.sig, row.side,
            row.bt_legs or "-", row.bt_src or "-",
            round(row.bt_entry, 2) if row.bt_entry else "-",
            _hm(row.bt_open), _hm(row.bt_close), row.bt_exit or "-", round(row.bt_pnl, 2),
            row.lv_legs or "-", round(row.lv_entry, 2) if row.lv_entry else "-",
            _hm(row.lv_open), _hm(row.lv_close), round(row.lv_pnl, 2),
            round(row.d_entry, 2) if row.d_entry is not None else "-",
            row.d_open_min if row.d_open_min is not None else "-",
            row.d_close_min if row.d_close_min is not None else "-",
            round(row.d_pnl, 2),
            DISCREPANCY[row.klass][0], row.description,
        ]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border = border
            c.alignment = Alignment(
                horizontal="left" if ci == len(_COLS) else "center",
                vertical="center", wrap_text=(ci == len(_COLS)),
            )
        # colour the d$ cell green/red, the Discrepancy cell by class
        d_cell = ws.cell(row=r, column=_CI["d$"])
        d_cell.fill = green if row.d_pnl >= 0 else red
        d_cell.font = Font(bold=True)
        ws.cell(row=r, column=_CI["Discrepancy"]).fill = klass_fill
        # flag any non-TICK backtest source in red (the reconcile wants TICK)
        src_cell = ws.cell(row=r, column=_CI["BT src"])
        if row.bt_src and row.bt_src != "TICK":
            src_cell.fill = red
            src_cell.font = Font(bold=True, color="FF9C0006")
        # tint LV $ vs BT $
        ws.cell(row=r, column=_CI["BT $"]).font = Font(color="FF006100" if row.bt_pnl >= 0 else "FF9C0006")
        ws.cell(row=r, column=_CI["LV $"]).font = Font(color="FF006100" if row.lv_pnl >= 0 else "FF9C0006")
        r += 1

    # totals row
    tot = ws.cell(row=r, column=1, value="TOTAL")
    tot.font = Font(bold=True)
    ws.cell(row=r, column=8, value=round(tot_bt, 2)).font = Font(bold=True)
    ws.cell(row=r, column=13, value=round(tot_lv, 2)).font = Font(bold=True)
    dcell = ws.cell(row=r, column=17, value=round(tot_lv - tot_bt, 2))
    dcell.font = Font(bold=True)
    dcell.fill = green if (tot_lv - tot_bt) >= 0 else red
    ws.cell(row=r, column=18, value="net")
    ws.cell(row=r, column=19,
            value=f"At matched lot {lot}: backtest {_money(tot_bt)}, live {_money(tot_lv)}, "
                  f"live-vs-model {_money(tot_lv - tot_bt)}.")

    ws.freeze_panes = ws.cell(row=hrow + 1, column=3)

    # legend sheet
    lg = wb.create_sheet("Legend")
    lg.cell(row=1, column=1, value="Discrepancy classes").font = Font(bold=True, size=12)
    lg.cell(row=2, column=1, value="Class").font = Font(bold=True)
    lg.cell(row=2, column=2, value="Meaning").font = Font(bold=True)
    lg.column_dimensions["A"].width = 20
    lg.column_dimensions["B"].width = 110
    rr = 3
    for key, (label, color, text) in DISCREPANCY.items():
        a = lg.cell(row=rr, column=1, value=label)
        a.fill = PatternFill("solid", fgColor=color)
        lg.cell(row=rr, column=2, value=text)
        rr += 1
    lg.cell(row=rr + 1, column=1, value="d$ cell").font = Font(bold=True)
    lg.cell(row=rr + 1, column=2,
            value="GREEN = live did better than the model at matched lot; RED = live did worse.")

    wb.save(out_path)


# --------------------------------------------------------------------------- #
# CLI                                                                          #
# --------------------------------------------------------------------------- #

def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        description="Per-signal LIVE vs BACKTEST reconcile -> colour-coded Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter, epilog=__doc__,
    )
    ap.add_argument("--report", required=True, help="MT5 Trade History Report HTML")
    ap.add_argument("--backtest", required=True, help="Per-Entry Detail backtest .xlsx")
    ap.add_argument("--tag", required=True, help="strategy tag, e.g. TSL18")
    ap.add_argument("--lot", type=float, default=0.01, help="live fixed lot (display + usd-per-point base)")
    ap.add_argument("--usd-per-point", type=float, default=None,
                    help="USD per 1.00 price move at the live lot (default lot/0.01 * 1.0)")
    ap.add_argument("--out", default=None, help="output .xlsx (default reports/<TAG>_recon_live_vs_bt.xlsx)")
    ap.add_argument("--require-tick", action="store_true",
                    help="fail if any matched signal's backtest source is not TICK "
                         "(the reconcile is meant to run on a TICK backtest; "
                         "generate the workbook with backtest_hybrid.py --ticks ...)")
    args = ap.parse_args(argv)

    upp = args.usd_per_point if args.usd_per_point is not None else (args.lot / 0.01) * 1.0
    out = args.out or f"reports/{args.tag}_recon_live_vs_bt.xlsx"

    text = _decode_report(args.report)
    live = parse_live(text, args.tag)
    bt = parse_backtest(args.backtest)
    if not live and not bt:
        print(f"[reconcile] no {args.tag} legs in report or backtest - nothing to do")
        return 1
    rows = build_rows(live, bt, upp, None)

    # window label from the data
    days = sorted({d.bt_open.date() for d in rows if d.bt_open} |
                  {d.lv_open.date() for d in rows if d.lv_open})
    window = f"{days[0]}..{days[-1]}" if days else "?"

    # backtest data-source coverage (the reconcile is meant to run on TICK)
    from collections import Counter
    traded = [r for r in rows if r.bt_legs]
    src_ct = Counter(r.bt_src or "?" for r in traded)
    non_tick = [r for r in traded if r.bt_src and r.bt_src != "TICK"]

    write_excel(rows, out, tag=args.tag, lot=args.lot, window=window)

    # terminal echo
    tb = sum(r.bt_pnl for r in rows)
    tl = sum(r.lv_pnl for r in rows)
    print(f"=== {args.tag} reconcile  ({window})  matched lot {args.lot} ===")
    print(f"signals: {len(rows)}   backtest {_money(tb)}   live {_money(tl)}   live-vs-model {_money(tl - tb)}")
    print("backtest data source: " + ", ".join(f"{s}={n}" for s, n in src_ct.most_common()))
    cc = Counter(r.klass for r in rows)
    for k, n in cc.most_common():
        print(f"  {DISCREPANCY[k][0]:<18} x{n}")
    print(f"workbook: {out}")

    if non_tick:
        msg = (f"[reconcile] {len(non_tick)} signal(s) are NOT on TICK data "
               f"(sources: {dict(src_ct)}). The reconcile is meant to run on a TICK "
               f"backtest - regenerate the workbook with backtest_hybrid.py --ticks ... "
               f"covering the live window.")
        if args.require_tick:
            print("ERROR: " + msg)
            return 2
        print("WARNING: " + msg)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
