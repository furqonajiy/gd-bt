"""tools/parity_reconcile.py -- measure the backtest<->live execution gap.

Joins three sources on (signal_key, entry_index) and quantifies, per leg and in
aggregate, *why* live realized P&L differs from the modeled backtest:

  live entries  <- forensic.jsonl   `reconcile_action`  (real fill price/time/lot)
  live exits    <- notifications.jsonl `position_closed` (real close price + P&L)
  modeled side  <- replay_signal(...) over the same signals + chart archive

The backtest's per-entry price levels (entry/exit/status) are lot-independent, so
each modeled leg is re-priced at the *live* lot. That isolates execution-price
quality (slippage + same-bar/catch-up flips) from the separate sizing question
(risk vs fixed), which this tool deliberately does not conflate.

Output: a console summary (operator-facing) and an optional Excel detail workbook.
Read-only: it never touches MT5, the registry, or any trading state.

Index bases differ across sources and are normalised to 0-based internally:
  reconcile_action.entry_index      -> already 0-based
  position_closed.details.entry_index -> 1-based (or "?"); see _norm_close_idx
"""
from __future__ import annotations

import argparse
import glob
import json
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional

_REPO_ROOT = Path(__file__).resolve().parents[1]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from trading.xauusd import (  # noqa: E402
    CONTRACT_SIZE_OZ,
    CsvChartSource,
    DEFAULT_CONFIG,
    StrategyConfig,
    parse_signals_file,
)
from trading.xauusd.strategy.backtest import replay_signal  # noqa: E402


# A live SL/TP fill landing within this many price points of the modeled level
# counts as the same outcome; only the dollar slippage is attributed, not a flip.
_FLAT_EPS = 1e-9
# Outcomes whose engine status means the leg actually opened and closed.
_BT_CLOSED = {
    "SL", "BEP", "LOCK_HALF_TP1", "LOCK_TP1", "LOCK_TP2",
    "TP1", "TP2", "TP3", "TIME_EXIT", "TRAILING_STOP",
}


# --------------------------------------------------------------------------
# pure helpers (no IO -- unit-tested in tests/test_parity_reconcile.py)
# --------------------------------------------------------------------------
def signed_move(side: str, entry: float, exit_price: float) -> float:
    """Price move in the position's P&L direction (positive = profit)."""
    return (exit_price - entry) if side == "BUY" else (entry - exit_price)


def entry_cost_pts(side: str, modeled_entry: float, live_fill: float) -> float:
    """Adverse entry slippage in price points (positive = live filled worse).

    BUY pays more when filled higher; SELL pays more when filled lower.
    """
    return (live_fill - modeled_entry) if side == "BUY" else (modeled_entry - live_fill)


def exit_cost_pts(side: str, modeled_exit: float, live_exit: float) -> float:
    """Adverse exit slippage in price points (positive = live exited worse).

    BUY loses when it sells lower than modeled; SELL loses when it buys higher.
    """
    return (modeled_exit - live_exit) if side == "BUY" else (live_exit - modeled_exit)


def classify_pnl(pnl: Optional[float]) -> str:
    if pnl is None:
        return "OPEN"
    if pnl > _FLAT_EPS:
        return "WIN"
    if pnl < -_FLAT_EPS:
        return "LOSS"
    return "FLAT"


def _norm_close_idx(raw: Any) -> Optional[int]:
    """position_closed carries a 1-based label (or '?'); return 0-based or None."""
    if isinstance(raw, bool):
        return None
    if isinstance(raw, int):
        return raw - 1 if raw >= 1 else None
    s = str(raw).strip()
    return int(s) - 1 if s.isdigit() and int(s) >= 1 else None


# --------------------------------------------------------------------------
# data containers
# --------------------------------------------------------------------------
@dataclass
class LiveEntry:
    fill_price: float
    fill_time: Optional[str]
    lot: float
    planned_price: float
    ticket: int


@dataclass
class LiveExit:
    close_price: float
    profit: float
    reason: str
    ticket: int


@dataclass
class BtLeg:
    entry_price: float
    exit_price: Optional[float]
    status: str
    fill_time: Optional[str]
    exit_time: Optional[str]
    filled: bool
    closed: bool


@dataclass
class CompRow:
    signal_key: str
    idx: int            # 0-based
    side: str
    flag: str           # MATCH | SLIP | FLIP | FILL_ONLY_LIVE | FILL_ONLY_BT | OPEN | NO_DATA
    lot: float
    bt_entry: Optional[float]
    live_entry: Optional[float]
    entry_slip: Optional[float]
    bt_exit: Optional[float]
    live_exit: Optional[float]
    exit_slip: Optional[float]
    bt_status: Optional[str]
    live_reason: Optional[str]
    bt_pnl_at_live_lot: Optional[float]
    live_pnl: Optional[float]
    pnl_delta: Optional[float]      # live - modeled, at the live lot
    note: str = ""


# --------------------------------------------------------------------------
# loaders
# --------------------------------------------------------------------------
def _iter_jsonl(path: Path) -> Iterable[dict]:
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except json.JSONDecodeError:
                continue  # tolerate a torn final line from a live append


def load_forensic_entries(path: Path) -> tuple[dict[tuple[str, int], LiveEntry], dict[int, tuple[str, int]]]:
    """Earliest real fill per (signal_key, idx) plus a ticket->key index for
    resolving '?' close labels back to their entry slot."""
    entries: dict[tuple[str, int], LiveEntry] = {}
    by_ticket: dict[int, tuple[str, int]] = {}
    for ev in _iter_jsonl(path):
        if ev.get("kind") != "reconcile_action":
            continue
        key = ev.get("signal_key")
        idx = ev.get("entry_index")
        if key is None or idx is None:
            continue
        k = (key, int(idx))
        ft = ev.get("fill_time")
        # Keep the first observed fill (reconcile patches PENDING->OPEN once).
        if k in entries and (entries[k].fill_time or "") <= (ft or ""):
            continue
        ticket = int(ev.get("mt5_ticket", 0) or 0)
        entries[k] = LiveEntry(
            fill_price=float(ev.get("fill_price", 0.0) or 0.0),
            fill_time=ft,
            lot=float(ev.get("lot", 0.0) or 0.0),
            planned_price=float(ev.get("planned_price", 0.0) or 0.0),
            ticket=ticket,
        )
        if ticket:
            by_ticket[ticket] = k
    return entries, by_ticket


def load_notifications_exits(
        path: Path, ticket_index: dict[int, tuple[str, int]],
) -> dict[tuple[str, int], LiveExit]:
    """Real broker close per (signal_key, idx) from `position_closed` events."""
    exits: dict[tuple[str, int], LiveExit] = {}
    for ev in _iter_jsonl(path):
        if ev.get("kind") != "position_closed":
            continue
        key = ev.get("signal_key")
        det = ev.get("details") or {}
        ticket = int(det.get("ticket", 0) or 0)
        idx = _norm_close_idx(det.get("entry_index"))
        if idx is None and ticket in ticket_index:
            # '?' label: recover the slot via the forensic fill ticket.
            _, idx = ticket_index[ticket]
        if key is None or idx is None:
            continue
        exits[(key, int(idx))] = LiveExit(
            close_price=float(det.get("close_price", 0.0) or 0.0),
            profit=float(det.get("profit", 0.0) or 0.0),
            reason=str(det.get("reason", "?")),
            ticket=ticket,
        )
    return exits


def build_backtest_legs(
        signals, chart: CsvChartSource, config: StrategyConfig,
        keys_of_interest: set[str],
) -> dict[tuple[str, int], BtLeg]:
    """Replay each live-traded signal standalone and capture per-entry price
    levels. Equity is irrelevant here -- only price/status are used."""
    chart_df = chart.dataframe
    legs: dict[tuple[str, int], BtLeg] = {}
    by_key = {s.signal_key: s for s in signals}
    for key in keys_of_interest:
        sig = by_key.get(key)
        if sig is None:
            continue
        pos = replay_signal(sig, chart_df, config.initial_capital, config)
        for e in pos.entries:
            legs[(key, int(e.entry_index))] = BtLeg(
                entry_price=float(e.entry_price),
                exit_price=(float(e.exit_price) if e.exit_price is not None else None),
                status=e.status,
                fill_time=(str(e.fill_time) if e.fill_time else None),
                exit_time=(str(e.exit_time) if e.exit_time else None),
                filled=e.fill_time is not None,
                closed=e.status in _BT_CLOSED,
            )
    return legs


# --------------------------------------------------------------------------
# row building + aggregation
# --------------------------------------------------------------------------
def build_rows(
        live_entries: dict[tuple[str, int], LiveEntry],
        live_exits: dict[tuple[str, int], LiveExit],
        bt_legs: dict[tuple[str, int], BtLeg],
        contract: float,
) -> list[CompRow]:
    side_of: dict[str, str] = {}
    keys = set(live_entries) | set(live_exits) | {
        k for k in bt_legs if k[0] in {kk[0] for kk in (set(live_entries) | set(live_exits))}
    }
    rows: list[CompRow] = []
    for key in sorted(keys, key=lambda x: (x[0], x[1])):
        sig_key, idx = key
        le = live_entries.get(key)
        lx = live_exits.get(key)
        bt = bt_legs.get(key)
        side = side_of.get(sig_key, "BUY")  # set below from any source carrying it

        lot = le.lot if le else 0.0
        bt_entry = bt.entry_price if bt else None
        live_entry = le.fill_price if le else None
        bt_exit = bt.exit_price if bt else None
        live_exit = lx.close_price if lx else None
        live_pnl = lx.profit if lx else None

        # modeled $ at the live lot isolates price execution from sizing.
        bt_pnl = None
        if bt and bt.filled and bt.exit_price is not None and lot > 0:
            bt_pnl = signed_move(side, bt.entry_price, bt.exit_price) * lot * contract

        entry_slip = (entry_cost_pts(side, bt_entry, live_entry)
                      if (bt_entry is not None and live_entry is not None) else None)
        exit_slip = (exit_cost_pts(side, bt_exit, live_exit)
                     if (bt_exit is not None and live_exit is not None) else None)
        pnl_delta = (live_pnl - bt_pnl) if (live_pnl is not None and bt_pnl is not None) else None

        flag, note = _flag_for(le, lx, bt, live_pnl, bt_pnl)
        if le is not None and abs(le.planned_price - (bt_entry or le.planned_price)) > 0.005 and bt is not None:
            note = (note + " " if note else "") + "planned!=modeled (config mismatch?)"

        rows.append(CompRow(
            signal_key=sig_key, idx=idx, side=side, flag=flag, lot=lot,
            bt_entry=bt_entry, live_entry=live_entry, entry_slip=entry_slip,
            bt_exit=bt_exit, live_exit=live_exit, exit_slip=exit_slip,
            bt_status=(bt.status if bt else None),
            live_reason=(lx.reason if lx else None),
            bt_pnl_at_live_lot=bt_pnl, live_pnl=live_pnl, pnl_delta=pnl_delta, note=note,
        ))
    return rows


def attach_sides(rows: list[CompRow], signals) -> None:
    """Fill side from the parsed signals (the loaders don't all carry it)."""
    side_by_key = {s.signal_key: s.side for s in signals}
    for r in rows:
        r.side = side_by_key.get(r.signal_key, r.side)
        # recompute side-dependent fields now that side is authoritative
        if r.bt_entry is not None and r.live_entry is not None:
            r.entry_slip = entry_cost_pts(r.side, r.bt_entry, r.live_entry)
        if r.bt_exit is not None and r.live_exit is not None:
            r.exit_slip = exit_cost_pts(r.side, r.bt_exit, r.live_exit)
        if (r.bt_entry is not None and r.bt_exit is not None and r.lot > 0):
            r.bt_pnl_at_live_lot = signed_move(r.side, r.bt_entry, r.bt_exit) * r.lot * CONTRACT_SIZE_OZ
            if r.live_pnl is not None:
                r.pnl_delta = r.live_pnl - r.bt_pnl_at_live_lot


def _flag_for(le, lx, bt, live_pnl, bt_pnl) -> tuple[str, str]:
    bt_filled = bool(bt and bt.filled)
    live_filled = le is not None
    if not bt_filled and not live_filled:
        return "NO_DATA", ""
    if live_filled and not bt_filled:
        return "FILL_ONLY_LIVE", "live filled; backtest NO_FILL"
    if bt_filled and not live_filled:
        return "FILL_ONLY_BT", "backtest filled; no live fill"
    if lx is None or bt is None or not bt.closed:
        return "OPEN", "still open in one side"
    if classify_pnl(live_pnl) != classify_pnl(bt_pnl):
        return "FLIP", "outcome sign differs (same-bar/catch-up/timing)"
    return "SLIP", ""


def decompose(rows: list[CompRow]) -> dict[str, float]:
    """Attribute the live-minus-modeled gap into buckets that sum to the gap."""
    agg = defaultdict(float)
    for r in rows:
        if r.flag in ("MATCH", "SLIP") and r.pnl_delta is not None:
            agg["slippage"] += r.pnl_delta
        elif r.flag == "FLIP" and r.pnl_delta is not None:
            agg["label_flip"] += r.pnl_delta
        elif r.flag == "FILL_ONLY_LIVE" and r.live_pnl is not None:
            agg["fill_only_live"] += r.live_pnl
        elif r.flag == "FILL_ONLY_BT" and r.bt_pnl_at_live_lot is not None:
            agg["fill_only_bt"] -= r.bt_pnl_at_live_lot
    agg["live_total"] = sum(r.live_pnl for r in rows if r.live_pnl is not None)
    agg["bt_total_at_live_lot"] = sum(
        r.bt_pnl_at_live_lot for r in rows if r.bt_pnl_at_live_lot is not None
    )
    agg["gap"] = agg["live_total"] - agg["bt_total_at_live_lot"]
    return dict(agg)


def slippage_stats(rows: list[CompRow]) -> dict[str, Any]:
    # Entry slip is comparable for any leg filled on both sides. Exit slip is
    # only meaningful when the outcome agrees -- a FLIP exits at a different
    # level (e.g. TP vs SL), so its "exit slip" is a flip, not slippage.
    e = sorted(r.entry_slip for r in rows if r.entry_slip is not None)
    agreeing = [r for r in rows if r.flag in ("MATCH", "SLIP")]
    x = sorted(r.exit_slip for r in agreeing if r.exit_slip is not None)
    sl_exits = [r.exit_slip for r in agreeing
                if r.exit_slip is not None and (r.bt_status == "SL" or r.live_reason in ("SL", "SO"))]

    def _med(v: list[float]) -> float:
        return 0.0 if not v else (v[len(v) // 2] if len(v) % 2 else (v[len(v) // 2 - 1] + v[len(v) // 2]) / 2)

    return {
        "n_entry": len(e), "entry_mean": (sum(e) / len(e) if e else 0.0), "entry_med": _med(e),
        "n_exit": len(x), "exit_mean": (sum(x) / len(x) if x else 0.0), "exit_med": _med(x),
        "n_sl_exit": len(sl_exits),
        "sl_exit_mean": (sum(sl_exits) / len(sl_exits) if sl_exits else 0.0),
    }


# --------------------------------------------------------------------------
# output
# --------------------------------------------------------------------------
def print_console_summary(rows: list[CompRow], comp: dict[str, float],
                          stats: dict[str, Any], top_n: int = 8) -> None:
    counts = defaultdict(int)
    for r in rows:
        counts[r.flag] += 1
    n = len(rows)

    print("=" * 70)
    print("BACKTEST <-> LIVE PARITY RECONCILIATION")
    print("=" * 70)
    print(f"Legs analysed: {n}")
    for fl in ("SLIP", "FLIP", "FILL_ONLY_LIVE", "FILL_ONLY_BT", "OPEN", "NO_DATA"):
        if counts[fl]:
            print(f"  {fl:<16} {counts[fl]}")
    print()
    print("P&L (modeled re-priced at the live lot, so this is execution-only):")
    print(f"  Live realized total      ${comp.get('live_total', 0.0):+,.2f}")
    print(f"  Backtest @ live lot      ${comp.get('bt_total_at_live_lot', 0.0):+,.2f}")
    print(f"  GAP (live - backtest)    ${comp.get('gap', 0.0):+,.2f}")
    print()
    print("Gap attribution:")
    print(f"  slippage (agreeing legs) ${comp.get('slippage', 0.0):+,.2f}")
    print(f"  label flips              ${comp.get('label_flip', 0.0):+,.2f}")
    print(f"  fills only live had      ${comp.get('fill_only_live', 0.0):+,.2f}")
    print(f"  fills only backtest had  ${comp.get('fill_only_bt', 0.0):+,.2f}")
    print()
    print("Slippage (price points, positive = live worse):")
    print(f"  entry  n={stats['n_entry']:<4} mean={stats['entry_mean']:+.3f} median={stats['entry_med']:+.3f}")
    print(f"  exit   n={stats['n_exit']:<4} mean={stats['exit_mean']:+.3f} median={stats['exit_med']:+.3f}")
    print(f"  SL-exit subset n={stats['n_sl_exit']:<4} mean={stats['sl_exit_mean']:+.3f}  <- the suspected leak")
    print()

    worst = sorted(
        (r for r in rows if r.pnl_delta is not None),
        key=lambda r: r.pnl_delta,
    )[:top_n]
    if worst:
        print(f"Worst {len(worst)} legs by live-minus-modeled $:")
        for r in worst:
            es = f"{r.entry_slip:+.2f}" if r.entry_slip is not None else "--"
            xs = f"{r.exit_slip:+.2f}" if r.exit_slip is not None else "--"
            print(f"  {r.signal_key} #{r.idx + 1} {r.side} {r.flag:<14} "
                  f"d=${r.pnl_delta:+8.2f}  e-slip {es}  x-slip {xs}  "
                  f"({r.bt_status}->{r.live_reason})")
    print("=" * 70)


def write_excel(rows: list[CompRow], comp: dict[str, float],
                stats: dict[str, Any], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    flag_fill = {
        "SLIP": PatternFill("solid", fgColor="FFEB9C"),
        "FLIP": PatternFill("solid", fgColor="FFC7CE"),
        "FILL_ONLY_LIVE": PatternFill("solid", fgColor="DDEBF7"),
        "FILL_ONLY_BT": PatternFill("solid", fgColor="EAEAEA"),
        "OPEN": PatternFill("solid", fgColor="F5F5F5"),
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Legs"
    cols = ["signal_key", "entry#", "side", "flag", "lot",
            "bt_entry", "live_entry", "entry_slip",
            "bt_exit", "live_exit", "exit_slip",
            "bt_status", "live_reason",
            "bt_pnl@live_lot", "live_pnl", "pnl_delta", "note"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append([
            r.signal_key, r.idx + 1, r.side, r.flag, round(r.lot, 2),
            _r(r.bt_entry), _r(r.live_entry), _r(r.entry_slip, 3),
            _r(r.bt_exit), _r(r.live_exit), _r(r.exit_slip, 3),
            r.bt_status, r.live_reason,
            _r(r.bt_pnl_at_live_lot, 2), _r(r.live_pnl, 2), _r(r.pnl_delta, 2), r.note,
                          ])
        fill = flag_fill.get(r.flag)
        if fill:
            for c in range(1, len(cols) + 1):
                ws.cell(row=ws.max_row, column=c).fill = fill
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(28, max(10, width + 2))

    s = wb.create_sheet("Summary")
    for k in ("live_total", "bt_total_at_live_lot", "gap",
              "slippage", "label_flip", "fill_only_live", "fill_only_bt"):
        s.append([k, round(comp.get(k, 0.0), 2)])
    s.append([])
    s.append(["entry_slip_mean", round(stats["entry_mean"], 3)])
    s.append(["exit_slip_mean", round(stats["exit_mean"], 3)])
    s.append(["sl_exit_slip_mean", round(stats["sl_exit_mean"], 3)])
    s.column_dimensions["A"].width = 24
    s.column_dimensions["B"].width = 16

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _r(v: Optional[float], nd: int = 2):
    return None if v is None else round(float(v), nd)


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------
def _expand(patterns: list[str]) -> list[Path]:
    out: list[Path] = []
    for pat in patterns:
        out.extend(sorted(Path(p) for p in glob.glob(pat)))
    return out


def _within(d: datetime, lo: Optional[date], hi: Optional[date]) -> bool:
    if lo and d.date() < lo:
        return False
    if hi and d.date() > hi:
        return False
    return True


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Reconcile live realized P&L against the fixed-edge backtest.")
    p.add_argument("--forensic", required=True, help="forensic.jsonl from auto runs.")
    p.add_argument("--notifications", required=True, help="notifications.jsonl.")
    p.add_argument("--signals", required=True, help="Signal text file.")
    p.add_argument("--charts", required=True, nargs="+", help="M1 chart CSV file(s)/glob(s).")
    p.add_argument("--out", default=None, help="Optional .xlsx detail report path.")
    p.add_argument("--signal", default=None, help="Restrict to one signal_key (e.g. 2026-05-28#03).")
    p.add_argument("--from", dest="date_from", default=None, help="YYYY-MM-DD (GMT+3 signal date).")
    p.add_argument("--to", dest="date_to", default=None, help="YYYY-MM-DD (GMT+3 signal date).")
    p.add_argument("--initial-capital", type=float, default=DEFAULT_CONFIG.initial_capital)
    p.add_argument("--bonus-per-lot", type=float, default=DEFAULT_CONFIG.bonus_per_closed_lot,
                   help="Informational only -- reported, not added to the execution gap.")
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    forensic_path = Path(args.forensic)
    notif_path = Path(args.notifications)
    for label, path in (("forensic", forensic_path), ("notifications", notif_path)):
        if not path.exists():
            print(f"{label} file not found: {path}", file=sys.stderr)
            return 2

    # The modeled side defaults to the DD40 contract; if the live run used other
    # params, the levels won't line up -- the 'planned!=modeled' note flags that.
    config = StrategyConfig(initial_capital=args.initial_capital,
                            bonus_per_closed_lot=args.bonus_per_lot)

    signals = parse_signals_file(Path(args.signals))
    lo = date.fromisoformat(args.date_from) if args.date_from else None
    hi = date.fromisoformat(args.date_to) if args.date_to else None
    signals = [s for s in signals if _within(s.signal_time_chart, lo, hi)
               and (args.signal is None or s.signal_key == args.signal)]
    if not signals:
        print("No signals in range after filtering.", file=sys.stderr)
        return 2

    live_entries, by_ticket = load_forensic_entries(forensic_path)
    live_exits = load_notifications_exits(notif_path, by_ticket)

    wanted = {s.signal_key for s in signals}
    live_entries = {k: v for k, v in live_entries.items() if k[0] in wanted}
    live_exits = {k: v for k, v in live_exits.items() if k[0] in wanted}
    if not live_entries and not live_exits:
        print("No live fills/closes matched the selected signals "
              "(check the date range and that the logs cover this period).",
              file=sys.stderr)
        return 2

    chart = CsvChartSource(_expand(args.charts))
    keys_of_interest = {k[0] for k in (set(live_entries) | set(live_exits))}
    bt_legs = build_backtest_legs(signals, chart, config, keys_of_interest)

    rows = build_rows(live_entries, live_exits, bt_legs, CONTRACT_SIZE_OZ)
    attach_sides(rows, signals)
    comp = decompose(rows)
    stats = slippage_stats(rows)

    print_console_summary(rows, comp, stats)
    if args.bonus_per_lot:
        closed_lots = sum(r.lot for r in rows if r.flag in ("SLIP", "FLIP") and r.lot)
        print(f"NOTE: $3/lot bonus on ~{closed_lots:.2f} closed lots = "
              f"${closed_lots * args.bonus_per_lot:+,.2f} of the backtest's apparent "
              f"edge -- live only earns this if ELEV8 actually credits it.")
    if args.out:
        out_path = Path(args.out)
        write_excel(rows, comp, stats, out_path)
        print(f"Detail workbook written: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())