#!/usr/bin/env python3
"""TSL18 quality-entry + collision-policy research sweep.

Compares the base TSL18 self-scalper feed against **quality-entry** variants
(`--quality-profile`, `--min-quality-score`, `--extreme-entry-mode` on
`tools/generate_scalper_signals.py`) AND **collision-policy** variants
(`--opposite-signal-policy` / `--same-side-overlap-policy`, the TSL18 collision
layer from PR #329, now wired into `tools/backtest_hybrid.py`) under the SAME
TSL18 execution geometry, on TICK where the archive covers the lifecycle. It
scores candidates with the **rebate-aware** objective (`tools/rebate_scoring.py`)
so a dense feed that is green only on the $3/closed-lot rebate — but flat/negative
on pure trading P&L — is never promoted.

It is a bounded research sweep (a small smoke grid + a curated full/validate grid,
not the full aggressive cross-product). Use the tiny `--mode smoke` (or
`--skeleton`, which emits the schema with placeholder rows and runs no backtests)
for a fast structural check.

Collision metrics are REAL whenever a non-baseline collision policy is active:
each candidate passes its `--opposite-signal-policy` / `--same-side-overlap-policy`
into `backtest_hybrid.py`, which now applies the collision layer on BOTH the M1
and the tick path and emits the collision counters in its `--score-json`. A
baseline candidate (`allow_hedge` + `allow_all`) makes zero interventions, so its
collision columns are all zero (parity). A non-baseline candidate whose score-json
carries no collision block is excluded from the ranking with a clear reason.

Modes / windows (``--end-date`` is EXCLUSIVE in backtest_hybrid, so a window's
end is the day AFTER the last day kept):

    smoke         2026-06-27 .. 2026-07-01   (last few June days, pure TICK)
    full_june     2026-06-01 .. 2026-07-01
    validate_top  2026-01-01 .. 2026-07-01   (re-score a prior run's top via --top-json)

Outputs (under ``reports/TSL18_QUALITY_<mode>/``):

    results.csv          one row per candidate, full schema (incl. the real
                         collision-policy columns + metrics)
    top_candidates.json  ranked survivors (gates passed) by score
    summary.md           human-readable leaderboard + the rebate-guard notes

    python tools/sweep_tsl18_quality_entry.py --mode smoke
    python tools/sweep_tsl18_quality_entry.py --mode smoke --skeleton   # no backtests
"""
from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools.rebate_scoring import (  # noqa: E402
    DEFAULT_REBATE_PER_LOT, SCORE_OBJECTIVES, compute_rebate_metrics,
    passes_rebate_guards, score_candidate,
)

# C160 self-scalper feed filter — identical to TSL18 / T818 (mirror; do not tune).
FEED_FILTER = [
    "--session-start", "0", "--session-end", "0", "--signal-tz", "7",
    "--rsi-buy-max", "70", "--rsi-sell-min", "30", "--bb-bandwidth-min", "0.0006",
    "--rr1", "1.2", "--rr2", "2.5", "--rr3", "5",
]

# TSL18 execution geometry — mirror of cli/candidate_TSL18_trailing_tick.txt.
TSL18_GEOMETRY = [
    "--sizing-mode", "risk", "--lot", "0.01", "--risk", "0.01",
    "--minimum-lot", "0.01", "--maximum-lot", "500.0", "--lot-step", "0.01",
    "--bonus-per-closed-lot", "3.0", "--entries", "8", "--entry-ladder", "range_to_sl",
    "--entry-sl-gap", "0.7", "--shared-sl", "false", "--activation-delay", "0",
    "--pending-expiry", "180", "--max-hold", "150", "--sl-multiplier", "1.8",
    "--final-target", "TP3", "--lock-after-tp1", "true", "--lock-after-tp2", "true",
    "--tp1-lock-delay-minutes", "24", "--tp2-lock-delay-minutes", "24",
    "--profit-lock-mode", "tp_levels", "--bep-trigger-distance", "3.0",
    "--tp1-lock-fraction", "0.75", "--tp2-lock-target", "TP1", "--tp3-lock-target", "TP2",
    "--runner-after-tp3", "false", "--trailing-open-distance", "0.5",
    "--trailing-close-distance", "0.5", "--trailing-close-after-stage", "2",
]
ERA_SLIP = ["--lock-tp1-exit-slippage", "2.0", "--lock-tp2-exit-slippage", "1.0"]  # R4

WINDOWS = {
    "smoke": ("2026-06-27", "2026-07-01"),
    "june": ("2026-06-01", "2026-07-01"),
    "jan_jun": ("2026-01-01", "2026-07-01"),
}
MODE_WINDOW = {"smoke": "smoke", "full_june": "june", "validate_top": "jan_jun"}

# Results schema. The collision columns are REAL: ``opposite_signal_policy`` /
# ``same_side_overlap_policy`` record the candidate's policy, and the metrics are
# read from backtest_hybrid's --score-json (zeros for a baseline candidate, which
# makes no interventions; populated for a non-baseline policy).
RESULTS_COLUMNS = [
    "label", "quality_profile", "min_quality_score", "extreme_entry_mode",
    "window", "window_start", "window_end",
    "pure_trading_pnl", "rebate_pnl", "net_pnl", "closed_lots",
    "pure_pnl_per_lot", "net_pnl_per_lot", "rebate_share_of_profit",
    "score_objective", "score",
    "passes_rebate_guards", "rebate_guard_reason",
    "max_drawdown_pct", "win_rate_pct", "signals_included",
    "tick_signals", "m1_signals", "data_source_mixed",
    "partial_tick_lifecycle_excluded", "open_or_pending_left", "included_in_ranking",
    # --- real collision-policy columns + metrics (PR #329, wired in #331-followup) ---
    "opposite_signal_policy", "same_side_overlap_policy", "collision_policy_pnl",
    "opposite_collisions_total", "opposite_collisions_allowed",
    "opposite_collisions_rejected", "opposite_collisions_flipped",
    "opposite_collisions_profit_bank_rearmed",
    "same_side_clusters_total", "same_side_clusters_accepted",
    "same_side_clusters_rejected", "same_side_clusters_downsized",
    "max_same_side_cluster_risk", "max_opposite_exposure",
]

# The collision counters read from backtest_hybrid's --score-json (order matches
# the CollisionPolicy.summary() keys; collision_policy_pnl is handled separately).
COLLISION_METRIC_COLUMNS = [
    "collision_policy_pnl",
    "opposite_collisions_total", "opposite_collisions_allowed",
    "opposite_collisions_rejected", "opposite_collisions_flipped",
    "opposite_collisions_profit_bank_rearmed",
    "same_side_clusters_total", "same_side_clusters_accepted",
    "same_side_clusters_rejected", "same_side_clusters_downsized",
    "max_same_side_cluster_risk", "max_opposite_exposure",
]

# Baseline (parity) collision config: the layer makes zero interventions here.
COLLISION_DEFAULTS = {
    "opposite_signal_policy": "allow_hedge",
    "same_side_overlap_policy": "allow_all",
    "same_side_cluster_window_minutes": 30,
    "same_side_cluster_entry_gap": 5.0,
    "max_cluster_risk_multiple": 1.0,
    "opposite_profit_threshold_r": 0.5,
    "hedge_lot_fraction": 0.5,
}

QUALITY_PROFILES = ["off", "trend_only", "reversal_extreme", "hybrid_quality",
                    "high_frequency_quality"]
EXTREME_MODES = ["support_demand", "supply_resistance", "both"]


def _candidate(label: str, *, quality_profile: str = "off",
               min_quality_score: float = 0.0, extreme_entry_mode: str = "off",
               **collision) -> dict:
    """Build a candidate dict: quality-layer keys + collision keys (baseline
    defaults filled, then overridden by ``collision``). One source of truth so a
    grid entry, the validate_top reconstruction, and ``_blank_row`` all agree."""
    cand = {"label": label, "quality_profile": quality_profile,
            "min_quality_score": min_quality_score,
            "extreme_entry_mode": extreme_entry_mode}
    cand.update(COLLISION_DEFAULTS)
    bad = set(collision) - set(COLLISION_DEFAULTS)
    if bad:
        raise ValueError(f"unknown collision keys for {label}: {sorted(bad)}")
    cand.update(collision)
    return cand


def _collision_is_baseline(cand: dict) -> bool:
    """True when the candidate runs the baseline (parity) collision policy, so its
    collision columns are all zero and no collision block is expected in score-json."""
    return (cand.get("opposite_signal_policy", "allow_hedge") == "allow_hedge"
            and cand.get("same_side_overlap_policy", "allow_all") == "allow_all")


def build_candidate_grid(mode: str) -> list[dict]:
    """Return the candidate list for a mode. ``base`` (no quality layer, baseline
    collision) is always first so the report can read every variant against it.

    The grid spans three families: quality-entry (feed-layer), collision-only
    (execution-layer, PR #329), and combined. It is a bounded curated grid, not
    the full aggressive cross-product."""
    if mode == "smoke":
        return [
            _candidate("base"),
            _candidate("hybrid_quality", quality_profile="hybrid_quality"),
            _candidate("profit_bank_rearm",
                       opposite_signal_policy="profit_bank_rearm",
                       opposite_profit_threshold_r=0.5),
            _candidate("hybrid_quality_profit_bank_scale_better",
                       quality_profile="hybrid_quality",
                       opposite_signal_policy="profit_bank_rearm",
                       opposite_profit_threshold_r=0.5,
                       same_side_overlap_policy="scale_in_better_entry_only",
                       same_side_cluster_entry_gap=5.0, max_cluster_risk_multiple=1.0),
        ]
    # full_june / validate_top: a curated research grid (quality + collision +
    # combined), still NOT the full aggressive cross-product.
    grid = [_candidate("base")]
    # --- quality-entry family ---
    grid += [
        _candidate("trend_only", quality_profile="trend_only"),
        _candidate("reversal_extreme", quality_profile="reversal_extreme"),
        _candidate("hybrid_quality", quality_profile="hybrid_quality"),
        _candidate("high_frequency_quality", quality_profile="high_frequency_quality"),
        _candidate("extreme_support_demand", extreme_entry_mode="support_demand"),
        _candidate("extreme_supply_resistance", extreme_entry_mode="supply_resistance"),
        _candidate("extreme_both", extreme_entry_mode="both"),
    ]
    # --- collision-only family (execution layer; feed unchanged) ---
    grid += [
        _candidate("reject_opposite", opposite_signal_policy="reject_opposite"),
        _candidate("profit_bank_rearm", opposite_signal_policy="profit_bank_rearm",
                   opposite_profit_threshold_r=0.5),
        _candidate("same_reject_overlap", same_side_overlap_policy="reject_overlap"),
        _candidate("same_scale_better_only",
                   same_side_overlap_policy="scale_in_better_entry_only",
                   same_side_cluster_entry_gap=5.0, max_cluster_risk_multiple=1.0),
        _candidate("same_scale_fixed_risk",
                   same_side_overlap_policy="scale_in_fixed_risk",
                   max_cluster_risk_multiple=1.0),
    ]
    # --- combined (quality feed + collision execution) family ---
    grid += [
        _candidate("hybrid_quality_profit_bank", quality_profile="hybrid_quality",
                   opposite_signal_policy="profit_bank_rearm",
                   opposite_profit_threshold_r=0.5),
        _candidate("hybrid_quality_scale_better", quality_profile="hybrid_quality",
                   same_side_overlap_policy="scale_in_better_entry_only",
                   same_side_cluster_entry_gap=5.0, max_cluster_risk_multiple=1.0),
        _candidate("hybrid_quality_profit_bank_scale_better",
                   quality_profile="hybrid_quality",
                   opposite_signal_policy="profit_bank_rearm",
                   opposite_profit_threshold_r=0.5,
                   same_side_overlap_policy="scale_in_better_entry_only",
                   same_side_cluster_entry_gap=5.0, max_cluster_risk_multiple=1.0),
        _candidate("hybrid_quality_extreme_both_profit_bank_scale_better",
                   quality_profile="hybrid_quality", extreme_entry_mode="both",
                   opposite_signal_policy="profit_bank_rearm",
                   opposite_profit_threshold_r=0.5,
                   same_side_overlap_policy="scale_in_better_entry_only",
                   same_side_cluster_entry_gap=5.0, max_cluster_risk_multiple=1.0),
    ]
    return grid


def _quality_flags(cand: dict) -> list[str]:
    """Render a candidate's quality-layer generator flags (empty for base)."""
    flags: list[str] = []
    if cand["quality_profile"] != "off":
        flags += ["--quality-profile", str(cand["quality_profile"])]
        if float(cand.get("min_quality_score", 0.0)) > 0.0:
            flags += ["--min-quality-score", str(cand["min_quality_score"])]
    if cand["extreme_entry_mode"] != "off":
        flags += ["--extreme-entry-mode", str(cand["extreme_entry_mode"])]
    return flags


def _collision_flags(cand: dict) -> list[str]:
    """Render a candidate's collision-policy flags for backtest_hybrid (empty for a
    baseline candidate, so a baseline run's command stays byte-identical). Only the
    flags relevant to the active policy are emitted, but explicitly enough that the
    run reproduces (the cluster sub-knobs that govern the chosen same-side mode and
    the profit/hedge knobs for the chosen opposite mode)."""
    flags: list[str] = []
    opp = cand.get("opposite_signal_policy", "allow_hedge")
    if opp != "allow_hedge":
        flags += ["--opposite-signal-policy", str(opp)]
        if opp == "profit_bank_rearm":
            flags += ["--opposite-profit-threshold-r",
                      str(cand.get("opposite_profit_threshold_r", 0.5))]
        elif opp == "reduce_then_hedge":
            flags += ["--hedge-lot-fraction", str(cand.get("hedge_lot_fraction", 0.5))]
    same = cand.get("same_side_overlap_policy", "allow_all")
    if same != "allow_all":
        flags += ["--same-side-overlap-policy", str(same)]
        flags += ["--same-side-cluster-window-minutes",
                  str(cand.get("same_side_cluster_window_minutes", 30))]
        if same in ("scale_in_better_entry_only", "scale_in_fixed_risk"):
            flags += ["--max-cluster-risk-multiple",
                      str(cand.get("max_cluster_risk_multiple", 1.0))]
        if same == "scale_in_better_entry_only":
            flags += ["--same-side-cluster-entry-gap",
                      str(cand.get("same_side_cluster_entry_gap", 5.0))]
    return flags


def _run(cmd: list[str]) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True)


def _gen_feed(out_txt: Path, charts: list[str], start: str, cand: dict) -> None:
    cmd = [sys.executable, "tools/generate_scalper_signals.py",
           "--charts", *charts, "--output", str(out_txt),
           "--start", start, "--progress-interval-seconds", "0",
           *FEED_FILTER, *_quality_flags(cand)]
    r = _run(cmd)
    if r.returncode != 0:
        raise SystemExit(f"feed generation failed for {cand['label']}:\n{r.stderr[-1500:]}")


def _run_backtest(feed: Path, charts: list[str], ticks: list[str],
                  start: str, end: str, out_dir: Path, score_json: Path,
                  collision_flags: list[str] | None = None) -> Path:
    cmd = [sys.executable, "tools/backtest_hybrid.py", "--signals", str(feed),
           "--charts", *charts, "--ticks", *ticks,
           "--sync-ticks", "false", "--sync-charts", "false",
           "--max-drawdown-limit-pct", "9999", "--progress-interval-seconds", "0",
           *TSL18_GEOMETRY, *ERA_SLIP, *(collision_flags or []),
           "--output-dir", str(out_dir), "--initial-capital", "50000",
           "--score-json", str(score_json),
           "--start-date", start, "--end-date", end]
    r = _run(cmd)
    if r.returncode != 0:
        raise SystemExit(f"backtest failed for {feed.name}:\n{r.stderr[-1500:]}")
    xlsx = out_dir.with_suffix(".xlsx")
    if not xlsx.exists():
        raise SystemExit(f"no workbook at {xlsx}")
    return xlsx


def _read_summary_pnl(xlsx: Path) -> tuple[float, float]:
    """(pure_trading_pnl, rebate_pnl) from the workbook Summary key/value cells."""
    from openpyxl import load_workbook

    ws = load_workbook(xlsx, data_only=True)["Summary"]
    pure = rebate = 0.0
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if not label:
            continue
        val = ws.cell(row=r, column=2).value
        text = str(label).strip()
        if text == "Trading P&L":
            pure = _money(val)
        elif text == "Closed-lot bonus":
            rebate = _money(val)
    return pure, rebate


def _count_open_pending(xlsx: Path) -> int:
    """Count entries still OPEN/PENDING at window end (incomplete P&L)."""
    from openpyxl import load_workbook

    ws = load_workbook(xlsx, data_only=True)["Per-Entry Detail"]
    hdr = {str(ws.cell(row=2, column=c).value).strip(): c
           for c in range(1, ws.max_column + 1)}
    status_col = hdr.get("Status")
    if not status_col:
        return 0
    n = 0
    for r in range(3, ws.max_row + 1):
        status = ws.cell(row=r, column=status_col).value
        if status and str(status).upper() in ("OPEN", "PENDING"):
            n += 1
    return n


def _money(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _blank_row(cand: dict, window: str, span: tuple[str, str], objective: str) -> dict:
    """A schema-complete placeholder row (skeleton / no-backtest mode). The
    collision policy columns reflect the candidate's policy; the metric columns are
    blank until a real run populates them. The full collision CONFIG (incl. the
    cluster sub-knobs that are not CSV columns) is stashed as extra keys so
    ``top_candidates.json`` preserves them for ``validate_top`` (write_results_csv
    only emits RESULTS_COLUMNS, so the extras are JSON-only)."""
    row = {
        "label": cand["label"], "quality_profile": cand["quality_profile"],
        "min_quality_score": cand["min_quality_score"],
        "extreme_entry_mode": cand["extreme_entry_mode"],
        "window": window, "window_start": span[0], "window_end": span[1],
        "pure_trading_pnl": "", "rebate_pnl": "", "net_pnl": "", "closed_lots": "",
        "pure_pnl_per_lot": "", "net_pnl_per_lot": "", "rebate_share_of_profit": "",
        "score_objective": objective, "score": "",
        "passes_rebate_guards": "", "rebate_guard_reason": "skeleton",
        "max_drawdown_pct": "", "win_rate_pct": "", "signals_included": "",
        "tick_signals": "", "m1_signals": "", "data_source_mixed": "",
        "partial_tick_lifecycle_excluded": "", "open_or_pending_left": "",
        "included_in_ranking": "",
        "opposite_signal_policy": cand.get("opposite_signal_policy", "allow_hedge"),
        "same_side_overlap_policy": cand.get("same_side_overlap_policy", "allow_all"),
    }
    for col in COLLISION_METRIC_COLUMNS:
        row[col] = ""
    # JSON-only collision sub-config (not CSV columns) so validate_top reproduces
    # the exact candidate.
    for k in COLLISION_DEFAULTS:
        if k not in row:
            row[k] = cand.get(k, COLLISION_DEFAULTS[k])
    return row


def evaluate_candidate(cand: dict, window: str, span: tuple[str, str], *,
                       charts: list[str], ticks: list[str], gen_start: str,
                       out_root: Path, args: argparse.Namespace) -> dict:
    """Run one candidate (or emit a placeholder row in --skeleton mode)."""
    if args.skeleton:
        return _blank_row(cand, window, span, args.score_objective)

    start, end = span
    feed = out_root / f"feed_{cand['label']}.txt"
    score_json = out_root / f"score_{cand['label']}.json"
    _gen_feed(feed, charts, gen_start, cand)
    xlsx = _run_backtest(feed, charts, ticks, start, end,
                         out_root / f"bt_{cand['label']}", score_json,
                         collision_flags=_collision_flags(cand))

    score = json.loads(score_json.read_text(encoding="utf-8"))
    pure, rebate = _read_summary_pnl(xlsx)
    closed_lots = rebate / args.rebate_per_lot if args.rebate_per_lot > 0 else 0.0
    m = compute_rebate_metrics(pure, closed_lots, rebate_per_lot=args.rebate_per_lot,
                               rebate_pnl=rebate)
    dd = float(score.get("max_drawdown_pct", 0.0))
    guard_ok, guard_reason = passes_rebate_guards(
        m, min_pure_trading_pnl=args.min_pure_trading_pnl,
        max_rebate_share_of_profit=args.max_rebate_share_of_profit)
    sc = score_candidate(m, args.score_objective, max_drawdown_pct=dd,
                         min_pure_trading_pnl=args.min_pure_trading_pnl,
                         max_rebate_share_of_profit=args.max_rebate_share_of_profit)
    tick_n = int(score.get("tick_signals", 0))
    m1_n = int(score.get("m1_signals", 0))
    mixed = tick_n > 0 and m1_n > 0
    partial_excluded = bool(args.require_full_tick_lifecycle and mixed)
    open_left = _count_open_pending(xlsx) > 0

    # Collision metrics: baseline candidates make zero interventions, so the
    # score-json carries no collision block -> populate zeros. A NON-baseline
    # candidate whose score-json has no collision block is broken (the policy did
    # not run): exclude it from ranking with a clear reason rather than score a
    # signal that silently ignored its collision flags.
    is_baseline = _collision_is_baseline(cand)
    has_block = "collision_policy_pnl" in score
    collision_metrics: dict = {}
    collision_missing = (not is_baseline) and (not has_block)
    if collision_missing:
        for col in COLLISION_METRIC_COLUMNS:
            collision_metrics[col] = ""
    else:
        for col in COLLISION_METRIC_COLUMNS:
            collision_metrics[col] = score.get(col, 0)

    included = (guard_ok and not partial_excluded and not collision_missing
                and not (args.exclude_open_or_pending and open_left))
    if collision_missing:
        guard_reason = "collision metrics missing for non-baseline policy"

    row = _blank_row(cand, window, span, args.score_objective)
    row.update({
        "pure_trading_pnl": m.pure_trading_pnl, "rebate_pnl": m.rebate_pnl,
        "net_pnl": m.net_pnl, "closed_lots": m.closed_lots,
        "pure_pnl_per_lot": m.pure_pnl_per_lot, "net_pnl_per_lot": m.net_pnl_per_lot,
        "rebate_share_of_profit": m.rebate_share_of_profit,
        "score": round(sc, 2), "passes_rebate_guards": guard_ok,
        "rebate_guard_reason": guard_reason,
        "max_drawdown_pct": dd, "win_rate_pct": float(score.get("win_rate_pct", 0.0)),
        "signals_included": int(score.get("signals_included", 0)),
        "tick_signals": tick_n, "m1_signals": m1_n, "data_source_mixed": mixed,
        "partial_tick_lifecycle_excluded": partial_excluded,
        "open_or_pending_left": open_left, "included_in_ranking": included,
    })
    row.update(collision_metrics)
    return row


def write_results_csv(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=RESULTS_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in RESULTS_COLUMNS})


def write_top_candidates_json(rows: list[dict], path: Path) -> list[dict]:
    """Rank ranking-included rows by score desc; write + return them."""
    ranked = [r for r in rows if r.get("included_in_ranking") is True]
    ranked.sort(key=lambda r: (r.get("score") if isinstance(r.get("score"), (int, float))
                               else float("-inf")), reverse=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(ranked, indent=2), encoding="utf-8")
    return ranked


def write_summary_md(rows: list[dict], ranked: list[dict], path: Path,
                     mode: str, span: tuple[str, str], objective: str) -> None:
    lines = [
        f"# TSL18 quality-entry + collision sweep — {mode} ({span[0]}..{span[1]})",
        "",
        f"Score objective: **{objective}**. Base TSL18 feed vs quality-entry AND "
        "collision-policy variants, same TSL18 geometry, TICK where covered. "
        "**Rebate-aware**: a candidate green only on the $3/closed-lot rebate (bad pure "
        "P&L) is guarded out, never promoted. **Collision policies are real** here — the "
        "opposite/same-side metrics come from backtest_hybrid's collision layer (zero for "
        "a baseline candidate, populated for a non-baseline policy).",
        "",
        "| label | profile | extreme | opp policy | same policy | pure $ | net $ | "
        "score | guards | DD% | coll $ | opp tot/rej/bank | same tot/rej/dsz | ranked |",
        "|---|---|---|---|---|--:|--:|--:|:--|--:|--:|:--:|:--:|:--:|",
    ]
    for r in rows:
        opp = (f"{r.get('opposite_collisions_total', '')}/"
               f"{r.get('opposite_collisions_rejected', '')}/"
               f"{r.get('opposite_collisions_profit_bank_rearmed', '')}")
        same = (f"{r.get('same_side_clusters_total', '')}/"
                f"{r.get('same_side_clusters_rejected', '')}/"
                f"{r.get('same_side_clusters_downsized', '')}")
        lines.append(
            f"| {r['label']} | {r['quality_profile']} | {r['extreme_entry_mode']} | "
            f"{r.get('opposite_signal_policy', '')} | {r.get('same_side_overlap_policy', '')} | "
            f"{r['pure_trading_pnl']} | {r['net_pnl']} | {r['score']} | "
            f"{r['rebate_guard_reason']} | {r['max_drawdown_pct']} | "
            f"{r.get('collision_policy_pnl', '')} | {opp} | {same} | "
            f"{r['included_in_ranking']} |")
    lines += [
        "",
        "Collision columns: **opp tot/rej/bank** = opposite collisions total / rejected / "
        "profit-bank-rearmed; **same tot/rej/dsz** = same-side clusters total / rejected / "
        "downsized; **coll $** = collision_policy_pnl (banked old-side delta).",
        "",
        "## Ranked survivors (gates passed)",
        "",
    ]
    if ranked:
        for i, r in enumerate(ranked, start=1):
            lines.append(f"{i}. **{r['label']}** — score {r['score']} "
                         f"(net ${r['net_pnl']}, pure ${r['pure_trading_pnl']}, "
                         f"opp={r.get('opposite_signal_policy', '')}, "
                         f"same={r.get('same_side_overlap_policy', '')})")
    else:
        lines.append("_No survivors (skeleton run, or all candidates failed the gates)._")
    lines += [
        "",
        "Gates: **rebate guards** (pure-P&L floor + max rebate share), "
        "**partial-tick-lifecycle exclusion** (mixed TICK/M1 windows when "
        "`--require-full-tick-lifecycle`), **open/pending-left** "
        "(`--exclude-open-or-pending`), and **collision-metrics-present** (a "
        "non-baseline policy whose run emitted no collision block is excluded).",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    # `full` is an automation alias for `full_june` (normalized in main()).
    ap.add_argument("--mode", choices=list(MODE_WINDOW) + ["full"], default="smoke",
                    help="smoke | full_june | validate_top ('full' = automation alias for full_june).")
    ap.add_argument("--window", choices=list(WINDOWS), default=None,
                    help="override the mode's default window.")
    ap.add_argument("--start-date", default=None,
                    help="override the scoring-window START (YYYY-MM-DD); default = the mode/window start.")
    ap.add_argument("--end-date", default=None,
                    help="override the scoring-window END (YYYY-MM-DD, EXCLUSIVE); default = the mode/window end.")
    ap.add_argument("--skeleton", action="store_true",
                    help="emit the results schema with placeholder rows; run NO backtests.")
    ap.add_argument("--charts", nargs="+",
                    default=["data/XAUUSD_M1_202606_ELEV8.csv"])
    ap.add_argument("--ticks", nargs="+", default=["data/ticks/XAUUSD_TICK_*_ELEV8.csv"])
    ap.add_argument("--gen-start", default="2026-06-01",
                    help="feed-generation start (indicator warmup precedes the window).")
    # --input-candidates is an automation alias for --top-json.
    ap.add_argument("--top-json", "--input-candidates", dest="top_json", default=None,
                    help="validate_top: a prior run's top_candidates.json to re-score "
                         "(--input-candidates is an automation alias).")
    # --output-dir is an automation alias for --out-root; the script ALWAYS writes
    # <out-root>/TSL18_QUALITY_<mode>/, so --output-dir is the ROOT, not the literal
    # final directory (kept simple + explicit to avoid confusing nested paths).
    ap.add_argument("--out-root", "--output-dir", dest="out_root", default="reports",
                    help="output ROOT; the script writes <out-root>/TSL18_QUALITY_<mode>/. "
                         "--output-dir is an automation alias (still creates the subfolder).")
    # rebate-aware scoring knobs (see tools/rebate_scoring.py)
    ap.add_argument("--rebate-per-lot", type=float, default=DEFAULT_REBATE_PER_LOT)
    ap.add_argument("--min-pure-trading-pnl", type=float, default=0.0,
                    help="reject candidates whose pure trading P&L < X.")
    ap.add_argument("--max-rebate-share-of-profit", type=float, default=0.50,
                    help="reject candidates whose net profit is >X rebate.")
    # --rank-objective is an automation alias for --score-objective.
    ap.add_argument("--score-objective", "--rank-objective", dest="score_objective",
                    choices=list(SCORE_OBJECTIVES), default="edge_plus_rebate_guarded",
                    help="rebate-aware ranking objective (--rank-objective is an automation alias).")
    # exclusion gates (each carries an automation-friendly alias).
    ap.add_argument("--require-full-tick-lifecycle", "--require-full-lifecycle-ticks",
                    dest="require_full_tick_lifecycle", action="store_true",
                    help="exclude candidates whose window mixes TICK and M1 (partial coverage). "
                         "(--require-full-lifecycle-ticks is an automation alias.)")
    ap.add_argument("--exclude-open-or-pending", "--fail-on-open-or-pending",
                    dest="exclude_open_or_pending", action="store_true",
                    help="exclude candidates that left positions OPEN/PENDING at window end. "
                         "(--fail-on-open-or-pending is an automation alias.)")
    return ap


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    if args.mode == "full":            # automation alias for full_june
        args.mode = "full_june"
    window = args.window or MODE_WINDOW[args.mode]
    span = WINDOWS[window]
    if args.start_date or args.end_date:   # explicit window overrides (--end-date EXCLUSIVE)
        span = (args.start_date or span[0], args.end_date or span[1])
    out_root = Path(args.out_root) / f"TSL18_QUALITY_{args.mode}"
    out_root.mkdir(parents=True, exist_ok=True)

    if args.mode == "validate_top" and args.top_json:
        prior = json.loads(Path(args.top_json).read_text(encoding="utf-8"))
        # Preserve BOTH the quality-layer AND the collision config from each top
        # record (the collision sub-knobs are stored in top_candidates.json even
        # though they are not CSV columns), so validate_top re-scores the EXACT
        # candidate that won full_june.
        candidates = [
            _candidate(c["label"], quality_profile=c.get("quality_profile", "off"),
                       min_quality_score=c.get("min_quality_score", 0.0),
                       extreme_entry_mode=c.get("extreme_entry_mode", "off"),
                       **{k: c[k] for k in COLLISION_DEFAULTS if k in c})
            for c in prior
        ]
    else:
        candidates = build_candidate_grid(args.mode)

    rows: list[dict] = []
    for cand in candidates:
        print(f"[quality-sweep] {cand['label']}: profile={cand['quality_profile']} "
              f"score>={cand['min_quality_score']} extreme={cand['extreme_entry_mode']} "
              f"opp={cand.get('opposite_signal_policy', 'allow_hedge')} "
              f"same={cand.get('same_side_overlap_policy', 'allow_all')}"
              f"{' [skeleton]' if args.skeleton else ''}", flush=True)
        rows.append(evaluate_candidate(cand, window, span, charts=args.charts,
                                       ticks=args.ticks, gen_start=args.gen_start,
                                       out_root=out_root, args=args))

    results_csv = out_root / "results.csv"
    top_json = out_root / "top_candidates.json"
    summary = out_root / "summary.md"
    write_results_csv(rows, results_csv)
    ranked = write_top_candidates_json(rows, top_json)
    write_summary_md(rows, ranked, summary, args.mode, span, args.score_objective)
    print(f"[quality-sweep] results: {results_csv}")
    print(f"[quality-sweep] top:     {top_json} ({len(ranked)} survivors)")
    print(f"[quality-sweep] summary: {summary}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
