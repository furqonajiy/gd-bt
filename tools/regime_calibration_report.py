#!/usr/bin/env python3
"""Calibrate XAUUSD regimes from the available 2021+ M1 chart history.

This report is evidence for the regime contract. It measures each month from
chart data, learns four volatility clusters from monthly M15 ATR, compares
those learned labels with the live router, and compares both with the calendar
windows used by the sweep workflows.

Example:

    python tools/regime_calibration_report.py \
      --charts data/XAUUSD_M1_*_ELEV8.csv \
      --out reports/regime_calibration.md \
      --csv-out reports/regime_calibration.csv \
      --json-out reports/regime_calibration.json
"""
from __future__ import annotations

import argparse
import csv
import glob
import json
import math
import sys
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from xauusd_trading.strategy.regime import (  # noqa: E402
    BULL_TREND_MIN,
    VOL_TIER_LOW_MAX,
    VOL_TIER_MID_MAX,
    detect_regime,
)

REGIME_BY_CLUSTER = ["R1quiet", "R2bull", "R3strong", "R4parab"]


@dataclass(frozen=True)
class MonthlyRegimeRow:
    month: str
    bars: int
    first_close: float
    last_close: float
    high: float
    low: float
    m15_atr_mean: float
    m15_atr_median: float
    m15_atr_p90: float
    monthly_return_pct: float
    range_pct: float
    trend_atr_multiple: float
    learned_regime: str
    learned_cluster: int
    learned_center_atr: float
    router_regime: str
    calendar_regime: str
    router_matches_learned: bool
    calendar_matches_learned: bool
    near_learned_boundary: bool
    nearest_boundary_pct: float | None


def _expand_chart_paths(patterns: Iterable[str]) -> list[Path]:
    out: list[Path] = []
    for pattern in patterns:
        if any(ch in pattern for ch in "*?["):
            matches = sorted(glob.glob(pattern))
            if not matches:
                raise SystemExit(f"No files match pattern: {pattern}")
            out.extend(Path(match) for match in matches)
        else:
            path = Path(pattern)
            if not path.exists():
                raise SystemExit(f"Chart file not found: {pattern}")
            out.append(path)
    if not out:
        raise SystemExit("No chart files provided")
    return out


def _read_chart(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, sep="\t")
    df.columns = [str(col).strip("<>").lower() for col in df.columns]
    required = {"date", "time", "open", "high", "low", "close"}
    missing = required - set(df.columns)
    if missing:
        raise SystemExit(f"{path} missing columns: {sorted(missing)}")
    df["dt"] = pd.to_datetime(
        df["date"].astype(str) + " " + df["time"].astype(str),
        format="%Y.%m.%d %H:%M:%S",
        errors="coerce",
    )
    for col in ("open", "high", "low", "close"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df.dropna(subset=["dt", "open", "high", "low", "close"]).set_index("dt")[
        ["open", "high", "low", "close"]
    ]


def load_charts(paths: Iterable[Path], *, start_date: str) -> pd.DataFrame:
    frames = [_read_chart(path) for path in paths]
    if not frames:
        raise SystemExit("No chart data loaded")
    m1 = pd.concat(frames).sort_index()
    m1 = m1[~m1.index.duplicated(keep="last")]
    if start_date:
        m1 = m1[m1.index >= pd.Timestamp(start_date)]
    if m1.empty:
        raise SystemExit("No chart rows remain after the start-date filter")
    return m1


def _m15(m1: pd.DataFrame) -> pd.DataFrame:
    return (
        m1.resample("15min")
        .agg(open=("open", "first"), high=("high", "max"), low=("low", "min"), close=("close", "last"))
        .dropna()
    )


def _true_range(m15: pd.DataFrame) -> pd.Series:
    prev_close = m15["close"].shift()
    return pd.concat(
        [
            m15["high"] - m15["low"],
            (m15["high"] - prev_close).abs(),
            (m15["low"] - prev_close).abs(),
        ],
        axis=1,
    ).max(axis=1)


def monthly_features(m1: pd.DataFrame, *, atr_period: int = 14) -> list[dict]:
    m15 = _m15(m1)
    tr = _true_range(m15)
    m15 = m15.copy()
    m15["atr"] = tr.rolling(atr_period, min_periods=atr_period).mean()
    m15["tr"] = tr
    rows: list[dict] = []
    for period, month_m1 in m1.groupby(m1.index.to_period("M")):
        month = str(period)
        month_m15 = m15[m15.index.to_period("M") == period]
        if month_m1.empty or month_m15.empty:
            continue
        atr = month_m15["atr"].dropna()
        if atr.empty:
            atr = month_m15["tr"].dropna()
        if atr.empty:
            continue
        first_close = float(month_m1["close"].iloc[0])
        last_close = float(month_m1["close"].iloc[-1])
        high = float(month_m1["high"].max())
        low = float(month_m1["low"].min())
        atr_mean = float(atr.mean())
        move = last_close - first_close
        rows.append(
            {
                "month": month,
                "bars": int(len(month_m1)),
                "first_close": first_close,
                "last_close": last_close,
                "high": high,
                "low": low,
                "m15_atr_mean": atr_mean,
                "m15_atr_median": float(atr.median()),
                "m15_atr_p90": float(atr.quantile(0.90)),
                "monthly_return_pct": (move / first_close * 100.0) if first_close else 0.0,
                "range_pct": ((high - low) / first_close * 100.0) if first_close else 0.0,
                "trend_atr_multiple": (move / atr_mean) if atr_mean else 0.0,
            }
        )
    return rows


def _quantile(sorted_values: list[float], q: float) -> float:
    if not sorted_values:
        raise ValueError("cannot take a quantile of an empty list")
    if len(sorted_values) == 1:
        return sorted_values[0]
    pos = max(0.0, min(1.0, q)) * (len(sorted_values) - 1)
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return sorted_values[lo]
    frac = pos - lo
    return sorted_values[lo] * (1.0 - frac) + sorted_values[hi] * frac


def kmeans_1d(values: list[float], *, k: int = 4, iterations: int = 50) -> tuple[list[float], list[int]]:
    """Deterministic 1D k-means. Returns sorted centers and assignment indexes."""
    clean = [float(v) for v in values if v is not None and math.isfinite(float(v))]
    if not clean:
        raise ValueError("no finite values to cluster")
    k = max(1, min(k, len(clean)))
    ordered = sorted(clean)
    centers = [_quantile(ordered, (i + 0.5) / k) for i in range(k)]
    for _ in range(iterations):
        buckets: list[list[float]] = [[] for _ in range(k)]
        for value in clean:
            idx = min(range(k), key=lambda i: abs(value - centers[i]))
            buckets[idx].append(value)
        new_centers = [
            (sum(bucket) / len(bucket)) if bucket else centers[i]
            for i, bucket in enumerate(buckets)
        ]
        if all(abs(a - b) < 1e-9 for a, b in zip(centers, new_centers)):
            break
        centers = new_centers
    centers = sorted(centers)
    assignments = [min(range(k), key=lambda i: abs(float(value) - centers[i])) for value in values]
    return centers, assignments


def learned_boundaries(centers: list[float]) -> list[float]:
    return [(centers[i] + centers[i + 1]) / 2.0 for i in range(len(centers) - 1)]


def nearest_boundary_pct(value: float, boundaries: list[float]) -> float | None:
    if not boundaries:
        return None
    nearest = min(boundaries, key=lambda boundary: abs(value - boundary))
    if nearest == 0:
        return None
    return abs(value - nearest) / nearest * 100.0


def calendar_regime(month: str) -> str:
    if "2021-11" <= month <= "2023-09":
        return "R1quiet"
    if "2023-10" <= month <= "2024-12":
        return "R2bull"
    if "2025-01" <= month <= "2025-12":
        return "R3strong"
    if month >= "2026-01":
        return "R4parab"
    return "pre_real_m1"


def annotate_months(rows: list[dict], *, boundary_ambiguity_pct: float = 10.0) -> tuple[list[MonthlyRegimeRow], dict]:
    values = [float(row["m15_atr_mean"]) for row in rows]
    centers, assignments = kmeans_1d(values, k=min(4, len(values)))
    boundaries = learned_boundaries(centers)
    annotated: list[MonthlyRegimeRow] = []
    for row, cluster in zip(rows, assignments):
        learned = REGIME_BY_CLUSTER[min(cluster, len(REGIME_BY_CLUSTER) - 1)]
        trend_fraction = float(row["monthly_return_pct"]) / 100.0
        router = detect_regime(float(row["m15_atr_mean"]), trend=trend_fraction)
        cal = calendar_regime(str(row["month"]))
        near_pct = nearest_boundary_pct(float(row["m15_atr_mean"]), boundaries)
        annotated.append(
            MonthlyRegimeRow(
                month=str(row["month"]),
                bars=int(row["bars"]),
                first_close=float(row["first_close"]),
                last_close=float(row["last_close"]),
                high=float(row["high"]),
                low=float(row["low"]),
                m15_atr_mean=float(row["m15_atr_mean"]),
                m15_atr_median=float(row["m15_atr_median"]),
                m15_atr_p90=float(row["m15_atr_p90"]),
                monthly_return_pct=float(row["monthly_return_pct"]),
                range_pct=float(row["range_pct"]),
                trend_atr_multiple=float(row["trend_atr_multiple"]),
                learned_regime=learned,
                learned_cluster=cluster + 1,
                learned_center_atr=float(centers[cluster]),
                router_regime=router,
                calendar_regime=cal,
                router_matches_learned=router == learned,
                calendar_matches_learned=cal == learned,
                near_learned_boundary=(near_pct is not None and near_pct <= boundary_ambiguity_pct),
                nearest_boundary_pct=near_pct,
            )
        )
    meta = {
        "learned_centers": centers,
        "learned_boundaries": boundaries,
        "boundary_ambiguity_pct": boundary_ambiguity_pct,
        "router_thresholds": {
            "vol_tier_low_max": VOL_TIER_LOW_MAX,
            "vol_tier_mid_max": VOL_TIER_MID_MAX,
            "bull_trend_min": BULL_TREND_MIN,
        },
    }
    return annotated, meta


def _fmt_money(value: float) -> str:
    return f"{value:.2f}"


def _fmt_pct(value: float | None) -> str:
    return "-" if value is None else f"{value:.1f}%"


def _comparison(rows: list[MonthlyRegimeRow], attr: str) -> tuple[int, int]:
    matches = sum(1 for row in rows if getattr(row, attr))
    return matches, len(rows) - matches


def render_markdown(rows: list[MonthlyRegimeRow], meta: dict) -> str:
    months = len(rows)
    start = rows[0].month if rows else "n/a"
    end = rows[-1].month if rows else "n/a"
    router_match, router_mismatch = _comparison(rows, "router_matches_learned")
    calendar_match, calendar_mismatch = _comparison(rows, "calendar_matches_learned")
    centers = meta["learned_centers"]
    boundaries = meta["learned_boundaries"]

    lines: list[str] = []
    lines.append("# Regime Calibration Report")
    lines.append("")
    lines.append(f"- months analyzed: {months} ({start} to {end})")
    lines.append("- learned from monthly mean M15 ATR using deterministic 1D clustering")
    lines.append(
        "- learned ATR centers: "
        + ", ".join(f"{REGIME_BY_CLUSTER[i]}={_fmt_money(center)}" for i, center in enumerate(centers))
    )
    if boundaries:
        lines.append("- learned ATR boundaries: " + ", ".join(_fmt_money(b) for b in boundaries))
    lines.append(
        f"- current router thresholds: low<{VOL_TIER_LOW_MAX:.2f}, "
        f"mid<{VOL_TIER_MID_MAX:.2f}, low-vol bull trend>={BULL_TREND_MIN:.3f}"
    )
    lines.append("")
    lines.append("## Agreement")
    lines.append("")
    lines.append("| comparison | matches | mismatches |")
    lines.append("|---|---:|---:|")
    lines.append(f"| live router vs learned clusters | {router_match} | {router_mismatch} |")
    lines.append(f"| sweep calendar vs learned clusters | {calendar_match} | {calendar_mismatch} |")
    lines.append("")

    ambiguous = [row for row in rows if row.near_learned_boundary]
    disagreements = [
        row for row in rows
        if not row.router_matches_learned or not row.calendar_matches_learned or row.near_learned_boundary
    ]
    lines.append("## Review Months")
    lines.append("")
    if disagreements:
        lines.append("| month | ATR | return | learned | router | calendar | boundary distance |")
        lines.append("|---|---:|---:|---|---|---|---:|")
        for row in disagreements:
            lines.append(
                f"| {row.month} | {_fmt_money(row.m15_atr_mean)} | "
                f"{row.monthly_return_pct:.1f}% | {row.learned_regime} | "
                f"{row.router_regime} | {row.calendar_regime} | "
                f"{_fmt_pct(row.nearest_boundary_pct)} |"
            )
    else:
        lines.append("No router/calendar disagreements or boundary-adjacent months.")
    lines.append("")

    lines.append("## Monthly Map")
    lines.append("")
    lines.append("| month | bars | ATR mean | ATR p90 | return | range | trend/ATR | learned | router | calendar |")
    lines.append("|---|---:|---:|---:|---:|---:|---:|---|---|---|")
    for row in rows:
        marker = " *" if row in ambiguous else ""
        lines.append(
            f"| {row.month}{marker} | {row.bars} | {_fmt_money(row.m15_atr_mean)} | "
            f"{_fmt_money(row.m15_atr_p90)} | {row.monthly_return_pct:.1f}% | "
            f"{row.range_pct:.1f}% | {row.trend_atr_multiple:.1f} | "
            f"{row.learned_regime} | {row.router_regime} | {row.calendar_regime} |"
        )
    lines.append("")
    lines.append("* Boundary-adjacent month.")
    lines.append("")
    return "\n".join(lines)


def write_csv(path: Path, rows: list[MonthlyRegimeRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(asdict(rows[0]).keys()) if rows else ["month"])
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def write_json(path: Path, rows: list[MonthlyRegimeRow], meta: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"meta": meta, "months": [asdict(row) for row in rows]}
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def write_xlsx(path: Path, rows: list[MonthlyRegimeRow], meta: dict) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "monthly_regimes"
    headers = list(asdict(rows[0]).keys()) if rows else ["month"]
    ws.append(headers)
    for row in rows:
        ws.append([asdict(row).get(header) for header in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    fills = {
        "R1quiet": PatternFill("solid", fgColor="D9EAD3"),
        "R2bull": PatternFill("solid", fgColor="FFF2CC"),
        "R3strong": PatternFill("solid", fgColor="CFE2F3"),
        "R4parab": PatternFill("solid", fgColor="F4CCCC"),
    }
    learned_col = headers.index("learned_regime") + 1 if "learned_regime" in headers else None
    for row_idx in range(2, ws.max_row + 1):
        regime = ws.cell(row_idx, learned_col).value if learned_col else None
        fill = fills.get(regime)
        if fill:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row_idx, col_idx).fill = fill

    summary = wb.create_sheet("summary")
    summary.append(["metric", "value"])
    summary.append(["learned_centers", ", ".join(f"{v:.4f}" for v in meta["learned_centers"])])
    summary.append(["learned_boundaries", ", ".join(f"{v:.4f}" for v in meta["learned_boundaries"])])
    summary.append(["boundary_ambiguity_pct", meta["boundary_ambiguity_pct"]])
    for cell in summary[1]:
        cell.font = Font(bold=True)
    wb.save(path)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--charts", nargs="+", required=True, help="M1 chart CSV files or globs.")
    p.add_argument("--start-date", default="2021-01-01", help="Ignore chart rows before this date.")
    p.add_argument("--atr-period", type=int, default=14)
    p.add_argument("--boundary-ambiguity-pct", type=float, default=10.0)
    p.add_argument("--out", default="reports/regime_calibration.md")
    p.add_argument("--csv-out", default=None)
    p.add_argument("--json-out", default=None)
    p.add_argument("--xlsx-out", default=None)
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    paths = _expand_chart_paths(args.charts)
    m1 = load_charts(paths, start_date=args.start_date)
    features = monthly_features(m1, atr_period=args.atr_period)
    rows, meta = annotate_months(features, boundary_ambiguity_pct=args.boundary_ambiguity_pct)
    if not rows:
        raise SystemExit("No monthly regime rows could be produced")

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(render_markdown(rows, meta), encoding="utf-8")
    print(f"wrote {out}")
    if args.csv_out:
        write_csv(Path(args.csv_out), rows)
        print(f"wrote {args.csv_out}")
    if args.json_out:
        write_json(Path(args.json_out), rows, meta)
        print(f"wrote {args.json_out}")
    if args.xlsx_out:
        write_xlsx(Path(args.xlsx_out), rows, meta)
        print(f"wrote {args.xlsx_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
