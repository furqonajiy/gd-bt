#!/usr/bin/env python3
"""Victor / V017 trailing-geometry staged sweep (bounded, NOT brute force).

Runs the Victor provider feed (``victor_signals.txt``, GMT+7 embedded) through
``tools/backtest_hybrid.py`` under the **V017 baseline execution geometry**, then
perturbs ONE (occasionally two) controlled trailing-geometry parameter(s) per
candidate — trailing-open distance, trailing-close distance + engage stage, SL
multiplier, max-hold, pending-expiry, entry-SL-gap, final target. It is a small
curated grid (each candidate is a single controlled step from ``base_v017``), NOT
the full Cartesian product.

Candidates are scored with the **rebate-aware** objective
(``tools/rebate_scoring.py``): a candidate that is green only on the $3/closed-lot
rebate — flat/negative on pure trading P&L — is guarded out and never promoted.
On TICK where the archive covers the lifecycle; the recent window uses May+June
tick data (July tick coverage is currently only one day).

Modes / windows (``--end-date`` is EXCLUSIVE in backtest_hybrid):

    smoke         2026-06-25 .. 2026-07-01   (short recent, tick-covered)
    full_recent   2026-05-01 .. 2026-07-01   (May + June -- the ``may_jun`` window)
    validate_top  2026-01-01 .. 2026-07-01   (re-score a prior run's top via --top-json)

Outputs (under ``<out-root>/VICTOR_TRAILING_<mode>/``):

    results.csv          one row per candidate, full schema (geometry + P&L + gates)
    top_candidates.json  ranked survivors (gates passed) by score
    summary.md           human-readable leaderboard + rebate-guard notes

    python tools/sweep_victor_trailing_geometry.py --mode smoke
    python tools/sweep_victor_trailing_geometry.py --mode smoke --skeleton   # no backtests

Research/backtest only. Never trades live; never promotes to live.
"""
from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools.rebate_scoring import (  # noqa: E402
    DEFAULT_REBATE_PER_LOT, SCORE_OBJECTIVES, compute_rebate_metrics,
    passes_rebate_guards, score_candidate,
)

# V017 baseline execution geometry — mirror of cli/candidate_V017_victor_july_trailing.txt.
# These are the ONLY parameters a candidate may perturb (one/two at a time).
V017_BASELINE = {
    "entries": 8,
    "entry_ladder": "range_to_sl",
    "entry_sl_gap": 0.5,
    "activation_delay": 0,
    "pending_expiry": 180,
    "max_hold": 150,
    "sl_multiplier": 1.7,
    "final_target": "TP2",
    "lock_after_tp1": True,
    "lock_after_tp2": True,
    "tp1_lock_delay_minutes": 12,
    "tp2_lock_delay_minutes": 2,
    "profit_lock_mode": "tp_levels",
    "bep_trigger_distance": 3.0,
    "tp1_lock_fraction": 0.5,
    "tp2_lock_target": "TP1",
    "tp3_lock_target": "TP2",
    "runner_after_tp3": False,
    "trailing_open_distance": 0.5,
    "trailing_close_distance": 0.5,
    "trailing_close_after_stage": 1,
}

# Fixed (non-swept) sizing/execution flags — mirror of the V017 snapshot. Risk
# sizing at 5% off capital; slippage overlay 2.0/1.0; DD limit 200 (research).
FIXED_EXEC = [
    "--sync-charts", "true", "--sync-ticks", "true",
    "--max-drawdown-limit-pct", "200", "--progress-interval-seconds", "0",
    "--sizing-mode", "risk", "--lot", "0.01", "--risk", "0.05",
    "--minimum-lot", "0.01", "--maximum-lot", "100.0", "--max-open-lots", "100",
    "--lot-step", "0.01", "--bonus-per-closed-lot", "3.0",
]
ERA_SLIP = ["--lock-tp1-exit-slippage", "2.0", "--lock-tp2-exit-slippage", "1.0"]

DEFAULT_FEED = "victor_signals.txt"

# Window spans (end EXCLUSIVE). ``may_jun`` is the tick-covered recent window
# (July tick coverage is a single day, so the recent sweep uses May+June).
WINDOWS = {
    "smoke": ("2026-06-25", "2026-07-01"),
    "may_jun": ("2026-05-01", "2026-07-01"),
    "jan_jun": ("2026-01-01", "2026-07-01"),
    "jan_jul": ("2026-01-01", "2026-08-01"),
}
MODE_WINDOW = {"smoke": "smoke", "full_recent": "may_jun", "validate_top": "jan_jun"}

# Result schema (geometry knobs + P&L/metrics + gates). One row per candidate.
RESULTS_COLUMNS = [
    "label", "window", "window_start", "window_end",
    "entries", "entry_sl_gap", "sl_multiplier", "pending_expiry", "max_hold",
    "final_target", "tp1_lock_delay_minutes", "tp2_lock_delay_minutes",
    "tp1_lock_fraction", "trailing_open_distance", "trailing_close_distance",
    "trailing_close_after_stage",
    "pure_trading_pnl", "rebate_pnl", "net_pnl", "closed_lots",
    "pure_pnl_per_lot", "net_pnl_per_lot",
    "max_drawdown_pct", "win_rate_pct", "signals_included",
    "tick_signals", "m1_signals", "open_or_pending_left", "included_in_ranking",
    "score", "guard_reason",
]

# Geometry knobs surfaced as CSV columns (subset of V017_BASELINE keys).
GEOMETRY_COLUMNS = [
    "entries", "entry_sl_gap", "sl_multiplier", "pending_expiry", "max_hold",
    "final_target", "tp1_lock_delay_minutes", "tp2_lock_delay_minutes",
    "tp1_lock_fraction", "trailing_open_distance", "trailing_close_distance",
    "trailing_close_after_stage",
]


def _candidate(label: str, **overrides) -> dict:
    """A candidate = the V017 baseline geometry with ``overrides`` applied. Only
    known geometry keys may be overridden (guards against typo'd knobs), and each
    curated candidate is expected to change only one/two parameters."""
    bad = set(overrides) - set(V017_BASELINE)
    if bad:
        raise ValueError(f"unknown geometry keys for {label}: {sorted(bad)}")
    geom = dict(V017_BASELINE)
    geom.update(overrides)
    return {"label": label, **geom}


def build_candidate_grid(mode: str) -> list[dict]:
    """Curated staged grid: ``base_v017`` first, then single-knob perturbations.
    ``smoke`` returns a small representative subset; every other mode returns the
    full curated grid (all required labels)."""
    grid = [
        _candidate("base_v017"),
        # trailing-open distance
        _candidate("open_0_25", trailing_open_distance=0.25),
        _candidate("open_0_50", trailing_open_distance=0.50),
        _candidate("open_0_75", trailing_open_distance=0.75),
        _candidate("open_1_00", trailing_open_distance=1.00),
        # trailing-close distance (+ engage stage)
        _candidate("close_0_25_stage1", trailing_close_distance=0.25, trailing_close_after_stage=1),
        _candidate("close_0_50_stage1", trailing_close_distance=0.50, trailing_close_after_stage=1),
        _candidate("close_0_75_stage1", trailing_close_distance=0.75, trailing_close_after_stage=1),
        _candidate("close_0_50_stage2", trailing_close_distance=0.50, trailing_close_after_stage=2),
        # SL multiplier
        _candidate("slm_1_60", sl_multiplier=1.60),
        _candidate("slm_1_70", sl_multiplier=1.70),
        _candidate("slm_1_80", sl_multiplier=1.80),
        # max-hold
        _candidate("hold_120", max_hold=120),
        _candidate("hold_150", max_hold=150),
        _candidate("hold_180", max_hold=180),
        # pending-expiry
        _candidate("expiry_120", pending_expiry=120),
        _candidate("expiry_180", pending_expiry=180),
        # entry-SL gap
        _candidate("entry_gap_0_50", entry_sl_gap=0.50),
        _candidate("entry_gap_0_70", entry_sl_gap=0.70),
        # final target
        _candidate("tp2_final", final_target="TP2"),
        _candidate("tp3_final", final_target="TP3"),
    ]
    if mode == "smoke":
        keep = {"base_v017", "open_0_75", "close_0_75_stage1", "slm_1_80"}
        return [c for c in grid if c["label"] in keep]
    return grid


def _b(v: bool) -> str:
    return "true" if v else "false"


def _geometry_flags(cand: dict) -> list[str]:
    """Render a candidate's full V017 geometry as backtest_hybrid flags."""
    g = cand
    return [
        "--entries", str(g["entries"]),
        "--entry-ladder", str(g["entry_ladder"]),
        "--entry-sl-gap", str(g["entry_sl_gap"]),
        "--shared-sl", "false",
        "--activation-delay", str(g["activation_delay"]),
        "--pending-expiry", str(g["pending_expiry"]),
        "--max-hold", str(g["max_hold"]),
        "--sl-multiplier", str(g["sl_multiplier"]),
        "--final-target", str(g["final_target"]),
        "--lock-after-tp1", _b(g["lock_after_tp1"]),
        "--lock-after-tp2", _b(g["lock_after_tp2"]),
        "--tp1-lock-delay-minutes", str(g["tp1_lock_delay_minutes"]),
        "--tp2-lock-delay-minutes", str(g["tp2_lock_delay_minutes"]),
        "--profit-lock-mode", str(g["profit_lock_mode"]),
        "--bep-trigger-distance", str(g["bep_trigger_distance"]),
        "--tp1-lock-fraction", str(g["tp1_lock_fraction"]),
        "--tp2-lock-target", str(g["tp2_lock_target"]),
        "--tp3-lock-target", str(g["tp3_lock_target"]),
        "--runner-after-tp3", _b(g["runner_after_tp3"]),
        "--trailing-open-distance", str(g["trailing_open_distance"]),
        "--trailing-close-distance", str(g["trailing_close_distance"]),
        "--trailing-close-after-stage", str(g["trailing_close_after_stage"]),
    ]


def _run(cmd: list[str]) -> subprocess.CompletedProcess:
    return subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True)


def _run_backtest(feed: str, charts: list[str], ticks: list[str], start: str, end: str,
                  out_dir: Path, score_json: Path, geometry_flags: list[str],
                  initial_capital: float) -> Path:
    cmd = [sys.executable, "tools/backtest_hybrid.py", "--signals", feed,
           "--charts", *charts, "--ticks", *ticks,
           *FIXED_EXEC, *geometry_flags, *ERA_SLIP,
           "--output-dir", str(out_dir), "--initial-capital", str(initial_capital),
           "--score-json", str(score_json),
           "--start-date", start, "--end-date", end]
    r = _run(cmd)
    if r.returncode != 0:
        raise SystemExit(f"backtest failed for {feed}:\n{r.stderr[-1500:]}")
    xlsx = out_dir.with_suffix(".xlsx")
    if not xlsx.exists():
        raise SystemExit(f"no workbook at {xlsx}")
    return xlsx


def _money(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _read_summary_pnl(xlsx: Path) -> tuple[float, float]:
    """(pure_trading_pnl, rebate_pnl) from the workbook Summary key/value cells."""
    from openpyxl import load_workbook

    ws = load_workbook(xlsx, data_only=True)["Summary"]
    pure = rebate = 0.0
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if not label:
            continue
        val = ws.cell(row=r, column=2).value
        text = str(label).strip()
        if text == "Trading P&L":
            pure = _money(val)
        elif text == "Closed-lot bonus":
            rebate = _money(val)
    return pure, rebate


def _count_open_pending(xlsx: Path) -> int:
    """Count entries still OPEN/PENDING at window end (incomplete P&L)."""
    from openpyxl import load_workbook

    ws = load_workbook(xlsx, data_only=True)["Per-Entry Detail"]
    hdr = {str(ws.cell(row=2, column=c).value).strip(): c
           for c in range(1, ws.max_column + 1)}
    status_col = hdr.get("Status")
    if not status_col:
        return 0
    n = 0
    for r in range(3, ws.max_row + 1):
        status = ws.cell(row=r, column=status_col).value
        if status and str(status).upper() in ("OPEN", "PENDING"):
            n += 1
    return n


def _blank_row(cand: dict, window: str, span: tuple[str, str]) -> dict:
    """A schema-complete placeholder row (skeleton / no-backtest mode); the geometry
    columns reflect the candidate, metrics stay blank until a real run fills them."""
    row = {"label": cand["label"], "window": window,
           "window_start": span[0], "window_end": span[1]}
    for col in GEOMETRY_COLUMNS:
        row[col] = cand[col]
    for col in ("pure_trading_pnl", "rebate_pnl", "net_pnl", "closed_lots",
                "pure_pnl_per_lot", "net_pnl_per_lot", "max_drawdown_pct",
                "win_rate_pct", "signals_included", "tick_signals", "m1_signals",
                "open_or_pending_left", "included_in_ranking", "score"):
        row[col] = ""
    row["guard_reason"] = "skeleton"
    return row


def evaluate_candidate(cand: dict, window: str, span: tuple[str, str], *,
                       charts: list[str], ticks: list[str], out_root: Path,
                       args: argparse.Namespace) -> dict:
    """Run one candidate (or emit a placeholder row in --skeleton mode)."""
    if args.skeleton:
        return _blank_row(cand, window, span)

    start, end = span
    score_json = out_root / f"score_{cand['label']}.json"
    xlsx = _run_backtest(args.signals, charts, ticks, start, end,
                         out_root / f"bt_{cand['label']}", score_json,
                         _geometry_flags(cand), args.initial_capital)

    score = json.loads(score_json.read_text(encoding="utf-8"))
    pure, rebate = _read_summary_pnl(xlsx)
    closed_lots = rebate / args.rebate_per_lot if args.rebate_per_lot > 0 else 0.0
    m = compute_rebate_metrics(pure, closed_lots, rebate_per_lot=args.rebate_per_lot,
                               rebate_pnl=rebate)
    dd = float(score.get("max_drawdown_pct", 0.0))
    guard_ok, guard_reason = passes_rebate_guards(
        m, min_pure_trading_pnl=args.min_pure_trading_pnl,
        max_rebate_share_of_profit=args.max_rebate_share_of_profit)
    sc = score_candidate(m, args.score_objective, max_drawdown_pct=dd,
                         min_pure_trading_pnl=args.min_pure_trading_pnl,
                         max_rebate_share_of_profit=args.max_rebate_share_of_profit)
    tick_n = int(score.get("tick_signals", 0))
    m1_n = int(score.get("m1_signals", 0))
    mixed = tick_n > 0 and m1_n > 0
    partial_excluded = bool(args.require_full_tick_lifecycle and mixed)
    open_left = _count_open_pending(xlsx) > 0
    included = (guard_ok and not partial_excluded
                and not (args.exclude_open_or_pending and open_left))
    if partial_excluded and guard_ok:
        guard_reason = "excluded: mixed tick/M1 (partial tick lifecycle)"
    elif args.exclude_open_or_pending and open_left and guard_ok:
        guard_reason = "excluded: open/pending left at window end"

    row = _blank_row(cand, window, span)
    row.update({
        "pure_trading_pnl": m.pure_trading_pnl, "rebate_pnl": m.rebate_pnl,
        "net_pnl": m.net_pnl, "closed_lots": m.closed_lots,
        "pure_pnl_per_lot": m.pure_pnl_per_lot, "net_pnl_per_lot": m.net_pnl_per_lot,
        "max_drawdown_pct": dd, "win_rate_pct": float(score.get("win_rate_pct", 0.0)),
        "signals_included": int(score.get("signals_included", 0)),
        "tick_signals": tick_n, "m1_signals": m1_n,
        "open_or_pending_left": open_left, "included_in_ranking": included,
        "score": round(sc, 2), "guard_reason": guard_reason,
    })
    return row


def write_results_csv(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=RESULTS_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in RESULTS_COLUMNS})


def write_top_candidates_json(rows: list[dict], path: Path) -> list[dict]:
    """Rank ranking-included rows by score desc; write + return them."""
    ranked = [r for r in rows if r.get("included_in_ranking") is True]
    ranked.sort(key=lambda r: (r.get("score") if isinstance(r.get("score"), (int, float))
                               else float("-inf")), reverse=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(ranked, indent=2), encoding="utf-8")
    return ranked


def write_summary_md(rows: list[dict], ranked: list[dict], path: Path,
                     mode: str, span: tuple[str, str], objective: str) -> None:
    lines = [
        f"# Victor/V017 trailing-geometry sweep — {mode} ({span[0]}..{span[1]})",
        "",
        f"Score objective: **{objective}**. Base V017 geometry vs single-knob "
        "trailing-geometry perturbations on the Victor provider feed, TICK where "
        "covered. **Rebate-aware**: a candidate green only on the $3/closed-lot "
        "rebate (bad pure P&L) is guarded out, never promoted. Research/backtest "
        "only — never live.",
        "",
        "| label | open | close | stage | slm | hold | expiry | gap | final | "
        "pure $ | net $ | DD% | win% | sigs | ranked | guard |",
        "|---|--:|--:|--:|--:|--:|--:|--:|:--|--:|--:|--:|--:|--:|:--:|:--|",
    ]
    for r in rows:
        lines.append(
            f"| {r['label']} | {r['trailing_open_distance']} | "
            f"{r['trailing_close_distance']} | {r['trailing_close_after_stage']} | "
            f"{r['sl_multiplier']} | {r['max_hold']} | {r['pending_expiry']} | "
            f"{r['entry_sl_gap']} | {r['final_target']} | {r['pure_trading_pnl']} | "
            f"{r['net_pnl']} | {r['max_drawdown_pct']} | {r['win_rate_pct']} | "
            f"{r['signals_included']} | {r['included_in_ranking']} | {r['guard_reason']} |")
    lines += ["", "## Ranked survivors (gates passed)", ""]
    if ranked:
        for i, r in enumerate(ranked, start=1):
            lines.append(f"{i}. **{r['label']}** — score {r['score']} "
                         f"(net ${r['net_pnl']}, pure ${r['pure_trading_pnl']}, "
                         f"DD {r['max_drawdown_pct']}%, open={r['trailing_open_distance']}, "
                         f"close={r['trailing_close_distance']}@stage{r['trailing_close_after_stage']})")
    else:
        lines.append("_No survivors (skeleton run, or all candidates failed the gates)._")
    lines += [
        "",
        "Gates: **rebate guards** (pure-P&L floor + max rebate share), "
        "**partial-tick-lifecycle exclusion** (mixed TICK/M1 windows when "
        "`--require-full-tick-lifecycle`), and **open/pending-left** "
        "(`--exclude-open-or-pending`). Prefer a survivor that beats `base_v017` on "
        "pure P&L at similar/lower DD, consistently across the validation window — "
        "not one that wins on a single lucky signal.",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--mode", choices=list(MODE_WINDOW), default="smoke",
                    help="smoke | full_recent (May+June) | validate_top.")
    ap.add_argument("--window", choices=list(WINDOWS), default=None,
                    help="override the mode's default window (smoke|may_jun|jan_jun|jan_jul).")
    ap.add_argument("--start-date", default=None,
                    help="override the window START (YYYY-MM-DD); default = the mode/window start.")
    ap.add_argument("--end-date", default=None,
                    help="override the window END (YYYY-MM-DD, EXCLUSIVE); default = the mode/window end.")
    ap.add_argument("--skeleton", action="store_true",
                    help="emit the results schema with placeholder rows; run NO backtests.")
    ap.add_argument("--signals", default=DEFAULT_FEED,
                    help="Victor provider feed (default: victor_signals.txt).")
    ap.add_argument("--charts", nargs="+",
                    default=["data/XAUUSD_M1_202605_ELEV8.csv", "data/XAUUSD_M1_202606_ELEV8.csv"])
    ap.add_argument("--ticks", nargs="+",
                    default=["data/ticks/XAUUSD_TICK_20260[56]*_ELEV8.csv"])
    ap.add_argument("--initial-capital", type=float, default=50000.0,
                    help="starting equity (run the workflow at 50000 and 5000 for the 50K/5K pair).")
    # --input-candidates is an automation alias for --top-json.
    ap.add_argument("--top-json", "--input-candidates", dest="top_json", default=None,
                    help="validate_top: a prior run's top_candidates.json to re-score.")
    # --output-dir is an automation alias for --out-root; the script ALWAYS writes
    # <out-root>/VICTOR_TRAILING_<mode>/ (the subfolder is still created under it).
    ap.add_argument("--out-root", "--output-dir", dest="out_root", default="reports",
                    help="output ROOT; the script writes <out-root>/VICTOR_TRAILING_<mode>/.")
    ap.add_argument("--rebate-per-lot", type=float, default=DEFAULT_REBATE_PER_LOT)
    ap.add_argument("--min-pure-trading-pnl", type=float, default=0.0,
                    help="reject candidates whose pure trading P&L < X.")
    ap.add_argument("--max-rebate-share-of-profit", type=float, default=0.50,
                    help="reject candidates whose net profit is >X rebate.")
    # --rank-objective is an automation alias for --score-objective.
    ap.add_argument("--score-objective", "--rank-objective", dest="score_objective",
                    choices=list(SCORE_OBJECTIVES), default="edge_plus_rebate_guarded",
                    help="rebate-aware ranking objective (--rank-objective is an automation alias).")
    ap.add_argument("--require-full-tick-lifecycle", "--require-full-lifecycle-ticks",
                    dest="require_full_tick_lifecycle", action="store_true",
                    help="exclude candidates whose window mixes TICK and M1 (partial coverage).")
    ap.add_argument("--exclude-open-or-pending", "--fail-on-open-or-pending",
                    dest="exclude_open_or_pending", action="store_true",
                    help="exclude candidates that left positions OPEN/PENDING at window end.")
    return ap


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    window = args.window or MODE_WINDOW[args.mode]
    span = WINDOWS[window]
    if args.start_date or args.end_date:   # explicit window overrides (--end-date EXCLUSIVE)
        span = (args.start_date or span[0], args.end_date or span[1])
    out_root = Path(args.out_root) / f"VICTOR_TRAILING_{args.mode}"
    out_root.mkdir(parents=True, exist_ok=True)

    if args.mode == "validate_top" and args.top_json:
        prior = json.loads(Path(args.top_json).read_text(encoding="utf-8"))
        candidates = [
            _candidate(c["label"], **{k: c[k] for k in V017_BASELINE if k in c})
            for c in prior
        ]
    else:
        candidates = build_candidate_grid(args.mode)

    rows: list[dict] = []
    for cand in candidates:
        print(f"[victor-sweep] {cand['label']}: open={cand['trailing_open_distance']} "
              f"close={cand['trailing_close_distance']}@stage{cand['trailing_close_after_stage']} "
              f"slm={cand['sl_multiplier']} hold={cand['max_hold']} final={cand['final_target']}"
              f"{' [skeleton]' if args.skeleton else ''}", flush=True)
        rows.append(evaluate_candidate(cand, window, span, charts=args.charts,
                                       ticks=args.ticks, out_root=out_root, args=args))

    results_csv = out_root / "results.csv"
    top_json = out_root / "top_candidates.json"
    summary = out_root / "summary.md"
    write_results_csv(rows, results_csv)
    ranked = write_top_candidates_json(rows, top_json)
    write_summary_md(rows, ranked, summary, args.mode, span, args.score_objective)
    print(f"[victor-sweep] results: {results_csv}")
    print(f"[victor-sweep] top:     {top_json} ({len(ranked)} survivors)")
    print(f"[victor-sweep] summary: {summary}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
