"""Excel report writer for backtest results.

Three sheets:

  Sheet 1 "Summary"          -- config, overall stats, monthly breakdown
                                (now with a P&L % column).
  Sheet 2 "Daily Breakdown"  -- one row per calendar day in the chart range
                                (zero-activity days included), with P&L %.
  Sheet 3 "Per-Entry Detail" -- one row per Entry slot (3 per signal),
                                color-coded by outcome.

Soft dependency on openpyxl. Importing this module raises ImportError if
openpyxl is missing; the backtest CLI catches that and proceeds without
Excel output.
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# styling
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(bold=True, color="000000", size=11)

# Status colors. Tuned to be readable but visually distinct.
STATUS_FILL = {
    # Signal-level statuses
    "WIN":      PatternFill("solid", fgColor="C6EFCE"),
    "LOSS":     PatternFill("solid", fgColor="FFC7CE"),
    "NO_FILL":  PatternFill("solid", fgColor="EAEAEA"),
    "OPEN":     PatternFill("solid", fgColor="FFEB9C"),
    "BREAKEVEN":PatternFill("solid", fgColor="DDEBF7"),
    # Entry-level statuses
    "TP1":       PatternFill("solid", fgColor="C6EFCE"),
    "TP2":       PatternFill("solid", fgColor="C6EFCE"),
    "TP3":       PatternFill("solid", fgColor="C6EFCE"),
    "SL":        PatternFill("solid", fgColor="FFC7CE"),
    "LOCK_TP1":  PatternFill("solid", fgColor="DDEBF7"),  # locked breakeven+
    "TIME_EXIT": PatternFill("solid", fgColor="FFEB9C"),
    "PENDING":   PatternFill("solid", fgColor="EAEAEA"),
}

# Lighter gray for "no signals on this day" rows in the Daily sheet -- visually
# distinct from the NO_FILL gray used elsewhere so a quiet day doesn't look
# like a failed entry.
QUIET_DAY_FILL = PatternFill("solid", fgColor="F5F5F5")

THIN = Side(border_style="thin", color="BFBFBF")
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _fmt_dt(value: Any) -> Optional[str]:
    """Return an ISO-ish string for datetimes; pass through None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def _autosize(ws: Worksheet, min_width: int = 10, max_width: int = 28) -> None:
    """Approximate auto-fit by inspecting cell content lengths."""
    for column_cells in ws.columns:
        try:
            length = max(
                len(str(c.value)) if c.value is not None else 0
                for c in column_cells
            )
        except ValueError:
            length = min_width
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, length + 2))


def _write_header(ws: Worksheet, row: int, columns: list[str]) -> None:
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = GRID


def _apply_row_fill(ws: Worksheet, row: int, ncols: int, fill: PatternFill) -> None:
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = fill


def _apply_status_fill(ws: Worksheet, row: int, ncols: int, status: str) -> None:
    fill = STATUS_FILL.get(status)
    if fill is None:
        return
    _apply_row_fill(ws, row, ncols, fill)


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _write_summary_sheet(ws: Worksheet, result: dict) -> None:
    ws.title = "Summary"
    cfg = result.get("config", {})

    # Title --------------------------------------------------------------
    ws["A1"] = "XAUUSD BACKTEST RESULTS"
    ws["A1"].font = Font(bold=True, size=14, color="305496")
    ws.merge_cells("A1:D1")

    # Configuration ------------------------------------------------------
    row = 3
    ws.cell(row=row, column=1, value="Configuration").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    config_pairs = [
        ("Initial capital", f"${cfg.get('initial_capital', 0):,.2f}"),
        ("Risk per signal", f"{cfg.get('risk_per_signal', 0) * 100:.2f}%"),
        ("Entry count", cfg.get("entry_count")),
        ("Activation delay (min)", cfg.get("activation_delay_minutes")),
        ("Pending expiry (min)", cfg.get("pending_expiry_minutes")),
        ("Max hold (min)", cfg.get("max_hold_minutes")),
        ("SL multiplier", cfg.get("sl_multiplier")),
        ("Final target", cfg.get("final_target")),
        ("Lock after TP1", cfg.get("lock_after_tp1")),
        ("Minimum lot", cfg.get("minimum_lot")),
        ("Lot step", cfg.get("lot_step")),
        ("Chart start", result.get("chart_start")),
        ("Chart end", result.get("chart_end")),
    ]
    for label, value in config_pairs:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Overall stats ------------------------------------------------------
    row += 1
    ws.cell(row=row, column=1, value="Overall Performance").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    initial = cfg.get("initial_capital", 0) or 1
    final = result.get("final_equity", 0)
    net = result.get("net_profit", 0)
    return_pct = (final / initial - 1) * 100 if initial else 0.0

    perf_pairs = [
        ("Final equity", f"${final:,.2f}"),
        ("Net profit", f"${net:,.2f}"),
        ("Return", f"{return_pct:+.2f}%"),
        ("Realized P&L", f"${result.get('realized_pnl', 0):,.2f}"),
        ("Max drawdown", f"{result.get('max_drawdown_pct', 0):.2f}%"),
        ("Signals parsed", result.get("signals_parsed")),
        ("Signals included", result.get("signals_included")),
        ("Signals excluded", result.get("signals_excluded")),
        ("Wins", result.get("wins")),
        ("Losses", result.get("losses")),
        ("No fills", result.get("no_fills")),
        ("Open", result.get("open")),
        ("Win rate", f"{result.get('win_rate_pct', 0):.2f}%"),
    ]
    for label, value in perf_pairs:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Monthly breakdown -------------------------------------------------
    row += 1
    ws.cell(row=row, column=1, value="Monthly Breakdown").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    # Subheader spans the 9 data columns now (added "P&L %").
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    headers = ["Month", "Signals", "Wins", "Losses", "No-fills",
               "Win rate", "P&L", "P&L %", "Equity end-of-month"]
    _write_header(ws, row, headers)
    row += 1

    money_fmt = '"$"#,##0.00;[Red]-"$"#,##0.00'

    monthly = result.get("monthly", []) or []
    for m in monthly:
        cells = [
            m.get("month"),
            m.get("signals"),
            m.get("wins"),
            m.get("losses"),
            m.get("no_fills"),
            f"{m.get('win_rate_pct', 0):.1f}%",
            m.get("pnl", 0.0),
            f"{m.get('pnl_pct', 0):+.2f}%",
            m.get("equity_end"),
        ]
        for c, v in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = GRID
            if c in (7, 9):  # P&L and Equity columns
                cell.number_format = money_fmt
        row += 1

    ws.freeze_panes = "A2"
    _autosize(ws)


# ---------------------------------------------------------------------------
# Daily Breakdown sheet
# ---------------------------------------------------------------------------

def _write_daily_sheet(ws: Worksheet, result: dict) -> None:
    ws.title = "Daily Breakdown"
    headers = ["Date", "Signals", "Wins", "Losses", "No-fills",
               "Win rate", "P&L", "P&L %", "Equity end-of-day"]
    _write_header(ws, 1, headers)

    money_fmt = '"$"#,##0.00;[Red]-"$"#,##0.00'
    daily = result.get("daily", []) or []

    for r_idx, d in enumerate(daily, start=2):
        signals = d.get("signals", 0) or 0
        pnl = d.get("pnl", 0.0) or 0.0

        cells = [
            d.get("date"),
            signals,
            d.get("wins", 0),
            d.get("losses", 0),
            d.get("no_fills", 0),
            f"{d.get('win_rate_pct', 0):.1f}%" if signals else "-",
            pnl,
            f"{d.get('pnl_pct', 0):+.2f}%",
            d.get("equity_end"),
        ]
        for c_idx, v in enumerate(cells, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.border = GRID
            if c_idx in (7, 9):
                cell.number_format = money_fmt

        # Row coloring: gray for zero-activity days, green for net positive,
        # red for net negative, default white for active-but-flat days.
        if signals == 0:
            _apply_row_fill(ws, r_idx, len(headers), QUIET_DAY_FILL)
        elif pnl > 0:
            _apply_row_fill(ws, r_idx, len(headers), STATUS_FILL["WIN"])
        elif pnl < 0:
            _apply_row_fill(ws, r_idx, len(headers), STATUS_FILL["LOSS"])

    ws.freeze_panes = "A2"
    if daily:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(daily) + 1}"
    _autosize(ws)


# ---------------------------------------------------------------------------
# Per-Entry Detail sheet
# ---------------------------------------------------------------------------

ENTRY_COLUMNS = [
    ("global_id", "Sig ID", None),
    ("signal_key", "Sig Key", None),
    ("signal_date", "Date", None),
    ("signal_time_source", "Time (src)", None),
    ("source_tz", "Src TZ", None),
    ("signal_time_chart", "Time (GMT+3)", "datetime"),
    ("side", "Side", None),
    ("range_low", "Range Low", "price"),
    ("range_high", "Range High", "price"),
    ("original_SL", "Orig SL", "price"),
    ("TP1", "TP1", "price"),
    ("TP2", "TP2", "price"),
    ("TP3", "TP3", "price"),
    ("final_target_label", "Tgt Lbl", None),
    ("final_target_price", "Tgt Price", "price"),
    ("entry_index", "Entry #", None),
    ("entry_price", "Entry Price", "price"),
    ("effective_SL", "Effective SL", "price"),
    ("SL_distance", "SL Dist ($)", "price"),
    ("lot", "Lot", "lot"),
    ("entry_status", "Entry Status", None),
    ("fill_time", "Fill Time", "datetime"),
    ("exit_time", "Exit Time", "datetime"),
    ("exit_price", "Exit Price", "price"),
    ("stop_at_exit", "Stop @ Exit", "price"),
    ("pnl", "P&L", "money"),
    ("first_fill_time", "1st Fill (sig)", "datetime"),
    ("time_exit_deadline", "Time Exit Deadline", "datetime"),
    ("signal_status", "Sig Status", None),
    ("equity_before", "Equity Before", "money"),
    ("equity_after", "Equity After", "money"),
]


def _write_entries_sheet(ws: Worksheet, result: dict) -> None:
    ws.title = "Per-Entry Detail"
    headers = [c[1] for c in ENTRY_COLUMNS]
    _write_header(ws, 1, headers)

    money_fmt = '"$"#,##0.00;[Red]-"$"#,##0.00'
    price_fmt = "#,##0.00"
    lot_fmt = "0.00"   # lots are floored to broker step (default 0.01)

    rows = result.get("entry_rows", []) or []
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (key, _label, kind) in enumerate(ENTRY_COLUMNS, start=1):
            value = row.get(key)
            if kind == "datetime":
                value = _fmt_dt(value)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = GRID
            if kind == "money":
                cell.number_format = money_fmt
            elif kind == "price":
                cell.number_format = price_fmt
            elif kind == "lot":
                cell.number_format = lot_fmt
        _apply_status_fill(ws, r_idx, len(headers), row.get("entry_status", ""))

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    _autosize(ws)


# ---------------------------------------------------------------------------
# entry point
# ---------------------------------------------------------------------------

def write_excel_report(result: dict, output_path: Path | str) -> Path:
    """Render the backtest result to a styled .xlsx file. Returns the path.

    Sheets, in order:
      1. Summary           (config + overall + monthly breakdown w/ P&L %)
      2. Daily Breakdown   (every calendar day in the chart range)
      3. Per-Entry Detail  (one row per Entry slot, color-coded)
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    summary_ws = wb.active
    _write_summary_sheet(summary_ws, result)

    daily_ws = wb.create_sheet("Daily Breakdown")
    _write_daily_sheet(daily_ws, result)

    entries_ws = wb.create_sheet("Per-Entry Detail")
    _write_entries_sheet(entries_ws, result)

    wb.save(output_path)
    return output_path
