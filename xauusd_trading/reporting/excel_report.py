"""Excel report writer for backtest results.

Three sheets:
  1. Summary           — config, overall stats, monthly breakdown with P&L %
  2. Daily Breakdown   — one row per calendar day in the chart range
  3. Per-Entry Detail  — one row per Entry slot, color-coded by outcome

Soft dependency on openpyxl; importing this module raises ImportError if
missing, and the backtest CLI catches that to skip Excel output gracefully.
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# styling
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(bold=True, color="000000", size=11)

STATUS_FILL = {
    # Signal-level
    "WIN":       PatternFill("solid", fgColor="C6EFCE"),
    "LOSS":      PatternFill("solid", fgColor="FFC7CE"),
    "NO_FILL":   PatternFill("solid", fgColor="EAEAEA"),
    "OPEN":      PatternFill("solid", fgColor="FFEB9C"),
    "BREAKEVEN": PatternFill("solid", fgColor="DDEBF7"),
    # Entry-level
    "TP1":       PatternFill("solid", fgColor="C6EFCE"),
    "TP2":       PatternFill("solid", fgColor="C6EFCE"),
    "TP3":       PatternFill("solid", fgColor="C6EFCE"),
    "SL":        PatternFill("solid", fgColor="FFC7CE"),
    "LOCK_TP1":  PatternFill("solid", fgColor="DDEBF7"),
    "TIME_EXIT": PatternFill("solid", fgColor="FFEB9C"),
    "PENDING":   PatternFill("solid", fgColor="EAEAEA"),
}

# Distinct from NO_FILL gray so a quiet day doesn't look like a failed entry.
QUIET_DAY_FILL = PatternFill("solid", fgColor="F5F5F5")
PROFIT_FILL = PatternFill("solid", fgColor="C6EFCE")
LOSS_FILL = PatternFill("solid", fgColor="FFC7CE")
WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")
BONUS_FILL = PatternFill("solid", fgColor="DDEBF7")
LOTS_FILL = PatternFill("solid", fgColor="E2F0D9")

THIN = Side(border_style="thin", color="BFBFBF")
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _fmt_dt(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def _autosize(ws: Worksheet, min_width: int = 10, max_width: int = 28) -> None:
    """Approximate auto-fit by inspecting cell content lengths."""
    for column_cells in ws.columns:
        try:
            length = max(
                len(str(c.value)) if c.value is not None else 0
                for c in column_cells
            )
        except ValueError:
            length = min_width
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, length + 2))


def _write_header(ws: Worksheet, row: int, columns: list[str]) -> None:
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = GRID


def _apply_row_fill(ws: Worksheet, row: int, ncols: int, fill: PatternFill) -> None:
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = fill


def _apply_status_fill(ws: Worksheet, row: int, ncols: int, status: str) -> None:
    fill = STATUS_FILL.get(status)
    if fill is None:
        return
    _apply_row_fill(ws, row, ncols, fill)


def _style_money_cell(cell, money_fmt: str, value: float | None = None) -> None:
    cell.number_format = money_fmt
    if value is None:
        return
    if value > 0:
        cell.font = Font(color="006100", bold=True)
    elif value < 0:
        cell.font = Font(color="9C0006", bold=True)


def _style_pct_cell(cell, value: float | None = None) -> None:
    if value is None:
        return
    if value > 0:
        cell.font = Font(color="006100", bold=True)
    elif value < 0:
        cell.font = Font(color="9C0006", bold=True)


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _write_summary_sheet(ws: Worksheet, result: dict) -> None:
    ws.title = "Summary"
    cfg = result.get("config", {})

    ws["A1"] = "XAUUSD BACKTEST RESULTS"
    ws["A1"].font = Font(bold=True, size=14, color="305496")
    ws.merge_cells("A1:D1")

    row = 3
    ws.cell(row=row, column=1, value="Configuration").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    config_pairs = [
        ("Initial capital", f"${cfg.get('initial_capital', 0):,.2f}"),
        ("Sizing mode", cfg.get("sizing_mode")),
        ("Risk per signal", f"{cfg.get('risk_per_signal', 0) * 100:.2f}%"),
        ("Entry count", cfg.get("entry_count")),
        ("Entry ladder", cfg.get("entry_ladder")),
        ("Activation delay (min)", cfg.get("activation_delay_minutes")),
        ("Pending expiry (min)", cfg.get("pending_expiry_minutes")),
        ("Max hold (min)", cfg.get("max_hold_minutes")),
        ("SL multiplier", cfg.get("sl_multiplier")),
        ("Final target", cfg.get("final_target")),
        ("Lock after TP1", cfg.get("lock_after_tp1")),
        ("Lock after TP2", cfg.get("lock_after_tp2")),
        ("Bonus / closed lot", f"${cfg.get('bonus_per_closed_lot', 0):,.2f}"),
        ("Minimum lot", cfg.get("minimum_lot")),
        ("Lot step", cfg.get("lot_step")),
        ("Chart start", result.get("chart_start")),
        ("Chart end", result.get("chart_end")),
    ]
    for label, value in config_pairs:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Overall Performance").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    initial = cfg.get("initial_capital", 0) or 1
    final = result.get("final_equity", 0)
    net = result.get("net_profit", 0)
    return_pct = (final / initial - 1) * 100 if initial else 0.0

    perf_pairs = [
        ("Final equity", f"${final:,.2f}"),
        ("Net profit incl. bonus", f"${net:,.2f}"),
        ("Trading P&L", f"${result.get('trading_pnl', 0):,.2f}"),
        ("Closed-lot bonus", f"${result.get('bonus', 0):,.2f}"),
        ("Closed lots", f"{result.get('closed_lots', 0):,.2f}"),
        ("Return", f"{return_pct:+.2f}%"),
        ("Realized P&L", f"${result.get('realized_pnl', 0):,.2f}"),
        ("Max drawdown", f"{result.get('max_drawdown_pct', 0):.2f}%"),
        ("Signals parsed", result.get("signals_parsed")),
        ("Signals included", result.get("signals_included")),
        ("Signals excluded", result.get("signals_excluded")),
        ("Wins", result.get("wins")),
        ("Losses", result.get("losses")),
        ("No fills", result.get("no_fills")),
        ("Open", result.get("open")),
        ("Win rate", f"{result.get('win_rate_pct', 0):.2f}%"),
    ]
    for label, value in perf_pairs:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Monthly Breakdown").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    row += 1
    headers = ["Month", "Signals", "Wins", "Losses", "No-fills",
               "Win rate", "Trading P&L", "Bonus", "Closed lots",
               "Net P&L", "P&L %", "Equity end-of-month"]
    _write_header(ws, row, headers)
    row += 1

    money_fmt = '"$"#,##0.00;[Red]-"$"#,##0.00'
    lot_fmt = "0.00"

    monthly = result.get("monthly", []) or []
    for m in monthly:
        pnl = m.get("pnl", 0.0) or 0.0
        pnl_pct = m.get("pnl_pct", 0.0) or 0.0
        signals = m.get("signals", 0) or 0
        cells = [
            m.get("month"),
            signals,
            m.get("wins"),
            m.get("losses"),
            m.get("no_fills"),
            f"{m.get('win_rate_pct', 0):.1f}%",
            m.get("trading_pnl", pnl - (m.get("bonus", 0.0) or 0.0)),
            m.get("bonus", 0.0),
            m.get("closed_lots", 0.0),
            pnl,
            f"{pnl_pct:+.2f}%",
            m.get("equity_end"),
        ]
        if signals == 0:
            _apply_row_fill(ws, row, len(headers), QUIET_DAY_FILL)
        elif pnl > 0:
            _apply_row_fill(ws, row, len(headers), PROFIT_FILL)
        elif pnl < 0:
            _apply_row_fill(ws, row, len(headers), LOSS_FILL)
        for c, v in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = GRID
            if c in (7, 8, 10, 12):
                _style_money_cell(cell, money_fmt, float(v or 0.0) if isinstance(v, (int, float)) else None)
            elif c == 9:
                cell.number_format = lot_fmt
                cell.fill = LOTS_FILL
            elif c == 11:
                _style_pct_cell(cell, pnl_pct)
            if c == 8:
                cell.fill = BONUS_FILL
        row += 1

    ws.freeze_panes = "A2"
    _autosize(ws)


# ---------------------------------------------------------------------------
# Daily Breakdown sheet
# ---------------------------------------------------------------------------

def _write_daily_sheet(ws: Worksheet, result: dict) -> None:
    ws.title = "Daily Breakdown"
    headers = ["Date", "Signals", "Wins", "Losses", "No-fills",
               "Win rate", "Trading P&L", "Bonus", "Closed lots",
               "Net P&L", "P&L %", "Equity end-of-day", "Drawdown %"]
    _write_header(ws, 1, headers)

    money_fmt = '"$"#,##0.00;[Red]-"$"#,##0.00'
    lot_fmt = "0.00"
    daily = result.get("daily", []) or []
    peak_equity = (result.get("config", {}) or {}).get("initial_capital", 0) or 0

    for r_idx, d in enumerate(daily, start=2):
        signals = d.get("signals", 0) or 0
        pnl = d.get("pnl", 0.0) or 0.0
        trading_pnl = d.get("trading_pnl", pnl - (d.get("bonus", 0.0) or 0.0)) or 0.0
        bonus = d.get("bonus", 0.0) or 0.0
        closed_lots = d.get("closed_lots", 0.0) or 0.0
        equity_end = d.get("equity_end", 0.0) or 0.0
        if equity_end > peak_equity:
            peak_equity = equity_end
        drawdown_pct = (equity_end - peak_equity) / peak_equity * 100.0 if peak_equity else 0.0
        pnl_pct = d.get("pnl_pct", 0.0) or 0.0

        cells = [
            d.get("date"),
            signals,
            d.get("wins", 0),
            d.get("losses", 0),
            d.get("no_fills", 0),
            f"{d.get('win_rate_pct', 0):.1f}%" if signals else "-",
            trading_pnl,
            bonus,
            closed_lots,
            pnl,
            f"{pnl_pct:+.2f}%",
            equity_end,
            f"{drawdown_pct:.2f}%",
        ]

        # Row coloring: gray = no activity, green = net positive, red = net negative.
        # Orange warning = active day with drawdown below -20%.
        if signals == 0:
            _apply_row_fill(ws, r_idx, len(headers), QUIET_DAY_FILL)
        elif drawdown_pct <= -20.0:
            _apply_row_fill(ws, r_idx, len(headers), WARNING_FILL)
        elif pnl > 0:
            _apply_row_fill(ws, r_idx, len(headers), PROFIT_FILL)
        elif pnl < 0:
            _apply_row_fill(ws, r_idx, len(headers), LOSS_FILL)

        for c_idx, v in enumerate(cells, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.border = GRID
            if c_idx in (7, 8, 10, 12):
                _style_money_cell(cell, money_fmt, float(v or 0.0) if isinstance(v, (int, float)) else None)
            elif c_idx == 9:
                cell.number_format = lot_fmt
                cell.fill = LOTS_FILL
            elif c_idx in (11, 13):
                _style_pct_cell(cell, pnl_pct if c_idx == 11 else drawdown_pct)
            if c_idx == 8:
                cell.fill = BONUS_FILL
            if c_idx == 13 and drawdown_pct <= -20.0:
                cell.fill = WARNING_FILL
                cell.font = Font(color="9C6500", bold=True)

    ws.freeze_panes = "A2"
    if daily:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(daily) + 1}"
    _autosize(ws)


# ---------------------------------------------------------------------------
# Per-Entry Detail sheet
# ---------------------------------------------------------------------------

ENTRY_COLUMNS = [
    ("global_id", "Sig ID", None),
    ("signal_key", "Sig Key", None),
    ("entry_key", "Entry Key", None),
    ("entry_number", "Entry #", None),
    ("signal_date", "Date", None),
    ("signal_time_source", "Time (src)", None),
    ("source_tz", "Src TZ", None),
    ("signal_time_chart", "Time (GMT+3)", "datetime"),
    ("side", "Side", None),
    ("range_low", "Range Low", "price"),
    ("range_high", "Range High", "price"),
    ("original_SL", "Orig SL", "price"),
    ("TP1", "TP1", "price"),
    ("TP2", "TP2", "price"),
    ("TP3", "TP3", "price"),
    ("final_target_label", "Tgt Lbl", None),
    ("final_target_price", "Tgt Price", "price"),
    ("entry_price", "Entry Price", "price"),
    ("effective_SL", "Effective SL", "price"),
    ("SL_distance", "SL Dist ($)", "price"),
    ("lot", "Lot", "lot"),
    ("entry_status", "Entry Status", None),
    ("fill_time", "Fill Time", "datetime"),
    ("exit_time", "Exit Time", "datetime"),
    ("exit_price", "Exit Price", "price"),
    ("stop_at_exit", "Stop @ Exit", "price"),
    ("trading_pnl", "Trading P&L", "money"),
    ("closed_lots", "Closed Lots", "lot"),
    ("bonus", "Bonus", "money"),
    ("pnl", "Net P&L", "money"),
    ("first_fill_time", "1st Fill (sig)", "datetime"),
    ("time_exit_deadline", "Time Exit Deadline", "datetime"),
    ("signal_status", "Sig Status", None),
    ("equity_before", "Equity Before", "money"),
    ("equity_after", "Equity After", "money"),
]


def _write_entries_sheet(ws: Worksheet, result: dict) -> None:
    ws.title = "Per-Entry Detail"
    headers = [c[1] for c in ENTRY_COLUMNS]
    _write_header(ws, 1, headers)

    money_fmt = '"$"#,##0.00;[Red]-"$"#,##0.00'
    price_fmt = "#,##0.00"
    lot_fmt = "0.00"

    rows = result.get("entry_rows", []) or []
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (key, _label, kind) in enumerate(ENTRY_COLUMNS, start=1):
            value = row.get(key)
            if kind == "datetime":
                value = _fmt_dt(value)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = GRID
            if kind == "money":
                _style_money_cell(cell, money_fmt, float(value or 0.0) if isinstance(value, (int, float)) else None)
            elif kind == "price":
                cell.number_format = price_fmt
            elif kind == "lot":
                cell.number_format = lot_fmt
            if key == "bonus":
                cell.fill = BONUS_FILL
            elif key == "closed_lots":
                cell.fill = LOTS_FILL
        _apply_status_fill(ws, r_idx, len(headers), row.get("entry_status", ""))
        # Re-highlight bonus and closed-lot cells after status row fill.
        bonus_col = [c[0] for c in ENTRY_COLUMNS].index("bonus") + 1
        lots_col = [c[0] for c in ENTRY_COLUMNS].index("closed_lots") + 1
        ws.cell(row=r_idx, column=bonus_col).fill = BONUS_FILL
        ws.cell(row=r_idx, column=lots_col).fill = LOTS_FILL

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    _autosize(ws)


# ---------------------------------------------------------------------------
# entry point
# ---------------------------------------------------------------------------

def write_excel_report(result: dict, output_path: Path | str) -> Path:
    """Render the backtest result to a styled .xlsx file."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    summary_ws = wb.active
    _write_summary_sheet(summary_ws, result)

    daily_ws = wb.create_sheet("Daily Breakdown")
    _write_daily_sheet(daily_ws, result)

    entries_ws = wb.create_sheet("Per-Entry Detail")
    _write_entries_sheet(entries_ws, result)

    wb.save(output_path)
    return output_path
