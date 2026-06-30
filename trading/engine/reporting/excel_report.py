"""Excel report writer for backtest results.

Four sheets:
  1. Summary           — config, overall stats, entry-outcome counts, realized
                         risk:reward, and the monthly breakdown
  2. Weekly Breakdown  — one row per feed-zone month-week (W1=days 1-7, W2=8-14,
                         ..., W5=29-31), the mid-grain view between Monthly and
                         Daily, with the same P&L / win-rate / equity columns
  3. Daily Breakdown   — one row per traded day (pre-start padding excluded),
                         with per-entry outcome counts and realized R per day
  4. Per-Entry Detail  — one row per Entry slot, split into ORIGINAL (signal)
                         vs EXECUTED (backtest result) column groups

Soft dependency on openpyxl; importing this module raises ImportError if
missing, and the backtest CLI catches that to skip Excel output gracefully.
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# styling
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(bold=True, color="000000", size=11)

# Per-Entry column-group identity: ORIGINAL (signal) vs EXECUTED (result).
ORIG_BANNER_FILL = PatternFill("solid", fgColor="2E75B6")     # blue
EXEC_BANNER_FILL = PatternFill("solid", fgColor="548235")     # green
ID_BANNER_FILL = PatternFill("solid", fgColor="305496")
ORIG_HEADER_FILL = PatternFill("solid", fgColor="8EAADB")     # lighter blue
EXEC_HEADER_FILL = PatternFill("solid", fgColor="A9D08E")     # lighter green
ORIGINAL_CELL_FILL = PatternFill("solid", fgColor="DDEBF7")   # constant tint for signal cols
LIVE_BANNER_FILL = PatternFill("solid", fgColor="7030A0")     # purple
LIVE_HEADER_FILL = PatternFill("solid", fgColor="CCC0DA")     # light purple
LIVE_CELL_FILL = PatternFill("solid", fgColor="E9E1F2")       # constant tint for live cols
LIVE_DIFF_FILL = PatternFill("solid", fgColor="FFC000")       # plan-vs-live mismatch
BANNER_FONT = Font(bold=True, color="FFFFFF", size=11)

# A live cell differs from the plan when their prices are apart by more than this.
LIVE_DIFF_TOL = 0.005

STATUS_FILL = {
    # Signal-level
    "WIN":       PatternFill("solid", fgColor="C6EFCE"),
    "LOSS":      PatternFill("solid", fgColor="FFC7CE"),
    "NO_FILL":   PatternFill("solid", fgColor="EAEAEA"),
    "OPEN":      PatternFill("solid", fgColor="FFEB9C"),
    "BREAKEVEN": PatternFill("solid", fgColor="DDEBF7"),
    # Entry-level exits
    "TP1":           PatternFill("solid", fgColor="C6EFCE"),
    "TP2":           PatternFill("solid", fgColor="C6EFCE"),
    "TP3":           PatternFill("solid", fgColor="C6EFCE"),
    "SL":            PatternFill("solid", fgColor="FFC7CE"),
    "LOCK_TP1":      PatternFill("solid", fgColor="E2EFDA"),
    "LOCK_TP2":      PatternFill("solid", fgColor="E2EFDA"),
    "LOCK_HALF_TP1": PatternFill("solid", fgColor="E2EFDA"),
    "BEP":           PatternFill("solid", fgColor="DDEBF7"),
    "TRAILING_STOP": PatternFill("solid", fgColor="C6EFCE"),
    "TIME_EXIT":     PatternFill("solid", fgColor="FFEB9C"),
    "PENDING":       PatternFill("solid", fgColor="EAEAEA"),
}

# Distinct from NO_FILL gray so a quiet day doesn't look like a failed entry.
QUIET_DAY_FILL = PatternFill("solid", fgColor="F5F5F5")
PROFIT_FILL = PatternFill("solid", fgColor="C6EFCE")
LOSS_FILL = PatternFill("solid", fgColor="FFC7CE")
WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")
BONUS_FILL = PatternFill("solid", fgColor="DDEBF7")
LOTS_FILL = PatternFill("solid", fgColor="E2F0D9")

THIN = Side(border_style="thin", color="BFBFBF")
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY_FMT = '"$"#,##0.00;[Red]-"$"#,##0.00'
PRICE_FMT = "#,##0.00"
LOT_FMT = "0.00"
RR_FMT = "0.00"


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _fmt_dt(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def _autosize(ws: Worksheet, min_width: int = 8, max_width: int = 24,
              header_rows: int = 1) -> None:
    """Approximate auto-fit by inspecting cell content lengths."""
    for column_cells in ws.columns:
        try:
            length = max(
                len(str(c.value)) if c.value is not None else 0
                for c in column_cells
            )
        except ValueError:
            length = min_width
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, length + 2))


def _write_header(ws: Worksheet, row: int, columns: list[str],
                  fills: list[PatternFill] | None = None) -> None:
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.fill = fills[c - 1] if fills else HEADER_FILL
        cell.font = HEADER_FONT if not fills else Font(bold=True, color="000000", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = GRID


def _apply_row_fill(ws: Worksheet, row: int, ncols: int, fill: PatternFill) -> None:
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = fill


def _style_money_cell(cell, value: float | None = None) -> None:
    cell.number_format = MONEY_FMT
    if value is None:
        return
    if value > 0:
        cell.font = Font(color="006100", bold=True)
    elif value < 0:
        cell.font = Font(color="9C0006", bold=True)


def _style_pct_cell(cell, value: float | None = None) -> None:
    if value is None:
        return
    if value > 0:
        cell.font = Font(color="006100", bold=True)
    elif value < 0:
        cell.font = Font(color="9C0006", bold=True)


def _fmt_rr_planned(r: float | None) -> str | None:
    """Planned setup R:R as a 1:N ratio (always positive), e.g. 1:2.4."""
    return f"1:{r:.1f}" if isinstance(r, (int, float)) else None


def _fmt_rr_realized(r: float | None) -> str | None:
    """Realized R as a ratio: a win is 1:N (e.g. 1:0.51), a loss is -N R (e.g. -1.0R)."""
    if not isinstance(r, (int, float)):
        return None
    return f"1:{r:.2f}" if r >= 0 else f"{r:.1f}R"


def _fmt_payoff(p: float | None) -> str:
    """Realized payoff (avg win / avg loss) as 1:N."""
    return f"1:{p:.2f}" if isinstance(p, (int, float)) else "-"


def _style_rr_planned_cell(cell) -> None:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")


def _style_rr_realized_cell(cell, value: float | None) -> None:
    cell.alignment = Alignment(horizontal="center")
    if value is None:
        return
    cell.font = Font(color="006100" if value >= 0 else "9C0006", bold=True)


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _kv_section(ws: Worksheet, row: int, title: str, pairs: list[tuple], span: int = 2) -> int:
    ws.cell(row=row, column=1, value=title).font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    row += 1
    for label, value in pairs:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    return row + 1


def _write_summary_sheet(ws: Worksheet, result: dict) -> None:
    ws.title = "Summary"
    cfg = result.get("config", {})

    ws["A1"] = "XAUUSD BACKTEST RESULTS"
    ws["A1"].font = Font(bold=True, size=14, color="305496")
    ws.merge_cells("A1:D1")

    row = 3
    row = _kv_section(ws, row, "Configuration", [
        ("Initial capital", f"${cfg.get('initial_capital', 0):,.2f}"),
        ("Sizing mode", cfg.get("sizing_mode")),
        ("Risk per signal", f"{cfg.get('risk_per_signal', 0) * 100:.2f}%"),
        ("Entry count", cfg.get("entry_count")),
        ("Entry ladder", cfg.get("entry_ladder")),
        ("Shared SL", cfg.get("shared_sl")),
        ("SL multiplier", cfg.get("sl_multiplier")),
        ("Final target", cfg.get("final_target")),
        ("Lock after TP1", cfg.get("lock_after_tp1")),
        ("Lock after TP2", cfg.get("lock_after_tp2")),
        ("Bonus / closed lot", f"${cfg.get('bonus_per_closed_lot', 0):,.2f}"),
        ("Chart start", result.get("chart_start")),
        ("Chart end", result.get("chart_end")),
    ])

    initial = cfg.get("initial_capital", 0) or 1
    final = result.get("final_equity", 0)
    return_pct = (final / initial - 1) * 100 if initial else 0.0
    overall_pairs = [
        ("Final equity", f"${final:,.2f}"),
        ("Net profit incl. bonus", f"${result.get('net_profit', 0):,.2f}"),
        ("Trading P&L", f"${result.get('trading_pnl', 0):,.2f}"),
        ("Closed-lot bonus", f"${result.get('bonus', 0):,.2f}"),
        ("Return", f"{return_pct:+.2f}%"),
        ("Max drawdown", f"{result.get('max_drawdown_pct', 0):.2f}%  "
                         f"(${abs(result.get('max_drawdown_usd', 0) or 0):,.2f})"),
        ("Signals included", result.get("signals_included")),
        ("Signal wins / losses", f"{result.get('wins', 0)} / {result.get('losses', 0)}"),
        ("Signal win rate", f"{result.get('win_rate_pct', 0):.2f}%"),
    ]
    # Drawdown trough context: WHEN the worst drawdown happened and how much had
    # executed by then (only when the engine recorded a trough).
    tr = result.get("drawdown_trough")
    if tr:
        overall_pairs.append(
            ("Max-DD trough",
             f"{tr.get('time_chart') or '?'}  "
             f"({tr.get('signals_executed_through', 0)} signals / "
             f"{tr.get('entries_filled_through', 0)} entries executed by then)"))
    # Hybrid (tick-preferred / M1-fallback) backtests carry a per-source split;
    # only shown when present, so pure M1/tick reports are unchanged.
    ds = result.get("data_sources")
    if ds:
        overall_pairs.append(
            ("Data source (signals)",
             f"TICK {ds.get('tick_signals', 0)} / M1 {ds.get('m1_signals', 0)}"))
    row = _kv_section(ws, row, "Overall Performance", overall_pairs)

    # Entry-outcome counts ("how many skipped / hit TP / SL / etc").
    ws.cell(row=row, column=1, value="Entry Outcomes").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    _write_header(ws, row, ["Outcome", "Count", "% of entries"])
    row += 1
    total_entries = result.get("entry_total", 0) or 0
    counts = result.get("entry_status_counts", {}) or {}
    for status in result.get("entry_statuses_present", []) or sorted(counts):
        n = counts.get(status, 0)
        pct = (n / total_entries * 100.0) if total_entries else 0.0
        ws.cell(row=row, column=1, value=status).border = GRID
        ws.cell(row=row, column=2, value=n).border = GRID
        c3 = ws.cell(row=row, column=3, value=f"{pct:.1f}%")
        c3.border = GRID
        fill = STATUS_FILL.get(status)
        if fill:
            for c in (1, 2, 3):
                ws.cell(row=row, column=c).fill = fill
        row += 1
    ws.cell(row=row, column=1, value="Total entries").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_entries).font = Font(bold=True)
    row += 2

    rr_avg = result.get("entry_rr_avg")
    rrp_avg = result.get("entry_rrp_avg")
    payoff = result.get("entry_payoff_ratio")
    row = _kv_section(ws, row, "Risk:Reward (executed)", [
        ("Filled entries", result.get("entry_filled", 0)),
        ("No-fill entries", result.get("entry_no_fill", 0)),
        ("Winning entries", result.get("entry_win_count", 0)),
        ("Losing entries", result.get("entry_loss_count", 0)),
        ("Avg planned R:R", _fmt_rr_planned(rrp_avg) or "-"),
        ("Avg realized R", _fmt_rr_realized(rr_avg) or "-"),
        ("Payoff (avg win / avg loss)", _fmt_payoff(payoff)),
    ])

    # TSL18 collision policies -- only when a policy was active (the result carries
    # a "collision_policy" block), so pure runs keep their exact Summary (parity).
    cp = result.get("collision_policy")
    if cp:
        ccfg = cp.get("config", {})
        row = _kv_section(ws, row, "TSL18 Collision Policies", [
            ("Opposite-side policy", ccfg.get("opposite_signal_policy")),
            ("Same-side overlap policy", ccfg.get("same_side_overlap_policy")),
            ("Opposite collisions (total)", cp.get("opposite_collisions_total", 0)),
            ("  allowed / rejected", f"{cp.get('opposite_collisions_allowed', 0)} / "
                                     f"{cp.get('opposite_collisions_rejected', 0)}"),
            ("  flipped / profit-banked",
             f"{cp.get('opposite_collisions_flipped', 0)} / "
             f"{cp.get('opposite_collisions_profit_bank_rearmed', 0)}"),
            ("Same-side clusters (total)", cp.get("same_side_clusters_total", 0)),
            ("  accepted / rejected / downsized",
             f"{cp.get('same_side_clusters_accepted', 0)} / "
             f"{cp.get('same_side_clusters_rejected', 0)} / "
             f"{cp.get('same_side_clusters_downsized', 0)}"),
            ("Max same-side cluster risk",
             f"${cp.get('max_same_side_cluster_risk', 0):,.2f}"),
            ("Max opposite exposure (lots)",
             f"{cp.get('max_opposite_exposure', 0):,.2f}"),
            ("Collision-policy P&L", f"${cp.get('collision_policy_pnl', 0):,.2f}"),
        ])

    # Monthly breakdown table.
    ws.cell(row=row, column=1, value="Monthly Breakdown").font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
    row += 1
    headers = ["Month", "Regime", "Signals", "Wins", "Losses", "No-fills",
               "Win rate", "Trading P&L", "Bonus", "Closed lots",
               "Net P&L", "P&L %", "Equity EoM", "Entries", "Drawdown %",
               "Drawdown $", "Worst DD at"]
    _write_header(ws, row, headers)
    row += 1
    for m in result.get("monthly", []) or []:
        pnl = m.get("pnl", 0.0) or 0.0
        pnl_pct = m.get("pnl_pct", 0.0) or 0.0
        signals = m.get("signals", 0) or 0
        equity_end = m.get("equity_end", 0.0) or 0.0
        dd_pct = m.get("max_drawdown_pct", 0.0) or 0.0
        dd_usd = m.get("max_drawdown_usd", 0.0) or 0.0
        cells = [
            m.get("month"), m.get("regime", "") or "-",
            signals, m.get("wins"), m.get("losses"), m.get("no_fills"),
            f"{m.get('win_rate_pct', 0):.1f}%",
            m.get("trading_pnl", 0.0), m.get("bonus", 0.0), m.get("closed_lots", 0.0),
            pnl, f"{pnl_pct:+.2f}%", equity_end,
            m.get("entries", 0) or 0, f"{dd_pct:.2f}%", dd_usd,
            m.get("max_drawdown_at") or "-",
        ]
        if signals == 0:
            _apply_row_fill(ws, row, len(headers), QUIET_DAY_FILL)
        elif pnl > 0:
            _apply_row_fill(ws, row, len(headers), PROFIT_FILL)
        elif pnl < 0:
            _apply_row_fill(ws, row, len(headers), LOSS_FILL)
        for c, v in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = GRID
            # Columns shifted +1 by the inserted Regime column (col 2):
            # Trading P&L=8, Bonus=9, Closed lots=10, Net P&L=11, P&L %=12, Equity=13.
            if c in (8, 9, 11):
                _style_money_cell(cell, float(v or 0.0) if isinstance(v, (int, float)) else None)
            elif c == 10:
                cell.number_format = LOT_FMT
            elif c == 12:
                _style_pct_cell(cell, pnl_pct)
            elif c == 13:
                # Equity at end of month, mirroring the Daily sheet's
                # Equity EoD column; no red-negative semantics, it's a level.
                cell.number_format = MONEY_FMT
            elif c == 15:                    # Drawdown % (worst during the month)
                _style_pct_cell(cell, dd_pct)
            elif c == 16:                    # Drawdown $ (matching the %)
                _style_money_cell(cell, dd_usd)
        row += 1

    ws.freeze_panes = "A2"
    _autosize(ws)


# ---------------------------------------------------------------------------
# Weekly (month-week) Breakdown sheet
# ---------------------------------------------------------------------------

def _write_weekly_sheet(ws: Worksheet, result: dict) -> None:
    """One row per feed-zone month-week (W1=days 1-7 .. W5=29-31). Mirrors the
    Summary's Monthly Breakdown columns (minus Regime, which is a month-level
    label) so the mid-grain view reads the same way."""
    ws.title = "Weekly Breakdown"
    headers = ["Month-Week", "Signals", "Wins", "Losses", "No-fills",
               "Win rate", "Trading P&L", "Bonus", "Closed lots",
               "Net P&L", "P&L %", "Equity EoW", "Entries", "Drawdown %",
               "Drawdown $", "Worst DD at"]
    _write_header(ws, 1, headers)

    weekly = result.get("weekly", []) or []
    for r_idx, w in enumerate(weekly, start=2):
        pnl = w.get("pnl", 0.0) or 0.0
        pnl_pct = w.get("pnl_pct", 0.0) or 0.0
        signals = w.get("signals", 0) or 0
        dd_pct = w.get("max_drawdown_pct", 0.0) or 0.0
        dd_usd = w.get("max_drawdown_usd", 0.0) or 0.0
        cells = [
            w.get("week"), signals, w.get("wins"), w.get("losses"), w.get("no_fills"),
            f"{w.get('win_rate_pct', 0):.1f}%",
            w.get("trading_pnl", 0.0), w.get("bonus", 0.0), w.get("closed_lots", 0.0),
            pnl, f"{pnl_pct:+.2f}%", w.get("equity_end", 0.0) or 0.0,
            w.get("entries", 0) or 0, f"{dd_pct:.2f}%", dd_usd,
            w.get("max_drawdown_at") or "-",
        ]
        if signals == 0:
            _apply_row_fill(ws, r_idx, len(headers), QUIET_DAY_FILL)
        elif pnl > 0:
            _apply_row_fill(ws, r_idx, len(headers), PROFIT_FILL)
        elif pnl < 0:
            _apply_row_fill(ws, r_idx, len(headers), LOSS_FILL)
        for c, v in enumerate(cells, start=1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.border = GRID
            if c in (7, 8, 10):              # Trading P&L, Bonus, Net P&L
                _style_money_cell(cell, float(v or 0.0) if isinstance(v, (int, float)) else None)
            elif c == 9:                     # Closed lots
                cell.number_format = LOT_FMT
            elif c == 11:                    # P&L %
                _style_pct_cell(cell, pnl_pct)
            elif c == 12:                    # Equity end-of-week (a level, no red)
                cell.number_format = MONEY_FMT
            elif c == 14:                    # Drawdown % (worst during the week)
                _style_pct_cell(cell, dd_pct)
            elif c == 15:                    # Drawdown $ (matching the %)
                _style_money_cell(cell, dd_usd)

    ws.freeze_panes = "A2"
    if weekly:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(weekly) + 1}"
    _autosize(ws)


# ---------------------------------------------------------------------------
# Daily Breakdown sheet
# ---------------------------------------------------------------------------

def _write_daily_sheet(ws: Worksheet, result: dict) -> None:
    ws.title = "Daily Breakdown"
    statuses = result.get("entry_statuses_present", []) or []
    base = ["Date", "Signals", "Wins", "Losses"]
    tail = ["Entries", "Avg R:R", "Avg R", "Trading P&L", "Bonus", "Closed lots",
            "Net P&L", "P&L %", "Equity EoD", "Drawdown %", "Drawdown $", "Worst DD at"]
    headers = base + statuses + tail
    _write_header(ws, 1, headers)

    n_base = len(base)
    n_status = len(statuses)
    daily = result.get("daily", []) or []
    peak_equity = (result.get("config", {}) or {}).get("initial_capital", 0) or 0

    for r_idx, d in enumerate(daily, start=2):
        signals = d.get("signals", 0) or 0
        pnl = d.get("pnl", 0.0) or 0.0
        trading_pnl = d.get("trading_pnl", 0.0) or 0.0
        bonus = d.get("bonus", 0.0) or 0.0
        closed_lots = d.get("closed_lots", 0.0) or 0.0
        equity_end = d.get("equity_end", 0.0) or 0.0
        if equity_end > peak_equity:
            peak_equity = equity_end
        drawdown_pct = (equity_end - peak_equity) / peak_equity * 100.0 if peak_equity else 0.0
        drawdown_usd = equity_end - peak_equity        # USD alongside %, matches the % shown
        pnl_pct = d.get("pnl_pct", 0.0) or 0.0
        rr_avg = d.get("entry_rr_avg")
        rrp_avg = d.get("entry_rrp_avg")
        status_counts = d.get("entry_status_counts", {}) or {}

        cells = [d.get("date"), signals, d.get("wins", 0), d.get("losses", 0)]
        cells += [status_counts.get(s, 0) for s in statuses]
        cells += [
            d.get("entry_total", 0),
            _fmt_rr_planned(rrp_avg), _fmt_rr_realized(rr_avg),
            trading_pnl, bonus, closed_lots, pnl,
            f"{pnl_pct:+.2f}%", equity_end, f"{drawdown_pct:.2f}%", drawdown_usd,
            d.get("max_drawdown_at") or "-",
        ]

        if signals == 0:
            _apply_row_fill(ws, r_idx, len(headers), QUIET_DAY_FILL)
        elif drawdown_pct <= -20.0:
            _apply_row_fill(ws, r_idx, len(headers), WARNING_FILL)
        elif pnl > 0:
            _apply_row_fill(ws, r_idx, len(headers), PROFIT_FILL)
        elif pnl < 0:
            _apply_row_fill(ws, r_idx, len(headers), LOSS_FILL)

        entries_col = n_base + n_status + 1        # Entries, then Avg R:R, Avg R
        rrp_col = entries_col + 1
        rrr_col = entries_col + 2
        money_cols = {entries_col + 3, entries_col + 4, entries_col + 6}  # Trading P&L, Bonus, Net P&L
        lots_col = entries_col + 5
        pnlpct_col = entries_col + 7
        eq_col = entries_col + 8
        dd_col = entries_col + 9
        dd_usd_col = entries_col + 10
        for c_idx, v in enumerate(cells, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.border = GRID
            if c_idx == rrp_col:
                _style_rr_planned_cell(cell)
            elif c_idx == rrr_col:
                _style_rr_realized_cell(cell, rr_avg)
            elif c_idx in money_cols:
                _style_money_cell(cell, float(v or 0.0) if isinstance(v, (int, float)) else None)
            elif c_idx == lots_col:
                cell.number_format = LOT_FMT
            elif c_idx == eq_col:
                cell.number_format = MONEY_FMT
            elif c_idx == dd_usd_col:
                _style_money_cell(cell, drawdown_usd)
            elif c_idx in (pnlpct_col, dd_col):
                _style_pct_cell(cell, pnl_pct if c_idx == pnlpct_col else drawdown_pct)

    ws.freeze_panes = "A2"
    if daily:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(daily) + 1}"
    _autosize(ws)


# ---------------------------------------------------------------------------
# Per-Entry Detail sheet
# ---------------------------------------------------------------------------

# (key, header, kind, group). group: "id" | "orig" (from signal) | "exec" (result).
ENTRY_LAYOUT = [
    ("entry_key", "Entry Key", None, "id"),
    ("signal_date", "Date", None, "id"),
    ("signal_time_source", "Time (src)", None, "id"),     # header set per source tz
    ("signal_time_chart", "Time (chart EET/EEST)", "datetime", "id"),
    ("side", "Side", None, "id"),
    ("range_low", "Range Low", "price", "orig"),
    ("range_high", "Range High", "price", "orig"),
    ("original_SL", "Orig SL", "price", "orig"),
    ("TP1", "TP1", "price", "orig"),
    ("TP2", "TP2", "price", "orig"),
    ("TP3", "TP3", "price", "orig"),
    ("final_target_price", "Target", "price", "exec"),
    ("entry_price", "Entry Price", "price", "exec"),
    ("effective_SL", "Effective SL", "price", "exec"),
    ("lot", "Lot", "lot", "exec"),
    ("entry_status", "Status", None, "exec"),
    ("fill_time", "Fill Time", "datetime", "exec"),
    ("exit_time", "Exit Time", "datetime", "exec"),
    ("exit_price", "Exit Price", "price", "exec"),
    ("trading_pnl", "P&L ($)", "money", "exec"),
    ("rr_planned", "R:R", "rrp", "exec"),
    ("rr", "R", "rrr", "exec"),
]

# Appended only on a TSL18 collision-policy run (any row stamped with a
# collision_type), so pure runs keep their exact column set (parity).
COLLISION_LAYOUT = [
    ("collision_type", "Collision", None),
    ("collision_policy", "Policy", None),
    ("collision_policy_action", "Action", None),
    ("cluster_id", "Cluster", None),
    ("cluster_risk_before", "Cluster Risk Before", "money"),
    ("cluster_risk_after", "Cluster Risk After", "money"),
    ("opposite_exposure_before", "Opp Exp Before", "lot"),
    ("opposite_exposure_after", "Opp Exp After", "lot"),
]

# Appended only when an MT5 history was merged. Each live column that mirrors a
# plan column names it for diff-highlighting (mismatch -> orange).
LIVE_LAYOUT = [
    ("live_entry", "Live Entry", "price", "live", "entry_price"),
    ("live_sl", "Live SL", "price", "live", "effective_SL"),
    ("live_exit", "Live Exit", "price", "live", "exit_price"),
    ("live_exit_time", "Live Exit Time", None, "live", None),
    ("live_pnl", "Live P&L", "money", "live", None),
    ("live_rr", "Live R", "rrr", "live", None),
]


def _source_time_header(rows: list[dict]) -> str:
    tzs = {r.get("source_tz") for r in rows if r.get("source_tz")}
    return f"Time ({tzs.pop()})" if len(tzs) == 1 else "Time (src)"


def _write_entries_sheet(ws: Worksheet, result: dict) -> None:
    ws.title = "Per-Entry Detail"
    rows = result.get("entry_rows", []) or []
    has_live = bool(result.get("has_live"))

    # (key, header, kind, group, compare_key) — compare_key only used by live cols.
    layout = [(k, h, kind, g, None) for k, h, kind, g in ENTRY_LAYOUT]
    if has_live:
        layout += list(LIVE_LAYOUT)
    # Dynamic source-time header (e.g. "Time (GMT+7)").
    layout[2] = (layout[2][0], _source_time_header(rows), layout[2][2], layout[2][3], None)
    # Hybrid backtests tag each row TICK/M1; add a SIGNAL-group column ONLY when
    # present so pure M1/tick reports keep their exact column set (parity).
    if any(r.get("data_source") for r in rows):
        side_i = next((i for i, (k, *_) in enumerate(layout) if k == "side"), 4)
        layout.insert(side_i + 1, ("data_source", "Data Source", None, "id", None))
    # Collision-policy run: append a COLLISION column group ONLY when a collision
    # policy stamped the rows, so pure runs keep their exact column set (parity).
    if any(r.get("collision_type") for r in rows):
        layout += [(k, h, kind, "collision", None) for k, h, kind in COLLISION_LAYOUT]

    groups = [g for _, _, _, g, _ in layout]
    headers = [h for _, h, _, _, _ in layout]

    def _span(group: str) -> tuple[int, int]:
        idx = [i for i, g in enumerate(groups) if g == group]
        return idx[0] + 1, idx[-1] + 1

    # Row 1: group banner.
    banner = {
        "id": ("SIGNAL", ID_BANNER_FILL),
        "orig": ("ORIGINAL  (from the signal)", ORIG_BANNER_FILL),
        "exec": ("EXECUTED  (backtest result)", EXEC_BANNER_FILL),
        "collision": ("COLLISION POLICY", SUBHEADER_FILL),
        "live": ("LIVE  (MT5 execution)", LIVE_BANNER_FILL),
    }
    for group, (text, fill) in banner.items():
        if group not in groups:
            continue
        c0, c1 = _span(group)
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
        cell = ws.cell(row=1, column=c0, value=text)
        cell.font = BANNER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(c0, c1 + 1):
            ws.cell(row=1, column=c).fill = fill
            ws.cell(row=1, column=c).border = GRID

    # Row 2: column headers, tinted by group.
    group_header_fill = {
        "orig": ORIG_HEADER_FILL, "exec": EXEC_HEADER_FILL, "live": LIVE_HEADER_FILL,
    }
    header_fills = [group_header_fill.get(g, SUBHEADER_FILL) for g in groups]
    _write_header(ws, 2, headers, fills=header_fills)

    # Data rows.
    for r_idx, row in enumerate(rows, start=3):
        status = row.get("entry_status", "")
        exec_fill = STATUS_FILL.get(status)
        for c_idx, (key, _label, kind, group, compare_key) in enumerate(layout, start=1):
            raw = row.get(key)
            num = raw if isinstance(raw, (int, float)) else None
            value = raw
            if kind == "datetime":
                value = _fmt_dt(raw)
            elif kind == "rrp":
                value = _fmt_rr_planned(raw)
            elif kind == "rrr":
                value = _fmt_rr_realized(raw)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = GRID
            if kind == "money":
                _style_money_cell(cell, num)
            elif kind == "price":
                cell.number_format = PRICE_FMT
            elif kind == "lot":
                cell.number_format = LOT_FMT
            elif kind == "rrp":
                _style_rr_planned_cell(cell)
            elif kind == "rrr":
                _style_rr_realized_cell(cell, num)
            # Group fills: ORIGINAL columns get a constant tint so the signal's
            # stated levels stand out; EXECUTED columns get the outcome colour;
            # LIVE columns get a constant tint, with mismatches vs the plan flagged.
            if group == "orig":
                cell.fill = ORIGINAL_CELL_FILL
            elif group == "exec" and exec_fill is not None:
                cell.fill = exec_fill
            elif group == "live":
                cell.fill = LIVE_CELL_FILL
                if compare_key is not None:
                    plan = row.get(compare_key)
                    if (isinstance(value, (int, float)) and isinstance(plan, (int, float))
                            and abs(value - plan) > LIVE_DIFF_TOL):
                        cell.fill = LIVE_DIFF_FILL
                        cell.font = Font(bold=True, color="7030A0")

    ws.freeze_panes = "A3"
    if rows:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(rows) + 2}"
    _autosize(ws)


# ---------------------------------------------------------------------------
# entry point
# ---------------------------------------------------------------------------

def write_excel_report(result: dict, output_path: Path | str) -> Path:
    """Render the backtest result to a styled .xlsx file."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _write_summary_sheet(wb.active, result)
    _write_weekly_sheet(wb.create_sheet("Weekly Breakdown"), result)
    _write_daily_sheet(wb.create_sheet("Daily Breakdown"), result)
    _write_entries_sheet(wb.create_sheet("Per-Entry Detail"), result)

    wb.save(output_path)
    return output_path
