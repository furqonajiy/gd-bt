"""Parse a MetaTrader 5 history export and attach the real fills to a backtest.

MT5's History tab ("Positions" view) exports one row per position with both the
open and close on the same line, e.g.:

    Time | Ticket | Type | Volume | Symbol | Price | S/L | T/P | Time | Price |
    Profit | ... | Comment

The executor writes each entry's key (``2026-06-08#02.1``) into the order
Comment, so we match live positions to backtest entries by that Comment. The
parser is deliberately tolerant: it locates the deals table by its header row,
maps columns by name (case/spacing-insensitive), and resolves the duplicated
``Time``/``Price`` columns as first = open, last = close. XLSX, CSV and HTML
exports are supported.

Non-trade rows (balance/credit/deposit adjustments) carry comments that don't
match any entry_key, so they're simply ignored at match time.
"""
from __future__ import annotations

import csv as _csv
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Optional


# Header synonyms (normalized: lowercased, whitespace removed -> "s / l" == "s/l").
_COMMENT_KEYS = {"comment"}
_PROFIT_KEYS = {"profit"}
_SL_KEYS = {"s/l", "sl", "stoploss"}
_TP_KEYS = {"t/p", "tp", "takeprofit"}
_TYPE_KEYS = {"type"}
_VOL_KEYS = {"volume", "lots", "size", "vol"}
_PRICE_KEYS = {"price"}
_TIME_KEYS = {"time", "opentime", "closetime"}


def _norm(value: Any) -> str:
    return "".join(str(value or "").split()).lower()


def _num(value: Any) -> Optional[float]:
    """Parse a numeric cell tolerating commas/spaces as thousands separators."""
    if value is None:
        return None
    s = str(value).strip().replace(",", "").replace(" ", "")
    if s in ("", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _live_rr(side: str, entry: Optional[float], sl: Optional[float],
             exit_price: Optional[float]) -> Optional[float]:
    """Realized R of the live fill (side-aware), or None if not computable."""
    if entry is None or sl is None or exit_price is None:
        return None
    risk = abs(entry - sl)
    if risk <= 0:
        return None
    favourable = (exit_price - entry) if side == "BUY" else (entry - exit_price)
    return favourable / risk


# ---------------------------------------------------------------------------
# raw-row readers (one list-of-cells per source format)
# ---------------------------------------------------------------------------

def _read_xlsx_rows(path: Path) -> list[list[Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[list[Any]] = []
    for ws in wb.worksheets:
        for r in ws.iter_rows(values_only=True):
            rows.append(list(r))
    return rows


def _read_export_text(path: Path) -> str:
    """Decode an MT5 export, sniffing the encoding from the BOM.

    The terminal's "Report History" HTML export is UTF-16-LE with a BOM; older
    exports and CSVs are UTF-8. Decoding UTF-16 bytes as UTF-8 yields NUL-laced
    garbage that the HTML parser walks without finding a single tag, so the
    parse used to fail silently with zero rows — sniff instead of assuming.
    """
    raw = Path(path).read_bytes()
    if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        return raw.decode("utf-16", errors="replace")
    return raw.decode("utf-8-sig", errors="replace")


def _read_csv_rows(path: Path) -> list[list[Any]]:
    # MT5 CSV exports are often tab- or semicolon-delimited; sniff, default comma.
    text = _read_export_text(path)
    sample = text[:4096]
    delim = ","
    for cand in ("\t", ";", ","):
        if cand in sample:
            delim = cand
            break
    return [row for row in _csv.reader(text.splitlines(), delimiter=delim)]


class _TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._row: Optional[list[str]] = None
        self._cell: Optional[list[str]] = None

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th"):
            self._cell = []

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in ("td", "th") and self._cell is not None and self._row is not None:
            self._row.append("".join(self._cell).strip())
            self._cell = None
        elif tag == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None


def _read_html_rows(path: Path) -> list[list[Any]]:
    parser = _TableParser()
    parser.feed(_read_export_text(path))
    return parser.rows


def _read_rows(path: Path) -> list[list[Any]]:
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx_rows(path)
    if ext in (".htm", ".html"):
        return _read_html_rows(path)
    if ext in (".csv", ".txt", ".tsv"):
        return _read_csv_rows(path)
    raise ValueError(f"unsupported MT5 history format: {ext} (use .xlsx/.csv/.html)")


# ---------------------------------------------------------------------------
# table extraction
# ---------------------------------------------------------------------------

def _find_header(rows: list[list[Any]]) -> Optional[int]:
    """Index of the deals header row (has Comment + Profit + Price), else None."""
    for i, row in enumerate(rows):
        cells = {_norm(c) for c in row}
        if _COMMENT_KEYS & cells and _PROFIT_KEYS & cells and _PRICE_KEYS & cells:
            return i
    return None


def parse_mt5_history(path: str | Path) -> dict[str, dict]:
    """Return {comment: {entry, sl, tp, exit, exit_time, profit, volume, type}}.

    ``comment`` is the raw order comment (the engine writes the entry_key there).
    Raises ValueError if no recognizable deals table is found.
    """
    path = Path(path)
    rows = _read_rows(path)
    header_idx = _find_header(rows)
    if header_idx is None:
        raise ValueError(
            f"could not find an MT5 deals table in {path.name} "
            f"(expected a header row containing Comment, Price and Profit). "
            f"Export the History tab with the Comment column visible."
        )

    norm = [_norm(c) for c in rows[header_idx]]

    def idxs(keys: set[str]) -> list[int]:
        return [i for i, h in enumerate(norm) if h in keys]

    comment_i = (idxs(_COMMENT_KEYS) or [None])[0]
    profit_i = (idxs(_PROFIT_KEYS) or [None])[-1]
    sl_i = (idxs(_SL_KEYS) or [None])[0]
    tp_i = (idxs(_TP_KEYS) or [None])[0]
    type_i = (idxs(_TYPE_KEYS) or [None])[0]
    vol_i = (idxs(_VOL_KEYS) or [None])[0]
    price_is = idxs(_PRICE_KEYS)
    time_is = idxs(_TIME_KEYS)
    open_price_i = price_is[0] if price_is else None
    close_price_i = price_is[-1] if len(price_is) > 1 else None
    close_time_i = time_is[-1] if len(time_is) > 1 else (time_is[0] if time_is else None)

    def cell(row: list[Any], i: Optional[int]) -> Any:
        return row[i] if (i is not None and i < len(row)) else None

    out: dict[str, dict] = {}
    for row in rows[header_idx + 1:]:
        comment = str(cell(row, comment_i) or "").strip()
        if not comment:
            continue
        entry = _num(cell(row, open_price_i))
        if entry is None:
            continue  # not a trade row
        exit_price = _num(cell(row, close_price_i)) if close_price_i is not None else None
        rec = {
            "entry": entry,
            "sl": _num(cell(row, sl_i)),
            "tp": _num(cell(row, tp_i)),
            "exit": exit_price,
            "exit_time": (str(cell(row, close_time_i)).strip() or None) if close_time_i is not None else None,
            "profit": _num(cell(row, profit_i)),
            "volume": _num(cell(row, vol_i)),
            "type": str(cell(row, type_i) or "").strip().upper() or None,
        }
        if comment in out:
            # Same comment seen twice (deal-by-deal export): keep first open, last
            # close, and sum the profit.
            prev = out[comment]
            prev["exit"] = rec["exit"] if rec["exit"] is not None else prev["exit"]
            prev["exit_time"] = rec["exit_time"] or prev["exit_time"]
            prev["profit"] = (prev["profit"] or 0.0) + (rec["profit"] or 0.0)
        else:
            out[comment] = rec
    return out


# ---------------------------------------------------------------------------
# attach to a backtest result
# ---------------------------------------------------------------------------

def _mt5_comment(signal_key: str, entry_index: int, max_len: int = 31) -> str:
    """MT5-safe per-entry comment (delegates to the executor's canonical builder
    so live-history matching uses the exact compact ``[TAG]#DD.N`` form that the
    executor stamped on the order)."""
    from trading.engine.execution.mt5_executor import mt5_entry_comment
    return mt5_entry_comment(signal_key, entry_index, max_len)


def attach_live_history(result: dict, live: dict[str, dict]) -> dict:
    """Merge parsed live fills into result['entry_rows'] by Comment/entry_key.

    Adds live_entry / live_sl / live_exit / live_exit_time / live_pnl / live_rr to
    each matched entry row, sets result['has_live'] = True, and returns a small
    match summary. Live risk uses the live SL when present, else the plan's SL.
    """
    matched = 0
    used_keys: set[str] = set()
    for er in result.get("entry_rows", []) or []:
        sig_key = er.get("signal_key", "")
        idx = er.get("entry_index", 0)
        candidates = [er.get("entry_key"), _mt5_comment(sig_key, idx)]
        lv = next((live[c] for c in candidates if c in live), None)
        if lv is None:
            continue
        matched += 1
        used_keys.update(c for c in candidates if c in live)
        sl_for_r = lv["sl"] if lv["sl"] is not None else er.get("effective_SL")
        er["live_entry"] = lv["entry"]
        er["live_sl"] = lv["sl"]
        er["live_exit"] = lv["exit"]
        er["live_exit_time"] = lv["exit_time"]
        er["live_pnl"] = lv["profit"]
        er["live_rr"] = _live_rr(er.get("side", ""), lv["entry"], sl_for_r, lv["exit"])

    result["has_live"] = matched > 0
    result["live_matched"] = matched
    result["live_unmatched"] = sorted(k for k in live if k not in used_keys)
    return {
        "matched": matched,
        "unmatched": result["live_unmatched"],
        "live_total": len(live),
    }
