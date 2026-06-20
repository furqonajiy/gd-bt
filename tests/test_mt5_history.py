"""Parse an MT5 history export and merge real fills into a backtest result.

Covers the tolerant table detection (duplicate Time/Price columns, "S / L"
spacing, comma/space thousands), CSV + XLSX readers, comment->entry_key match,
side-aware live R, and that the Per-Entry sheet gains a LIVE group.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from trading.engine.reporting.mt5_history import (
    attach_live_history, parse_mt5_history,
)
from trading.engine.reporting.excel_report import write_excel_report


# MT5 "Positions" history layout: Time/Price appear twice (open then close).
_HEADER = ["Time", "Ticket", "Type", "Volume", "Symbol", "Price", "S / L",
           "T / P", "Time", "Price", "Profit", "Comment"]
_DEALS = [
    ["2026.06.08 10:24:36", "3494230238", "sell", "0.01", "xauusd", "4314.36",
     "4329.75", "4276.00", "2026.06.08 10:39:07", "4307.35", "7.01", "2026-06-08#02.1"],
    ["2026.06.08 10:25:09", "3494230260", "sell", "0.01", "xauusd", "4315.05",
     "4330.75", "4276.00", "2026.06.08 10:39:07", "4307.35", "7.70", "2026-06-08#02.2"],
    # comma + space thousands separators must still parse.
    ["2026.06.08 10:25:45", "3494230279", "sell", "0.01", "xauusd", "4,316.56",
     "4331.75", "4276.00", "2026.06.08 10:39:07", "4 307.38", "9.18", "2026-06-08#02.3"],
    # a non-trade balance row: ignored at match time.
    ["2026.06.08 16:45:28", "3494130365", "balance", "", "", "", "", "",
     "", "", "1126.32", "Z/negative balance adjustment"],
]


def _write_csv(tmp_path) -> Path:
    # MT5 CSV exports are tab-delimited, so comma/space thousands stay intact.
    p = tmp_path / "history.csv"
    lines = ["\t".join(_HEADER)] + ["\t".join(c) for c in _DEALS]
    p.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return p


def _write_xlsx(tmp_path) -> Path:
    p = tmp_path / "history.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Account: 123  Broker: X"])     # noise header line above the table
    ws.append(_HEADER)
    for d in _DEALS:
        ws.append(d)
    wb.save(p)
    return p


def test_parse_csv_resolves_open_close_and_skips_noise(tmp_path):
    live = parse_mt5_history(_write_csv(tmp_path))
    assert set(live) == {"2026-06-08#02.1", "2026-06-08#02.2", "2026-06-08#02.3"}
    one = live["2026-06-08#02.1"]
    assert one["entry"] == 4314.36 and one["sl"] == 4329.75
    assert one["exit"] == 4307.35 and one["profit"] == 7.01
    assert one["type"] == "SELL"
    # thousands separators
    assert live["2026-06-08#02.3"]["entry"] == 4316.56
    assert live["2026-06-08#02.3"]["exit"] == 4307.38


def test_parse_xlsx_finds_table_under_noise(tmp_path):
    live = parse_mt5_history(_write_xlsx(tmp_path))
    assert live["2026-06-08#02.2"]["entry"] == 4315.05
    assert live["2026-06-08#02.2"]["exit"] == 4307.35


def _fake_result():
    rows = []
    for i, (entry, sl) in enumerate([(4314.0, 4329.75), (4315.0, 4330.75),
                                     (4316.0, 4331.75), (4317.0, 4332.75)]):
        rows.append({
            "signal_key": "2026-06-08#02", "entry_index": i,
            "entry_key": f"2026-06-08#02.{i + 1}", "side": "SELL",
            "entry_price": entry, "effective_SL": sl, "exit_price": 4306.0,
            "entry_status": "LOCK_TP1", "trading_pnl": (entry - 4306.0) * 0.01 * 100,
            "rr": (entry - 4306.0) / abs(entry - sl),
        })
    return {"entry_rows": rows}


def test_attach_matches_and_computes_live_rr(tmp_path):
    result = _fake_result()
    live = parse_mt5_history(_write_csv(tmp_path))
    info = attach_live_history(result, live)

    assert result["has_live"] is True
    assert info["matched"] == 3                      # #02.4 has no live row
    er = result["entry_rows"][0]
    assert er["live_entry"] == 4314.36 and er["live_exit"] == 4307.35
    assert er["live_pnl"] == 7.01
    # SELL win -> +R (favourable move / risk distance), not negative.
    assert round(er["live_rr"], 4) == round((4314.36 - 4307.35) / abs(4314.36 - 4329.75), 4)
    # the unmatched live total is 0 here (all 3 live rows matched)
    assert info["unmatched"] == []
    # entry without a live row gets no live fields.
    assert "live_entry" not in result["entry_rows"][3]


def test_report_adds_live_group_when_present(tmp_path):
    result = _fake_result()
    attach_live_history(result, parse_mt5_history(_write_csv(tmp_path)))
    # minimal extra keys the writer reads
    result.update({"config": {}, "monthly": [], "daily": [], "entry_statuses_present": []})
    out = write_excel_report(result, tmp_path / "r.xlsx")
    wb = load_workbook(out)
    ws = wb["Per-Entry Detail"]
    banner = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert any(v and "LIVE" in str(v) for v in banner)
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    assert "Live Entry" in headers and "Live R" in headers
