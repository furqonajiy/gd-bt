"""Report data layer: realized R per entry, entry-outcome aggregates, and the
Daily Breakdown trimmed to the traded window (no pre-start padding).

Deterministic: a tiny synthetic M1 chart, so it runs everywhere.
"""
from __future__ import annotations

from dataclasses import replace

from trading.engine import CsvChartSource, DEFAULT_CONFIG, parse_one_signal, run_backtest
from trading.engine.strategy.backtest import _payoff_ratio, _planned_rr, _realized_rr
from trading.engine.reporting.excel_report import (
    _fmt_payoff, _fmt_rr_planned, _fmt_rr_realized,
)


# --- pure helpers -----------------------------------------------------------

def test_realized_rr_is_side_aware():
    # BUY win: up 10 on 5 risk -> +2R; BUY loss: down 5 -> -1R.
    assert _realized_rr("BUY", 100.0, 95.0, 110.0, filled=True) == 2.0
    assert _realized_rr("BUY", 100.0, 95.0, 95.0, filled=True) == -1.0
    # SELL win: price falls 8 on 15.75 risk -> +R (positive, not negative).
    assert round(_realized_rr("SELL", 4314.0, 4329.75, 4306.0, filled=True), 4) == round(8 / 15.75, 4)
    assert _realized_rr("SELL", 4314.0, 4329.75, 4329.75, filled=True) == -1.0


def test_realized_rr_none_when_not_filled_or_no_risk():
    assert _realized_rr("BUY", 100.0, 95.0, None, filled=True) is None     # not closed
    assert _realized_rr("BUY", 100.0, 95.0, 110.0, filled=False) is None   # not filled
    assert _realized_rr("BUY", 100.0, 100.0, 110.0, filled=True) is None   # zero risk


def test_payoff_ratio_avg_win_over_avg_loss():
    assert _payoff_ratio([10.0, 20.0], [-5.0, -5.0]) == 3.0   # avg win 15 / avg loss 5
    assert _payoff_ratio([10.0], []) is None
    assert _payoff_ratio([], [-5.0]) is None


def test_planned_rr_reward_over_risk():
    # reward 10 (100->110) over risk 5 (100->95) = 2.0; positive regardless of side.
    assert _planned_rr(100.0, 95.0, 110.0) == 2.0
    assert _planned_rr(4314.0, 4329.75, 4276.0) == 38.0 / 15.75
    assert _planned_rr(100.0, 100.0, 110.0) is None     # zero risk
    assert _planned_rr(100.0, 95.0, None) is None


def test_rr_formats_as_ratio():
    # planned -> 1:N (one decimal)
    assert _fmt_rr_planned(2.41) == "1:2.4"
    assert _fmt_rr_planned(None) is None
    # realized -> win 1:N (two decimals), loss -N R
    assert _fmt_rr_realized(0.51) == "1:0.51"
    assert _fmt_rr_realized(-1.0) == "-1.0R"
    assert _fmt_rr_realized(None) is None
    # payoff -> 1:N, or "-" when undefined
    assert _fmt_payoff(1.5) == "1:1.50"
    assert _fmt_payoff(None) == "-"


# --- end-to-end through run_backtest ----------------------------------------

_HEADER = "<DATE>\t<TIME>\t<OPEN>\t<HIGH>\t<LOW>\t<CLOSE>\t<TICKVOL>\t<VOL>\t<SPREAD>"


def _bar(d, t, o, h, l, c):
    return f"{d}\t{t}\t{o}\t{h}\t{l}\t{c}\t100.0\t0.0\t2"


def _chart(tmp_path):
    rows = [
        # 2026-06-01: pre-signal day (must be trimmed from Daily Breakdown).
        _bar("2026.06.01", "10:00:00", 100, 100.5, 99.5, 100),
        _bar("2026.06.01", "10:01:00", 100, 100.5, 99.5, 100),
        # 2026-06-02: signal at 11:00, fills ~100, runs to TP1 110.
        _bar("2026.06.02", "11:00:00", 100, 100.2, 99.9, 100),
        _bar("2026.06.02", "11:01:00", 100, 101.0, 98.0, 100),     # fill the BUY
        _bar("2026.06.02", "11:02:00", 100, 103.0, 100.0, 102),
        _bar("2026.06.02", "11:05:00", 105, 111.0, 104.0, 110),    # hit TP1 110
        # 2026-06-03: post-signal day (also outside the traded window).
        _bar("2026.06.03", "09:00:00", 110, 110.5, 109.5, 110),
    ]
    p = tmp_path / "XAUUSD_M1_TEST.csv"
    p.write_text("\n".join([_HEADER, *rows]) + "\n", encoding="utf-8")
    return CsvChartSource([p])


def _config():
    return replace(
        DEFAULT_CONFIG, entry_count=1, entry_ladder="range_uniform",
        sl_multiplier=1.0, activation_delay_minutes=0, pending_expiry_minutes=60,
        max_hold_minutes=120, lock_after_tp1=False, final_target="TP1",
    )


def test_run_backtest_emits_entry_rr_and_outcome_aggregates(tmp_path):
    sig = parse_one_signal(
        "1. BUY XAUUSD 100 - 100 SL 95 TP1 110 TP2 115 TP3 120 11:00 AM",
        source_date="2026-06-02", source_offset=3,
    )
    result = run_backtest([sig], _chart(tmp_path), _config())

    # Every entry row carries a realized-R field.
    assert all("rr" in er for er in result["entry_rows"])
    # New summary aggregates exist.
    for key in ("entry_total", "entry_status_counts", "entry_statuses_present",
                "entry_filled", "entry_rr_avg", "entry_payoff_ratio"):
        assert key in result
    # The single entry filled and hit TP1 -> +2R realized; planned R:R 110/95 = 2.0.
    er = result["entry_rows"][0]
    assert er["entry_status"] == "TP1"
    assert er["rr"] == 2.0
    assert er["rr_planned"] == 2.0
    assert result["entry_rrp_avg"] == 2.0
    assert result["entry_status_counts"]["TP1"] == 1


def test_daily_breakdown_trims_pre_and_post_signal_days(tmp_path):
    sig = parse_one_signal(
        "1. BUY XAUUSD 100 - 100 SL 95 TP1 110 TP2 115 TP3 120 11:00 AM",
        source_date="2026-06-02", source_offset=3,
    )
    result = run_backtest([sig], _chart(tmp_path), _config())
    dates = [d["date"] for d in result["daily"]]
    # Chart spans 06-01..06-03 but the only signal is 06-02, so that's the only day.
    assert dates == ["2026-06-02"]
    day = result["daily"][0]
    assert day["entry_total"] == 1
    assert day["entry_status_counts"].get("TP1") == 1
    assert day["entry_rr_avg"] == 2.0


def _chart_gmt7(tmp_path):
    # A GMT+7 signal at 2026-07-01 01:00 maps to chart-local 2026-06-30 21:00
    # (summer EEST = UTC+3): 01:00 - 7h + 3h. So its chart day/month (06-30 /
    # 2026-06) differ from its feed day/month (07-01 / 2026-07).
    rows = [
        _bar("2026.06.30", "21:00:00", 100, 100.2, 99.9, 100),
        _bar("2026.06.30", "21:01:00", 100, 101.0, 98.0, 100),   # fill the BUY ~100
        _bar("2026.06.30", "21:05:00", 105, 111.0, 104.0, 110),  # hit TP1 110
    ]
    p = tmp_path / "XAUUSD_M1_TEST.csv"
    p.write_text("\n".join([_HEADER, *rows]) + "\n", encoding="utf-8")
    return CsvChartSource([p])


def test_daily_and_monthly_breakdown_use_signal_feed_zone_date(tmp_path):
    # Regression: the Daily/Monthly breakdowns must group by the signal's own
    # feed-zone (GMT+7) date, so a report day lines up with the signal codes
    # (e.g. SQZ6-0623) -- not the chart (EET/EEST) day, which can be the day
    # before for an early-morning GMT+7 signal.
    sig = parse_one_signal(
        "1. BUY XAUUSD 100 - 100 SL 95 TP1 110 TP2 115 TP3 120 1:00 AM",
        source_date="2026-07-01", source_offset=7,
    )
    assert sig.signal_time_chart.date().isoformat() == "2026-06-30"   # chart day
    assert sig.signal_time_source.date().isoformat() == "2026-07-01"  # feed day

    result = run_backtest([sig], _chart_gmt7(tmp_path), _config())

    # The single TP1 win is bucketed on the FEED day/month, not the chart's.
    assert result["entry_rows"][0]["entry_status"] == "TP1"
    assert [d["date"] for d in result["daily"]] == ["2026-07-01"]
    assert [m["month"] for m in result["monthly"]] == ["2026-07"]
    day = result["daily"][0]
    assert day["signals"] == 1 and day["wins"] == 1
    assert day["entry_total"] == 1
    assert day["entry_status_counts"].get("TP1") == 1


def test_summary_monthly_breakdown_shows_equity_end_of_month(tmp_path):
    # The Monthly Breakdown table must show the equity level at month end
    # (mirroring the Daily sheet's Equity EoD), not just the month's P&L.
    from openpyxl import load_workbook

    from trading.engine import write_backtest_outputs

    sig = parse_one_signal(
        "1. BUY XAUUSD 100 - 100 SL 95 TP1 110 TP2 115 TP3 120 11:00 AM",
        source_date="2026-06-02", source_offset=3,
    )
    result = run_backtest([sig], _chart(tmp_path), _config())
    monthly = result["monthly"]
    assert monthly and monthly[0]["equity_end"] == result["final_equity"]

    path = write_backtest_outputs(result, tmp_path / "reports" / "eom_check")
    ws = load_workbook(path)["Summary"]
    header_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Month"
    )
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, 14)]
    assert headers[1] == "Regime"                       # inserted after Month
    assert headers[11:] == ["P&L %", "Equity EoM"]
    month_cells = [ws.cell(row=header_row + 1, column=c).value for c in range(1, 14)]
    assert month_cells[0] == "2026-06"
    assert month_cells[1] == (monthly[0].get("regime") or "-")   # regime label cell
    assert month_cells[12] == monthly[0]["equity_end"]


def test_weekly_breakdown_groups_by_feed_zone_month_week(tmp_path):
    # 2026-06-02 falls in week-of-month W1 (days 1-7), keyed on the FEED-zone date
    # (same zone the monthly/daily breakdowns use).
    sig = parse_one_signal(
        "1. BUY XAUUSD 100 - 100 SL 95 TP1 110 TP2 115 TP3 120 11:00 AM",
        source_date="2026-06-02", source_offset=3,
    )
    result = run_backtest([sig], _chart(tmp_path), _config())
    weekly = result["weekly"]
    assert [w["week"] for w in weekly] == ["2026-06 W1"]
    w = weekly[0]
    assert w["signals"] == 1 and w["wins"] == 1
    # A single-week run: the week's P&L and end equity match the month and the run.
    assert w["equity_end"] == result["final_equity"]
    assert round(w["pnl"], 6) == round(result["monthly"][0]["pnl"], 6)


def test_weekly_breakdown_sheet_written_with_headers(tmp_path):
    from openpyxl import load_workbook

    from trading.engine import write_backtest_outputs

    sig = parse_one_signal(
        "1. BUY XAUUSD 100 - 100 SL 95 TP1 110 TP2 115 TP3 120 11:00 AM",
        source_date="2026-06-02", source_offset=3,
    )
    result = run_backtest([sig], _chart(tmp_path), _config())
    path = write_backtest_outputs(result, tmp_path / "reports" / "weekly_check")
    wb = load_workbook(path)
    assert "Weekly Breakdown" in wb.sheetnames
    ws = wb["Weekly Breakdown"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 13)]
    assert headers[0] == "Month-Week"
    assert headers[-1] == "Equity EoW"
    assert ws.cell(row=2, column=1).value == "2026-06 W1"   # first month-week row
